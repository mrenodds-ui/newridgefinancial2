from __future__ import annotations

import csv
import json
import logging
import os
import shutil
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable


try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - optional dependency
    load_workbook = None

try:
    import xlrd
except ImportError:  # pragma: no cover - optional dependency
    xlrd = None


SOFTDENT_IMPORT_FILE_NAMES = frozenset(
    {
        "softdent_dashboard_data.json",
        "softdent_dashboard_data.csv",
        "softdent_claims_export.csv",
        "softdent_clinical_notes_data.json",
        "outstanding_claims_by_company.csv",
        "unsubmitted_claims.csv",
        "softdent_ar_aging.csv",
        "softdent_accounts_receivable.csv",
        "insurance_income.csv",
        "insurance_payment_distribution.csv",
        "insurance_check_distribution.csv",
        "treatment_plan_summary.csv",
        "payment_plans.csv",
    }
)

SOFTDENT_PASSTHROUGH_IMPORT_FILES = {
    name.casefold(): name
    for name in (
        "outstanding_claims_by_company.csv",
        "unsubmitted_claims.csv",
        "softdent_ar_aging.csv",
        "softdent_accounts_receivable.csv",
        "insurance_income.csv",
        "insurance_payment_distribution.csv",
        "insurance_check_distribution.csv",
        "treatment_plan_summary.csv",
        "payment_plans.csv",
    )
}

SUPPORTED_IMPORT_SUFFIXES = frozenset({".csv", ".json", ".txt", ".xlsx", ".xlsm", ".xls"})
SUPPORTED_TEXT_IMPORT_ENCODINGS = ("utf-8-sig", "utf-8", "cp1252")

logger = logging.getLogger(__name__)


@dataclass
class RuntimeImportSettings:
    softdent_source_dir: Path | None
    quickbooks_source_dir: Path | None
    softdent_import_dir: Path
    quickbooks_import_dir: Path
    softdent_auto_pull_enabled: bool
    quickbooks_auto_pull_enabled: bool
    financial_daily_refresh_enabled: bool
    ai_workspace_dir: Path


@dataclass
class PullSection:
    enabled: bool
    status: str
    summary: str
    source_dir: str = ""
    import_dir: str = ""
    scanned: int = 0
    copied: int = 0
    files: list[str] = field(default_factory=list)
    last_refresh_utc: str = ""
    last_error: str = ""

    def to_dict(self) -> dict[str, object]:
        return {
            "enabled": self.enabled,
            "status": self.status,
            "summary": self.summary,
            "source_dir": self.source_dir,
            "import_dir": self.import_dir,
            "scanned": self.scanned,
            "copied": self.copied,
            "files": list(self.files),
            "last_refresh_utc": self.last_refresh_utc,
            "last_error": self.last_error,
        }


def recompute_cache(app) -> dict[str, object]:
    settings = ensure_runtime_state(app)
    settings.softdent_import_dir.mkdir(parents=True, exist_ok=True)
    settings.quickbooks_import_dir.mkdir(parents=True, exist_ok=True)
    settings.ai_workspace_dir.mkdir(parents=True, exist_ok=True)

    evaluated_at = _utc_now_iso()
    softdent_section = _pull_softdent_sources(settings, evaluated_at=evaluated_at)
    quickbooks_section = _pull_quickbooks_sources(settings, evaluated_at=evaluated_at)
    practice_central_section = PullSection(
        enabled=False,
        status="idle",
        summary="Practice Central pull is not configured in this repo.",
        last_refresh_utc=evaluated_at,
    )

    payload = {
        "softdent": softdent_section.to_dict(),
        "quickbooks": quickbooks_section.to_dict(),
        "practice_central": practice_central_section.to_dict(),
    }
    app.state.report_pull_status = payload
    app.state.last_refresh_date = datetime.now(timezone.utc).date().isoformat()
    app.state.last_refresh_utc = evaluated_at
    _refresh_current_kpis(app)
    try:
        from app.hal.widget_builder import refresh_import_driven_widget_feed

        refresh_import_driven_widget_feed()
    except Exception as exc:
        logger.warning("Import-driven widget feed refresh failed after cache recompute: %s", exc)
    return payload


def ensure_runtime_state(app):
    settings = getattr(app.state, "settings", None)
    if settings is None:
        settings = get_runtime_settings()
        app.state.settings = settings
    if not hasattr(app.state, "report_pull_status"):
        app.state.report_pull_status = {}
    if not hasattr(app.state, "current_kpis"):
        app.state.current_kpis = []
    if not hasattr(app.state, "last_refresh_date"):
        app.state.last_refresh_date = ""
    if not hasattr(app.state, "last_refresh_utc"):
        app.state.last_refresh_utc = ""
    return settings


def get_pull_status_payload(app) -> dict[str, object]:
    settings = ensure_runtime_state(app)
    status_sections = getattr(app.state, "report_pull_status", None) or recompute_cache(app)
    return {
        "daily_refresh_enabled": settings.financial_daily_refresh_enabled,
        "last_refresh_date": getattr(app.state, "last_refresh_date", "") or "",
        "status": status_sections,
    }


def get_runtime_settings() -> RuntimeImportSettings:
    project_root = _project_root()
    return RuntimeImportSettings(
        softdent_source_dir=_coalesce_path(
            _read_env_path("SOFTDENT_SOURCE_DIR", project_root=project_root),
            _read_env_path("SOFTDENT_SENSEI_DATASYNC_ROOT", project_root=project_root),
        ),
        quickbooks_source_dir=_read_env_path("QUICKBOOKS_SOURCE_DIR", project_root=project_root),
        softdent_import_dir=_read_env_path("SOFTDENT_IMPORT_DIR", project_root=project_root, default="app/data/imports/softdent") or (project_root / "app" / "data" / "imports" / "softdent"),
        quickbooks_import_dir=_read_env_path("QUICKBOOKS_IMPORT_DIR", project_root=project_root, default="app/data/imports/quickbooks") or (project_root / "app" / "data" / "imports" / "quickbooks"),
        softdent_auto_pull_enabled=_read_env_flag("SOFTDENT_AUTO_PULL_ENABLED", default=False),
        quickbooks_auto_pull_enabled=_read_env_flag("QUICKBOOKS_AUTO_PULL_ENABLED", default=False),
        financial_daily_refresh_enabled=_read_env_flag("FINANCIAL_DAILY_REFRESH_ENABLED", default=True),
        ai_workspace_dir=project_root / "AI_Workspace",
    )


def import_uploaded_file(*, app, source: str, file_name: str, content: bytes) -> dict[str, object]:
    settings = ensure_runtime_state(app)
    _validate_import_file_name(file_name)
    rows = _read_rows_from_bytes(file_name=file_name, content=content)
    if source == "softdent":
        passthrough_name = _softdent_passthrough_target_name(file_name)
        if passthrough_name:
            written_paths = [_write_bytes(settings.softdent_import_dir / passthrough_name, content)]
        else:
            written_paths = _write_softdent_rows(rows, settings.softdent_import_dir)
    elif source == "quickbooks":
        written_paths = _write_quickbooks_rows(rows, settings.quickbooks_import_dir, file_name=file_name)
    else:
        raise ValueError(f"Unsupported import source: {source}")

    recompute_cache(app)
    return {
        "source": source,
        "files_written": [str(path) for path in written_paths],
        "row_count": len(rows),
    }


def _pull_softdent_sources(settings: RuntimeImportSettings, *, evaluated_at: str) -> PullSection:
    section = PullSection(
        enabled=settings.softdent_auto_pull_enabled,
        status="idle",
        summary="No SoftDent source files were processed.",
        source_dir=str(settings.softdent_source_dir or ""),
        import_dir=str(settings.softdent_import_dir),
        last_refresh_utc=evaluated_at,
    )
    candidate_files = _list_source_files(settings.softdent_source_dir) if settings.softdent_auto_pull_enabled else []
    section.scanned = len(candidate_files)

    copied: list[Path] = []
    seen_targets: set[str] = set()
    try:
        for source_file in candidate_files:
            copied.extend(_copy_softdent_source_file(source_file, settings.softdent_import_dir, seen_targets))
    except Exception as exc:
        section.status = "error"
        section.summary = f"SoftDent import failed: {exc}"
        section.last_error = str(exc)
        return section

    section.copied = len(copied)
    current_files = _current_softdent_import_files(settings.softdent_import_dir)
    section.files = sorted({path.name for path in copied} | set(current_files))
    if copied:
        section.status = "ready"
        section.summary = f"Processed {len(copied)} SoftDent import file(s)."
    elif current_files:
        section.status = "ready"
        section.summary = f"Using {len(current_files)} SoftDent import file(s) already present in the import directory."
    elif settings.softdent_auto_pull_enabled and settings.softdent_source_dir and settings.softdent_source_dir.exists():
        section.status = "warning"
        section.summary = "SoftDent auto-pull is enabled, but no supported source files were found."
    elif settings.softdent_auto_pull_enabled:
        section.status = "warning"
        section.summary = "SoftDent auto-pull is enabled, but the configured source directory is unavailable."
    else:
        section.status = "idle"
        section.summary = "SoftDent auto-pull is disabled."
    return section


def _pull_quickbooks_sources(settings: RuntimeImportSettings, *, evaluated_at: str) -> PullSection:
    section = PullSection(
        enabled=settings.quickbooks_auto_pull_enabled,
        status="idle",
        summary="No QuickBooks source files were processed.",
        source_dir=str(settings.quickbooks_source_dir or ""),
        import_dir=str(settings.quickbooks_import_dir),
        last_refresh_utc=evaluated_at,
    )
    candidate_files = _list_source_files(settings.quickbooks_source_dir) if settings.quickbooks_auto_pull_enabled else []
    section.scanned = len(candidate_files)

    copied: list[Path] = []
    seen_targets: set[str] = set()
    try:
        for source_file in candidate_files:
            copied.extend(_copy_quickbooks_file(source_file, settings.quickbooks_import_dir, seen_targets))
    except Exception as exc:
        section.status = "error"
        section.summary = f"QuickBooks import failed: {exc}"
        section.last_error = str(exc)
        return section

    section.copied = len(copied)
    current_files = _current_quickbooks_import_files(settings.quickbooks_import_dir)
    section.files = sorted({path.name for path in copied} | set(current_files))
    if copied:
        section.status = "ready"
        section.summary = f"Processed {len(copied)} QuickBooks import file(s)."
    elif current_files:
        section.status = "ready"
        section.summary = f"Using {len(current_files)} QuickBooks import file(s) already present in the import directory."
    elif settings.quickbooks_auto_pull_enabled and settings.quickbooks_source_dir and settings.quickbooks_source_dir.exists():
        section.status = "warning"
        section.summary = "QuickBooks auto-pull is enabled, but no supported source files were found."
    elif settings.quickbooks_auto_pull_enabled:
        section.status = "warning"
        section.summary = "QuickBooks auto-pull is enabled, but the configured source directory is unavailable."
    else:
        section.status = "idle"
        section.summary = "QuickBooks auto-pull is disabled."
    return section


def _copy_softdent_source_file(source_path: Path, import_dir: Path, seen_targets: set[str]) -> list[Path]:
    passthrough_name = _softdent_passthrough_target_name(source_path.name)
    if passthrough_name:
        import_dir.mkdir(parents=True, exist_ok=True)
        target_path = import_dir / passthrough_name
        key = str(target_path).casefold()
        if key in seen_targets:
            return []
        seen_targets.add(key)
        shutil.copy2(source_path, target_path)
        return [target_path]

    rows = _read_rows_from_path(source_path)
    written = _write_softdent_rows(rows, import_dir)
    output: list[Path] = []
    for target_path in written:
        key = str(target_path).casefold()
        if key in seen_targets:
            continue
        seen_targets.add(key)
        output.append(target_path)
    return output


def _copy_quickbooks_file(source_path: Path, import_dir: Path, seen_targets: set[str]) -> list[Path]:
    import_dir.mkdir(parents=True, exist_ok=True)
    target_name = _quickbooks_target_name(source_path.name)
    target_path = import_dir / target_name
    key = str(target_path).casefold()
    if key in seen_targets:
        return []
    seen_targets.add(key)
    shutil.copy2(source_path, target_path)
    return [target_path]


def _current_softdent_import_files(import_dir: Path) -> list[str]:
    return sorted(path.name for path in import_dir.iterdir() if path.is_file() and path.name in SOFTDENT_IMPORT_FILE_NAMES) if import_dir.exists() else []


def _current_quickbooks_import_files(import_dir: Path) -> list[str]:
    return sorted(path.name for path in import_dir.glob("quickbooks_*.*")) if import_dir.exists() else []


def _write_softdent_rows(rows: list[dict[str, object]], import_dir: Path) -> list[Path]:
    import_dir.mkdir(parents=True, exist_ok=True)
    if _looks_like_dashboard_rows(rows):
        dashboard_rows = _merge_dashboard_rows(_load_existing_softdent_dashboard_rows(), _normalize_dashboard_rows(rows))
        return [
            _write_json(import_dir / "softdent_dashboard_data.json", dashboard_rows),
            _write_csv(import_dir / "softdent_dashboard_data.csv", dashboard_rows),
        ]

    if _looks_like_claim_rows(rows):
        claims_rows = _normalize_claim_rows(rows)
        return [_write_csv(import_dir / "softdent_claims_export.csv", claims_rows)]

    if _looks_like_note_rows(rows):
        note_rows = _normalize_note_rows(rows)
        return [_write_json(import_dir / "softdent_clinical_notes_data.json", {"notes": note_rows})]

    raise ValueError("Uploaded SoftDent file did not match a supported dashboard, claims, or clinical-note shape")


def _write_quickbooks_rows(rows: list[dict[str, object]], import_dir: Path, *, file_name: str) -> list[Path]:
    import_dir.mkdir(parents=True, exist_ok=True)
    target_path = import_dir / _quickbooks_target_name(file_name)
    if not rows:
        raise ValueError("Uploaded QuickBooks file did not contain any rows")
    return [_write_csv(target_path, rows)]


def _quickbooks_target_name(file_name: str) -> str:
    stem = Path(file_name).stem.lower()
    suffix = Path(file_name).suffix.lower() or ".csv"
    normalized = "_".join(part for part in stem.replace("-", " ").replace("&", " ").split() if part)
    normalized = normalized or "export"
    if not normalized.startswith("quickbooks_"):
        normalized = f"quickbooks_{normalized}"
    return f"{normalized}{suffix if suffix in SUPPORTED_IMPORT_SUFFIXES else '.csv'}"


def _normalize_dashboard_rows(rows: list[dict[str, object]]) -> list[dict[str, object]]:
    if _looks_like_metric_amount_rows(rows):
        grouped: dict[tuple[str, str], dict[str, object]] = {}
        for row in rows:
            period = _first_value(row, "period", "year_month", "month", "report_period")
            metric = _first_value(row, "metric", "measure", "name").lower()
            provider = _first_value(row, "provider", "provider_name", "doctor") or "Unknown"
            amount = _to_float(_first_value(row, "amount", "value", "total"))
            key = (provider or "Unknown", period)
            aggregate = grouped.setdefault(
                key,
                {"provider": provider or "Unknown", "period": period, "production": 0.0, "collections": 0.0, "insurance": 0.0, "patient": 0.0},
            )
            if "production" in metric:
                aggregate["production"] = round(float(aggregate["production"]) + amount, 2)
            elif "collection" in metric:
                aggregate["collections"] = round(float(aggregate["collections"]) + amount, 2)
            elif "insurance" in metric:
                aggregate["insurance"] = round(float(aggregate["insurance"]) + amount, 2)
            elif "patient" in metric:
                aggregate["patient"] = round(float(aggregate["patient"]) + amount, 2)
        return list(grouped.values())

    normalized: list[dict[str, object]] = []
    for row in rows:
        provider = _first_value(row, "provider", "provider_name", "doctor") or "Unknown"
        period = _first_value(row, "period", "year_month", "month", "report_period")
        production = _to_float(_first_value(row, "production", "gross_production", "net_production"))
        collections = _to_float(_first_value(row, "collections", "collection_total", "deposit_total"))
        insurance = _to_float(_first_value(row, "insurance", "insurance_amount"))
        patient = _to_float(_first_value(row, "patient", "patient_amount"))
        if provider or period or production or collections or insurance or patient:
            normalized.append(
                {
                    "provider": provider,
                    "period": period,
                    "production": production,
                    "collections": collections,
                    "insurance": insurance,
                    "patient": patient,
                }
            )
    return normalized


def _normalize_claim_rows(rows: list[dict[str, object]]) -> list[dict[str, object]]:
    normalized: list[dict[str, object]] = []
    for row in rows:
        claim_id = _first_value(row, "ClaimId", "claim_number", "claimid", "claim", "refnumber")
        claim_status = _first_value(row, "ClaimStatus", "status")
        payer = _first_value(row, "Payer", "carrier", "insurance", "plan")
        service_date = _first_value(row, "ServiceDate", "dos", "date", "dateofservice")
        amount = _to_float(_first_value(row, "ClaimAmount", "amount", "balance", "current_balance", "total_ar"))
        if not (claim_id or claim_status or payer or service_date or amount):
            continue
        normalized.append(
            {
                "PatientName": _first_value(row, "PatientName", "patient_name", "patient", "name"),
                "MRN": _first_value(row, "MRN", "patient_id", "chartnumber"),
                "ClaimId": claim_id,
                "ClaimStatus": claim_status,
                "Payer": payer,
                "Procedure": _first_value(row, "Procedure", "procdesc", "description", "treatment"),
                "ServiceDate": service_date,
                "DenialReason": _first_value(row, "DenialReason", "reason", "note", "remark"),
                "ClaimAmount": amount,
            }
        )
    return normalized


def _normalize_note_rows(rows: list[dict[str, object]]) -> list[dict[str, object]]:
    normalized: list[dict[str, object]] = []
    for row in rows:
        note_text = _first_value(row, "ClinicalNote", "note", "narrative", "chartnote", "assessment")
        note_date = _first_value(row, "NoteDate", "entrydate", "date", "servicedate")
        if not (note_text or note_date):
            continue
        normalized.append(
            {
                "PatientName": _first_value(row, "PatientName", "patient_name", "patient", "name"),
                "MRN": _first_value(row, "MRN", "patient_id", "chartnumber"),
                "NoteDate": note_date,
                "Provider": _first_value(row, "Provider", "doctor", "clinician"),
                "Procedure": _first_value(row, "Procedure", "procdesc", "description", "treatment"),
                "ClinicalNote": note_text,
            }
        )
    return normalized


def _looks_like_metric_amount_rows(rows: list[dict[str, object]]) -> bool:
    if not rows:
        return False
    first_row = rows[0]
    has_metric = bool(_first_value(first_row, "metric", "measure", "name"))
    has_amount = bool(_first_value(first_row, "amount", "value", "total"))
    has_period = bool(_first_value(first_row, "period", "year_month", "month", "report_period"))
    return has_metric and has_amount and has_period


def _looks_like_dashboard_rows(rows: list[dict[str, object]]) -> bool:
    if _looks_like_metric_amount_rows(rows):
        return True
    if not rows:
        return False
    first_row = rows[0]
    has_provider = bool(_first_value(first_row, "provider", "provider_name", "doctor"))
    has_production = bool(_first_value(first_row, "production", "gross_production", "net_production"))
    has_collections = bool(_first_value(first_row, "collections", "collection_total", "deposit_total"))
    return has_provider and (has_production or has_collections)


def _looks_like_claim_rows(rows: list[dict[str, object]]) -> bool:
    if not rows:
        return False
    first_row = rows[0]
    explicit_claim = bool(_first_value(first_row, "ClaimId", "claim_number", "claimid", "claim", "refnumber"))
    claim_status = bool(_first_value(first_row, "ClaimStatus", "status"))
    payer = bool(_first_value(first_row, "Payer", "carrier", "insurance", "plan"))
    aging_columns = any(
        _first_value(first_row, alias)
        for alias in ("current_balance", "balance_30", "balance_60", "balance_90", "total_ar")
    )
    return (explicit_claim and (claim_status or payer)) or bool(aging_columns)


def _looks_like_note_rows(rows: list[dict[str, object]]) -> bool:
    if not rows:
        return False
    first_row = rows[0]
    note_text = bool(_first_value(first_row, "ClinicalNote", "note", "narrative", "chartnote", "assessment"))
    note_date = bool(_first_value(first_row, "NoteDate", "entrydate", "date", "servicedate"))
    return note_text and note_date


def _load_existing_softdent_dashboard_rows() -> list[dict[str, object]]:
    try:
        from app.services import load_softdent_dashboard_rows
    except Exception:
        return []
    return _normalize_dashboard_rows(load_softdent_dashboard_rows())


def _merge_dashboard_rows(existing_rows: list[dict[str, object]], incoming_rows: list[dict[str, object]]) -> list[dict[str, object]]:
    merged: dict[tuple[str, str], dict[str, object]] = {}
    for row in existing_rows:
        provider = str(row.get("provider") or "Unknown")
        period = str(row.get("period") or "")
        key = (provider, period)
        merged[key] = {
            "provider": provider,
            "period": period,
            "production": round(_to_float(row.get("production")), 2),
            "collections": round(_to_float(row.get("collections")), 2),
            "insurance": round(_to_float(row.get("insurance")), 2),
            "patient": round(_to_float(row.get("patient")), 2),
        }

    incoming_by_key: dict[tuple[str, str], dict[str, object]] = {}
    for row in incoming_rows:
        provider = str(row.get("provider") or "Unknown")
        period = str(row.get("period") or "")
        key = (provider, period)
        aggregate = incoming_by_key.setdefault(
            key,
            {
                "provider": provider,
                "period": period,
                "production": 0.0,
                "collections": 0.0,
                "insurance": 0.0,
                "patient": 0.0,
            },
        )
        for field in ("production", "collections", "insurance", "patient"):
            aggregate[field] = round(float(aggregate[field]) + _to_float(row.get(field)), 2)

    merged.update(incoming_by_key)
    return list(merged.values())


def _list_source_files(root: Path | None) -> list[Path]:
    if root is None or not root.exists():
        return []
    files = [path for path in root.rglob("*") if path.is_file() and path.suffix.lower() in SUPPORTED_IMPORT_SUFFIXES]
    return sorted(files, key=lambda path: path.stat().st_mtime, reverse=True)


def _read_rows_from_path(path: Path) -> list[dict[str, object]]:
    suffix = path.suffix.lower()
    if suffix in {".csv", ".txt"}:
        with path.open("r", encoding=_detect_text_encoding(path.read_bytes(), file_name=path.name), newline="") as handle:
            return _materialize_rows(csv.DictReader(handle))
    if suffix == ".json":
        return _extract_json_rows(json.loads(_decode_text_import(path.read_bytes(), file_name=path.name, allow_windows_1252=False)))
    if suffix in {".xlsx", ".xlsm"}:
        if load_workbook is None:
            raise ValueError(f"Uploaded file '{path.name}' requires openpyxl support for .xlsx/.xlsm imports")
        workbook = load_workbook(filename=path, read_only=True, data_only=True)
        try:
            sheet = workbook[workbook.sheetnames[0]] if workbook.sheetnames else None
            return _rows_from_iterable(sheet.iter_rows(values_only=True)) if sheet is not None else []
        finally:
            workbook.close()
    if suffix == ".xls":
        if xlrd is None:
            raise ValueError(f"Uploaded file '{path.name}' requires xlrd support for .xls imports")
        workbook = xlrd.open_workbook(path)
        sheet = workbook.sheet_by_index(0) if workbook.nsheets else None
        return _rows_from_iterable(sheet.row_values(index) for index in range(sheet.nrows)) if sheet is not None else []
    return []


def _read_rows_from_bytes(*, file_name: str, content: bytes) -> list[dict[str, object]]:
    suffix = Path(file_name).suffix.lower()
    if suffix in {".csv", ".txt", ""}:
        text = _decode_text_import(content, file_name=file_name)
        return _materialize_rows(csv.DictReader(text.splitlines()))
    if suffix == ".json":
        return _extract_json_rows(json.loads(_decode_text_import(content, file_name=file_name, allow_windows_1252=False)))
    temp_path = _project_root() / "AI_Workspace" / f".tmp-import-{Path(file_name).name}"
    temp_path.parent.mkdir(parents=True, exist_ok=True)
    temp_path.write_bytes(content)
    try:
        return _read_rows_from_path(temp_path)
    finally:
        temp_path.unlink(missing_ok=True)


def _materialize_rows(rows: Iterable[dict[str, object]]) -> list[dict[str, object]]:
    return [{str(key): value for key, value in row.items()} for row in rows if any(str(value or "").strip() for value in row.values())]


def _validate_import_file_name(file_name: str) -> None:
    suffix = Path(file_name).suffix.lower()
    if suffix not in SUPPORTED_IMPORT_SUFFIXES:
        supported = ", ".join(sorted(SUPPORTED_IMPORT_SUFFIXES))
        raise ValueError(f"Unsupported import file type '{suffix or '<none>'}'. Supported file types: {supported}")


def _detect_text_encoding(content: bytes, *, file_name: str, allow_windows_1252: bool = True) -> str:
    if content.startswith((b"\xff\xfe", b"\xfe\xff", b"\xff\xfe\x00\x00", b"\x00\x00\xfe\xff")) or b"\x00" in content:
        if allow_windows_1252:
            raise ValueError(f"Uploaded file '{file_name}' must use UTF-8 or Windows-1252 text encoding")
        raise ValueError(f"Uploaded JSON file '{file_name}' must use UTF-8 text encoding")
    encodings = SUPPORTED_TEXT_IMPORT_ENCODINGS if allow_windows_1252 else ("utf-8-sig", "utf-8")
    for encoding in encodings:
        try:
            content.decode(encoding)
            return encoding
        except UnicodeDecodeError:
            continue
    if allow_windows_1252:
        raise ValueError(f"Uploaded file '{file_name}' must use UTF-8 or Windows-1252 text encoding")
    raise ValueError(f"Uploaded JSON file '{file_name}' must use UTF-8 text encoding")


def _decode_text_import(content: bytes, *, file_name: str, allow_windows_1252: bool = True) -> str:
    encoding = _detect_text_encoding(content, file_name=file_name, allow_windows_1252=allow_windows_1252)
    decoded = content.decode(encoding)
    if "\x00" in decoded:
        if allow_windows_1252:
            raise ValueError(f"Uploaded file '{file_name}' must use UTF-8 or Windows-1252 text encoding")
        raise ValueError(f"Uploaded JSON file '{file_name}' must use UTF-8 text encoding")
    return decoded


def _rows_from_iterable(rows: Iterable[Iterable[object]]) -> list[dict[str, object]]:
    materialized = [list(row) for row in rows if any(cell not in (None, "") for cell in row)]
    if not materialized:
        return []
    headers = [str(cell).strip() if cell is not None else "" for cell in materialized[0]]
    if not any(headers):
        return []
    normalized_headers = [header or f"column_{index + 1}" for index, header in enumerate(headers)]
    records: list[dict[str, object]] = []
    for row in materialized[1:]:
        padded = row + [""] * max(0, len(normalized_headers) - len(row))
        records.append({normalized_headers[index]: padded[index] for index in range(len(normalized_headers))})
    return records


def _extract_json_rows(payload: object) -> list[dict[str, object]]:
    if isinstance(payload, list):
        return [row for row in payload if isinstance(row, dict)]
    if isinstance(payload, dict):
        for key in ("rows", "items", "data", "claims", "notes"):
            value = payload.get(key)
            if isinstance(value, list):
                return [row for row in value if isinstance(row, dict)]
        for value in payload.values():
            if isinstance(value, list) and all(isinstance(row, dict) for row in value):
                return list(value)
    return []


def _write_csv(path: Path, rows: list[dict[str, object]]) -> Path:
    if not rows:
        raise ValueError(f"Cannot write empty CSV to {path}")
    path.parent.mkdir(parents=True, exist_ok=True)
    headers = list(rows[0].keys())
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)
    return path


def _write_json(path: Path, payload: object) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    return path


def _write_bytes(path: Path, content: bytes) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(content)
    return path


def _refresh_current_kpis(app) -> None:
    try:
        from app.services import load_softdent_dashboard_rows
    except Exception:
        app.state.current_kpis = []
        return

    rows = load_softdent_dashboard_rows()
    grouped: dict[str, dict[str, object]] = {}
    for row in rows:
        period = str(row.get("period") or row.get("Period") or row.get("year_month") or row.get("month") or "").strip()
        if not period:
            continue
        aggregate = grouped.setdefault(period, {"period": period, "production": 0.0, "collections": 0.0})
        aggregate["production"] = round(float(aggregate["production"]) + _to_float(row.get("production") or row.get("Production") or row.get("gross_production")), 2)
        aggregate["collections"] = round(float(aggregate["collections"]) + _to_float(row.get("collections") or row.get("Collections")), 2)
    app.state.current_kpis = sorted(grouped.values(), key=lambda item: str(item["period"]))


def _first_value(row: dict[str, object], *aliases: str) -> str:
    normalized = {_normalize_alias(alias) for alias in aliases}
    for key, value in row.items():
        if value in (None, ""):
            continue
        if _normalize_alias(str(key)) in normalized:
            return str(value).strip()
    return ""


def _softdent_passthrough_target_name(file_name: str) -> str | None:
    return SOFTDENT_PASSTHROUGH_IMPORT_FILES.get(Path(file_name).name.casefold())


def _normalize_alias(value: str) -> str:
    return "".join(character for character in value.lower() if character.isalnum())


def _to_float(value: object) -> float:
    try:
        return float(str(value or "").replace(",", "").replace("$", "").strip())
    except ValueError:
        return 0.0


def _read_env_flag(name: str, *, default: bool) -> bool:
    value = os.getenv(name)
    if value is None:
        return default
    return value.strip().lower() not in {"", "0", "false", "no", "off"}


def _read_env_path(name: str, *, project_root: Path, default: str | None = None) -> Path | None:
    configured = os.getenv(name, default or "").strip()
    if not configured:
        return None
    candidate = Path(configured)
    if not candidate.is_absolute():
        candidate = project_root / candidate
    return candidate.expanduser().resolve()


def _coalesce_path(*values: Path | None) -> Path | None:
    for value in values:
        if value is not None:
            return value
    return None


def _utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _project_root() -> Path:
    return Path(__file__).resolve().parents[1]