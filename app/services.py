import os
from calendar import monthrange
from contextlib import contextmanager
import csv
import json
import gc
from datetime import datetime, timezone
from pathlib import Path
import queue
import subprocess
import sys
import threading
import sqlite3
from typing import Iterable
import winreg
import xml.etree.ElementTree as ET

from fastapi import Request
from pydantic import BaseModel, ConfigDict, Field
from .auth import authenticate, get_service_user
from app.hal.safety import append_ai_activity_log, ensure_within_ai_workspace, get_ai_workspace_path, resolve_within_hal_allowed_base
from app.hal.storage import get_hal_storage_path

try:
    import pyodbc
except ImportError:  # pragma: no cover - exercised in environments without QODBC support
    pyodbc = None

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - exercised in environments without workbook support
    load_workbook = None

try:
    import xlrd
except ImportError:  # pragma: no cover - exercised in environments without workbook support
    xlrd = None

try:
    import win32com.client  # type: ignore[import-untyped]
    import pythoncom  # type: ignore[import-untyped]
except ImportError:  # pragma: no cover - exercised in environments without pywin32
    win32com = None
    pythoncom = None


QUICKBOOKS_SDK_APP_NAME = os.getenv("QUICKBOOKS_SDK_APP_NAME", "New Ridge HAL")
QUICKBOOKS_SDK_TIMEOUT_SECONDS = float(os.getenv("QUICKBOOKS_SDK_TIMEOUT_SECONDS", "8"))
QUICKBOOKS_QBXML_VERSION = os.getenv("QUICKBOOKS_QBXML_VERSION", "13.0")
MAX_QUICKBOOKS_DIAGNOSTIC_SQL_LENGTH = 10_000
QUICKBOOKS_EXPORT_ENV_BY_TOPIC: dict[str, str] = {
    "revenue": "QUICKBOOKS_REVENUE_EXPORT_PATH",
    "expenses": "QUICKBOOKS_EXPENSES_EXPORT_PATH",
    "ar": "QUICKBOOKS_AR_EXPORT_PATH",
}
QUICKBOOKS_EXPORT_DEFAULT_NAMES_BY_TOPIC: dict[str, tuple[str, ...]] = {
    "revenue": (
        "quickbooks_revenue.csv",
        "quickbooks_revenue.json",
        "quickbooks_profit_and_loss.csv",
        "quickbooks_profit_loss.csv",
    ),
    "expenses": (
        "quickbooks_expenses.csv",
        "quickbooks_expense_detail.csv",
        "quickbooks_expense_categories.csv",
    ),
    "ar": (
        "quickbooks_ar.csv",
        "quickbooks_accounts_receivable.csv",
        "quickbooks_aging.csv",
    ),
}
SOFTDENT_OUTSTANDING_CLAIMS_EXPORT_ENV = "SOFTDENT_OUTSTANDING_CLAIMS_EXPORT_PATH"
SOFTDENT_UNSUBMITTED_CLAIMS_EXPORT_ENV = "SOFTDENT_UNSUBMITTED_CLAIMS_EXPORT_PATH"
SOFTDENT_INSURANCE_INCOME_EXPORT_ENV = "SOFTDENT_INSURANCE_INCOME_EXPORT_PATH"
SOFTDENT_INSURANCE_PAYMENT_DISTRIBUTION_EXPORT_ENV = "SOFTDENT_INSURANCE_PAYMENT_DISTRIBUTION_EXPORT_PATH"
SOFTDENT_INSURANCE_CHECK_DISTRIBUTION_EXPORT_ENV = "SOFTDENT_INSURANCE_CHECK_DISTRIBUTION_EXPORT_PATH"
SOFTDENT_TREATMENT_PLAN_EXPORT_ENV = "SOFTDENT_TREATMENT_PLAN_EXPORT_PATH"
SOFTDENT_PAYMENT_PLAN_EXPORT_ENV = "SOFTDENT_PAYMENT_PLAN_EXPORT_PATH"
HAL_ALLOWED_STAGED_IMPORT_FILES = frozenset(
    file_name
    for default_names in QUICKBOOKS_EXPORT_DEFAULT_NAMES_BY_TOPIC.values()
    for file_name in default_names
)


class ProviderProductionRow(BaseModel):
    provider_id: str = Field(..., description="Unique provider identifier or alias")
    provider_name: str = Field(..., description="Provider display name")
    production_amount: float
    collection_amount: float


class SoftDentAggregateSnapshot(BaseModel):
    source_file: str = Field(..., description="Approved local snapshot file name")
    period_start: str = Field(..., description="Start date in YYYY-MM-DD format when derivable")
    period_end: str = Field(..., description="End date in YYYY-MM-DD format when derivable")
    provider_count: int
    provider_rows: list[ProviderProductionRow] = Field(default_factory=list)
    totals: dict[str, float] = Field(
        default_factory=dict,
        description="Strict aggregate totals keyed by production, collections, insurance, and patient",
    )
    data_complete: bool


class PatientClaimRecord(BaseModel):
    model_config = ConfigDict(extra="allow")

    PatientName: str = ""
    MRN: str = ""
    ClaimId: str = ""
    ClaimStatus: str = ""
    Payer: str = ""
    Procedure: str = ""
    ServiceDate: str = ""
    DenialReason: str = ""
    ClaimAmount: float = 0.0


class ClinicalNoteEntry(BaseModel):
    model_config = ConfigDict(extra="allow")

    PatientName: str = ""
    MRN: str = ""
    NoteDate: str = ""
    Provider: str = ""
    Procedure: str = ""
    ClinicalNote: str = ""


def get_quickbooks_company_file() -> str:
    configured = os.getenv("QUICKBOOKS_COMPANY_FILE", "").strip()
    if configured:
        return configured

    try:
        with winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE, r"SOFTWARE\Intuit\QuickBooksCommon\QBFinder") as key:
            value, _ = winreg.QueryValueEx(key, "0")
            if isinstance(value, str) and "|" in value:
                return value.split("|", 1)[0].strip()
    except OSError:
        pass
    return ""


def get_quickbooks_sdk_status() -> dict[str, object]:
    company_file = get_quickbooks_company_file()
    return {
        "com_available": win32com is not None,
        "company_file": company_file,
        "company_file_exists": bool(company_file) and Path(company_file).exists(),
        "app_name": QUICKBOOKS_SDK_APP_NAME,
        "timeout_seconds": QUICKBOOKS_SDK_TIMEOUT_SECONDS,
    }


def fetch_quickbooks_sdk_summary(topic: str, period_dict: dict[str, str] | None = None) -> list[dict]:
    if topic not in {"revenue", "expenses", "ar"}:
        raise ValueError(f"Unsupported QuickBooks SDK topic: {topic}")

    command = [sys.executable, "-m", "app.quickbooks_sdk_runner", topic]
    if period_dict is not None:
        start_date, end_date = _validate_quickbooks_sdk_period(period_dict)
        command.extend([start_date, end_date])
    completed = subprocess.run(
        command,
        cwd=str(_project_root()),
        capture_output=True,
        text=True,
        timeout=max(int(QUICKBOOKS_SDK_TIMEOUT_SECONDS), 1) + 10,
        check=False,
    )
    if completed.returncode != 0:
        stderr = completed.stderr.strip()
        stdout = completed.stdout.strip()
        detail = stderr or stdout or f"subprocess exit code {completed.returncode}"
        raise RuntimeError(f"QuickBooks SDK subprocess failed: {detail}")

    payload = completed.stdout.strip()
    if not payload:
        return []
    try:
        parsed = json.loads(payload)
    except json.JSONDecodeError as exc:
        raise RuntimeError("QuickBooks SDK subprocess returned invalid JSON output") from exc
    if not isinstance(parsed, list):
        raise RuntimeError("QuickBooks SDK subprocess returned invalid JSON output")
    return parsed


def fetch_quickbooks_sdk_summary_direct(topic: str, period_dict: dict[str, str] | None = None) -> list[dict]:
    if win32com is None:
        raise RuntimeError("pywin32 is required for QuickBooks Desktop SDK access but is not installed in this environment")

    company_file = get_quickbooks_company_file()
    if not company_file:
        raise RuntimeError("QuickBooks company file could not be discovered for SDK access")

    request_xml = _build_quickbooks_sdk_request(topic, period_dict=period_dict)
    response_xml = _run_quickbooks_sdk_request(request_xml=request_xml, company_file=company_file)
    return _parse_quickbooks_sdk_summary(topic=topic, response_xml=response_xml)


def run_quickbooks_sdk_request_with_timeout(*, request_xml: str, company_file: str) -> str:
    result_queue: queue.Queue[tuple[str, object]] = queue.Queue(maxsize=1)

    def worker() -> None:
        initialized = False
        try:
            if pythoncom is not None:
                pythoncom.CoInitialize()
                initialized = True
            result_queue.put(("ok", _run_quickbooks_sdk_request(request_xml=request_xml, company_file=company_file)))
        except Exception as exc:  # pragma: no cover - exercised in integration environments
            result_queue.put(("error", exc))
        finally:
            if initialized and pythoncom is not None:
                pythoncom.CoUninitialize()

    thread = threading.Thread(target=worker, name="quickbooks-sdk-request", daemon=True)
    thread.start()
    thread.join(QUICKBOOKS_SDK_TIMEOUT_SECONDS)
    if thread.is_alive():
        raise RuntimeError("QuickBooks SDK request timed out or is blocked by the QuickBooks UI")

    status, payload = result_queue.get_nowait()
    if status == "error":
        raise payload  # type: ignore[misc]
    return str(payload)


def _build_quickbooks_sdk_request(topic: str, period_dict: dict[str, str] | None = None) -> str:
    report_date_tags = _build_quickbooks_sdk_report_date_tags(period_dict)
    if topic == "revenue":
        body = (
            "<GeneralSummaryReportQueryRq requestID=\"1\">"
            "<GeneralSummaryReportType>ProfitAndLossStandard</GeneralSummaryReportType>"
            f"{report_date_tags}"
            "</GeneralSummaryReportQueryRq>"
        )
    elif topic == "expenses":
        body = (
            "<GeneralSummaryReportQueryRq requestID=\"1\">"
            "<GeneralSummaryReportType>ProfitAndLossStandard</GeneralSummaryReportType>"
            f"{report_date_tags}"
            "</GeneralSummaryReportQueryRq>"
        )
    elif topic == "ar":
        if period_dict is not None:
            _validate_quickbooks_sdk_period(period_dict)
        body = (
            "<GeneralSummaryReportQueryRq requestID=\"1\">"
            "<GeneralSummaryReportType>ARAgingSummary</GeneralSummaryReportType>"
            f"{report_date_tags}"
            "</GeneralSummaryReportQueryRq>"
        )
    else:
        raise ValueError(f"Unsupported QuickBooks SDK topic: {topic}")

    return (
        f'<?xml version="1.0"?><?qbxml version="{QUICKBOOKS_QBXML_VERSION}"?>'
        f'<QBXML><QBXMLMsgsRq onError="stopOnError">{body}</QBXMLMsgsRq></QBXML>'
    )


def _run_quickbooks_sdk_request(*, request_xml: str, company_file: str) -> str:
    initialized = False
    if pythoncom is not None:
        pythoncom.CoInitialize()
        initialized = True
    request_processor = win32com.client.Dispatch("QBXMLRP2.RequestProcessor")
    ticket = None
    try:
        request_processor.OpenConnection2("", QUICKBOOKS_SDK_APP_NAME, 1)
        ticket = request_processor.BeginSession(company_file, 2)
        return str(request_processor.ProcessRequest(ticket, request_xml))
    finally:
        try:
            if ticket:
                request_processor.EndSession(ticket)
        except Exception:
            pass
        try:
            request_processor.CloseConnection()
        except Exception:
            pass
        request_processor = None
        gc.collect()
        if initialized and pythoncom is not None:
            pythoncom.CoUninitialize()


def _parse_quickbooks_sdk_summary(*, topic: str, response_xml: str) -> list[dict]:
    root = ET.fromstring(response_xml)

    if topic == "revenue":
        report = root.find(".//ReportRet")
        total_income = _find_report_amount(root, "Total Income")
        if report is None or total_income == "":
            return []
        return [
            {
                "ReportTitle": _xml_text(report, "ReportTitle"),
                "ReportPeriod": _xml_text(report, "ReportSubtitle"),
                "ReportBasis": _xml_text(report, "ReportBasis"),
                "TotalIncome": total_income,
            }
        ]

    if topic == "expenses":
        report = root.find(".//ReportRet")
        total_expense = _find_report_amount(root, "Total Expense")
        if report is None or total_expense == "":
            return []
        return [
            {
                "ReportTitle": _xml_text(report, "ReportTitle"),
                "ReportPeriod": _xml_text(report, "ReportSubtitle"),
                "ReportBasis": _xml_text(report, "ReportBasis"),
                "TotalExpense": total_expense,
            }
        ]

    if topic == "ar":
        report = root.find(".//ReportRet")
        if report is None:
            return []
        report_subtitle = _xml_text(report, "ReportSubtitle")
        rows: list[dict] = []
        for data_row in report.findall(".//DataRow"):
            customer_ref = _coldata_value(data_row, "1")
            outstanding_ar = _coldata_value(data_row, "6")
            if not customer_ref or not outstanding_ar:
                continue
            rows.append(
                {
                    "CustomerRef": customer_ref,
                    "OutstandingAR": outstanding_ar,
                    "ReportDate": report_subtitle,
                    "RefNumber": "",
                }
            )
        return rows

    return []


def _xml_text(node: ET.Element, tag_name: str) -> str:
    if tag_name == "FullName":
        child = node.find(f".//{tag_name}")
    else:
        child = node.find(tag_name)
    return child.text.strip() if child is not None and child.text else ""


def _find_report_amount(root: ET.Element, label: str) -> str:
    for subtotal in root.findall(".//SubtotalRow"):
        label_node = subtotal.find("./ColData[@colID='1']")
        amount_node = subtotal.find("./ColData[@colID='2']")
        if label_node is None or amount_node is None:
            continue
        if (label_node.attrib.get("value") or "").strip() == label:
            return (amount_node.attrib.get("value") or "").strip()
    return ""


def _coldata_value(node: ET.Element, col_id: str) -> str:
    child = node.find(f"./ColData[@colID='{col_id}']")
    return (child.attrib.get("value") or "").strip() if child is not None else ""


def _build_quickbooks_sdk_report_date_tags(period_dict: dict[str, str] | None) -> str:
    if period_dict is None:
        return ""

    start_date, end_date = _validate_quickbooks_sdk_period(period_dict)
    return (
        "<ReportPeriod>"
        f"<FromReportDate>{start_date}</FromReportDate>"
        f"<ToReportDate>{end_date}</ToReportDate>"
        "</ReportPeriod>"
    )


def _validate_quickbooks_sdk_period(period_dict: dict[str, str]) -> tuple[str, str]:
    start_date = str(period_dict.get("start_date") or "").strip()
    end_date = str(period_dict.get("end_date") or "").strip()
    if not start_date or not end_date:
        raise ValueError("QuickBooks SDK period_dict must include non-empty start_date and end_date values")

    try:
        start = datetime.strptime(start_date, "%Y-%m-%d")
        end = datetime.strptime(end_date, "%Y-%m-%d")
    except ValueError as exc:
        raise ValueError("QuickBooks SDK period_dict dates must use YYYY-MM-DD format") from exc

    if end < start:
        raise ValueError("QuickBooks SDK period_dict end_date must be on or after start_date")

    return start_date, end_date

def _get_quickbooks_dsn() -> str:
    configured_dsn = str(os.getenv("QUICKBOOKS_DSN", "QuickBooks Data QRemote") or "").strip()
    return configured_dsn or "QuickBooks Data QRemote"


def fetch_quickbooks_data(sql_query: str) -> list[dict]:
    """
    Fetch data from QuickBooks via QODBC.
    Args:
        sql_query: SQL query string (e.g., 'SELECT * FROM Invoice')
    Returns:
        List of rows as dicts.
    """
    if pyodbc is None:
        raise RuntimeError("pyodbc is required for QuickBooks ODBC access but is not installed in this environment")
    conn_str = f"DSN={_get_quickbooks_dsn()};"
    with pyodbc.connect(conn_str, autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute(sql_query)
        columns = [col[0] for col in cursor.description]
        results = [dict(zip(columns, row)) for row in cursor.fetchall()]
    return results
# Place business logic here as you migrate from scripts/


def validate_quickbooks_diagnostic_sql(sql_query: str) -> str:
    normalized = str(sql_query or "").strip()
    if not normalized:
        raise ValueError("QuickBooks diagnostic SQL only allows read-only SELECT statements")
    if len(normalized) > MAX_QUICKBOOKS_DIAGNOSTIC_SQL_LENGTH:
        raise ValueError(
            f"QuickBooks diagnostic SQL must be {MAX_QUICKBOOKS_DIAGNOSTIC_SQL_LENGTH} characters or fewer"
        )

    lowered = normalized.lower()
    if not lowered.startswith("select"):
        raise ValueError("QuickBooks diagnostic SQL only allows read-only SELECT statements")

    forbidden_tokens = (
        " insert ",
        " update ",
        " delete ",
        " merge ",
        " drop ",
        " alter ",
        " truncate ",
        " create ",
        " replace ",
        " grant ",
        " revoke ",
        " execute ",
        " exec ",
    )
    tokenized = f" {lowered.replace(chr(10), ' ').replace(chr(13), ' ').replace(chr(9), ' ')} "
    if ";" in normalized[:-1] or any(token in tokenized for token in forbidden_tokens):
        raise ValueError("QuickBooks diagnostic SQL only allows read-only SELECT statements")

    return normalized


def _quickbooks_import_dir() -> Path:
    configured = _read_project_env_path("QUICKBOOKS_IMPORT_DIR")
    if configured is not None:
        return configured
    return _project_root() / "app" / "data" / "imports" / "quickbooks"


def _softdent_import_dir() -> Path:
    configured = _read_project_env_path("SOFTDENT_IMPORT_DIR")
    if configured is not None:
        return configured
    return _project_root() / "app" / "data" / "imports" / "softdent"


def _softdent_bridge_export_root() -> Path:
    configured = _read_project_env_path("SOFTDENT_BRIDGE_EXPORT_ROOT")
    if configured is not None:
        return configured
    return _project_root() / "bridge" / "exports"


def _quickbooks_export_candidates(topic: str) -> list[Path]:
    if topic not in QUICKBOOKS_EXPORT_ENV_BY_TOPIC:
        raise ValueError(f"Unsupported QuickBooks export topic: {topic}")

    env_var = QUICKBOOKS_EXPORT_ENV_BY_TOPIC[topic]
    default_names = QUICKBOOKS_EXPORT_DEFAULT_NAMES_BY_TOPIC[topic]
    candidates: list[Path] = []
    configured = os.getenv(env_var, "").strip()
    if configured:
        candidates.append(Path(configured).expanduser().resolve())

    import_dir = _quickbooks_import_dir()
    candidates.extend(import_dir / name for name in default_names)

    deduped: list[Path] = []
    seen: set[Path] = set()
    for candidate in candidates:
        if candidate in seen:
            continue
        seen.add(candidate)
        deduped.append(candidate)
    return deduped


def _load_quickbooks_export(topic: str) -> tuple[list[dict], Path | None, str]:
    for path in _quickbooks_export_candidates(topic):
        if not path.exists() or not path.is_file():
            continue
        suffix = path.suffix.lower()
        if suffix == ".json":
            try:
                payload = json.loads(path.read_text(encoding="utf-8"))
            except (OSError, json.JSONDecodeError):
                continue
            rows = _extract_json_rows(payload)
            if rows:
                return rows, path, "json"
        elif suffix == ".csv":
            try:
                with path.open("r", encoding="utf-8-sig", newline="") as handle:
                    rows = [row for row in csv.DictReader(handle) if isinstance(row, dict)]
            except OSError:
                continue
            if rows:
                return rows, path, "csv"
        elif suffix in {".xlsx", ".xlsm", ".xls"}:
            rows = _load_excel_export_rows(path)
            if rows:
                return rows, path, "excel"
    return [], None, "missing"


def load_quickbooks_export_rows(topic: str) -> list[dict]:
    rows, _, _ = _load_quickbooks_export(topic)
    return rows


def get_quickbooks_source_status(topic: str) -> dict[str, object]:
    _, source_path, source_backend = _load_quickbooks_export(topic)
    return _build_optional_source_status(source_path, source_backend)


def get_local_accounting_documents_status() -> dict[str, object]:
    db_path = _local_accounting_db_path()
    if not db_path.exists():
        return {
            "available": False,
            "source_backend": "sqlite",
            "source_file": db_path.name,
            "document_count": 0,
            "modified_at_utc": "",
        }

    with sqlite3.connect(db_path) as connection:
        try:
            row = connection.execute(
                "SELECT COUNT(*) AS document_count, MAX(processed_at_utc) AS latest_processed_at FROM local_accounting_documents"
            ).fetchone()
        except sqlite3.OperationalError as exc:
            if "no such table" not in str(exc).lower():
                raise
            row = None

    document_count = int(row[0] or 0) if row is not None else 0
    latest_processed_at = str(row[1] or "") if row is not None else ""
    return {
        "available": True,
        "source_backend": "sqlite",
        "source_file": db_path.name,
        "document_count": document_count,
        "modified_at_utc": latest_processed_at,
    }

def _project_root() -> Path:
    return Path(__file__).resolve().parents[1]


def _read_project_env_path(name: str) -> Path | None:
    configured = os.getenv(name, "").strip()
    if not configured:
        return None
    candidate = Path(configured).expanduser()
    if not candidate.is_absolute():
        candidate = _project_root() / candidate
    return candidate.resolve()


def stage_hal_import_files(files: list[dict[str, str]], *, actor: str) -> dict[str, object]:
    written_files: list[dict[str, object]] = []
    workspace_root = get_ai_workspace_path()
    staging_root = ensure_within_ai_workspace(workspace_root / "hal_staged_imports")
    staging_root.mkdir(parents=True, exist_ok=True)

    for file in files:
        file_name = str(file.get("file_name") or "").strip()
        if file_name not in HAL_ALLOWED_STAGED_IMPORT_FILES:
            raise ValueError(f"File is not approved for HAL staging: {file_name}")

        content = str(file.get("content") or "")
        destination = ensure_within_ai_workspace(staging_root / file_name)
        destination.write_text(content, encoding="utf-8", newline="")
        written_files.append(
            {
                "file_name": file_name,
                "bytes_written": len(content.encode("utf-8")),
                "destination_path": str(destination),
            }
        )

    append_ai_activity_log(
        tier="tier_1",
        actor=actor,
        action="stage-hal-import-files",
        detail=f"Wrote {len(written_files)} approved staged import file(s) into AI_Workspace.",
    )

    return {
        "message": "Approved HAL staging files were written to AI_Workspace.",
        "actor": actor,
        "file_count": len(written_files),
        "files": written_files,
    }


def _local_accounting_db_path() -> Path:
    configured = os.getenv("LOCAL_AI_ACCOUNTING_DB_PATH", "").strip()
    if configured:
        return resolve_within_hal_allowed_base(Path(configured), label="Local accounting database path")
    return get_hal_storage_path()


def _local_accounting_table_columns(connection: sqlite3.Connection) -> set[str]:
    return {
        str(row[1])
        for row in connection.execute("PRAGMA table_info(local_accounting_documents)").fetchall()
    }


def _deserialize_correction_flags(value: object) -> list[str]:
    if isinstance(value, list):
        return [str(item) for item in value if str(item).strip()]
    if not isinstance(value, str) or not value.strip():
        return []
    try:
        payload = json.loads(value)
    except json.JSONDecodeError:
        return []
    if not isinstance(payload, list):
        return []
    return [str(item) for item in payload if str(item).strip()]


def _fallback_local_accounting_correction_flags(item: dict[str, object]) -> list[str]:
    flags: list[str] = []
    raw_upper = str(item.get("raw_text") or "").upper()
    vendor_name = str(item.get("vendor_name") or "").strip()
    invoice_number = str(item.get("invoice_number") or "").strip()
    document_date = str(item.get("document_date") or "").strip()
    if vendor_name and vendor_name.upper() not in raw_upper:
        flags.append("vendor_normalized")
    if invoice_number and invoice_number.upper() not in raw_upper:
        flags.append("invoice_corrected")
    if document_date and document_date.upper() not in raw_upper:
        flags.append("date_corrected")
    return flags


def _fallback_local_accounting_confidence(item: dict[str, object], correction_flags: list[str]) -> str:
    correction_count = len(correction_flags)
    if correction_count > 2:
        return "manual review"
    if correction_count > 0:
        return "review suggested"
    if str(item.get("extractor") or "") in {"plain_text", "pdf_text"}:
        return "high confidence"
    if len(str(item.get("raw_text") or "").strip()) > 40:
        return "medium confidence"
    return "review suggested"


def list_local_accounting_documents(
    *,
    limit: int = 20,
    document_type: str | None = None,
    search: str | None = None,
    review_only: bool = False,
) -> dict[str, object]:
    db_path = _local_accounting_db_path()
    if not db_path.exists():
        return {
            "count": 0,
            "limit": limit,
            "document_type": document_type,
            "search": search,
            "review_only": review_only,
            "items": [],
        }

    with sqlite3.connect(db_path) as connection:
        connection.row_factory = sqlite3.Row
        try:
            columns = _local_accounting_table_columns(connection)
            where_clauses: list[str] = []
            params: list[object] = []
            if document_type:
                where_clauses.append("document_type = ?")
                params.append(document_type)
            if search:
                where_clauses.append("(source_name LIKE ? OR vendor_name LIKE ? OR invoice_number LIKE ? OR raw_text LIKE ?)")
                needle = f"%{search}%"
                params.extend([needle, needle, needle, needle])
            if review_only and "review_required" in columns:
                where_clauses.append("review_required = 1")

            where_sql = f"WHERE {' AND '.join(where_clauses)}" if where_clauses else ""
            correction_flags_sql = "correction_flags_json" if "correction_flags_json" in columns else "'[]' AS correction_flags_json"
            confidence_sql = "confidence_label" if "confidence_label" in columns else "'' AS confidence_label"
            review_sql = "review_required" if "review_required" in columns else "0 AS review_required"
            query = f"""
                SELECT
                    id,
                    source_path,
                    source_name,
                    sha256,
                    processed_at_utc,
                    extractor,
                    document_type,
                    vendor_name,
                    invoice_number,
                    document_date,
                    total_amount,
                    subtotal_amount,
                    tax_amount,
                    currency,
                    text_preview,
                    raw_text,
                    {correction_flags_sql},
                    {confidence_sql},
                    {review_sql}
                FROM local_accounting_documents
                {where_sql}
                ORDER BY processed_at_utc DESC, id DESC
                LIMIT ?
            """
            params.append(limit)
            rows = connection.execute(query, params).fetchall()
        except sqlite3.OperationalError as exc:
            if "no such table" in str(exc).lower():
                return {
                    "count": 0,
                    "limit": limit,
                    "document_type": document_type,
                    "search": search,
                    "review_only": review_only,
                    "items": [],
                }
            raise

    items = [dict(row) for row in rows]
    has_review_column = any("review_required" in item for item in items)
    for item in items:
        correction_flags = _deserialize_correction_flags(item.get("correction_flags_json"))
        if not correction_flags:
            correction_flags = _fallback_local_accounting_correction_flags(item)
        item["correction_flags"] = correction_flags
        item["review_required"] = bool(item.get("review_required")) or bool(correction_flags)
        item["confidence_label"] = str(item.get("confidence_label") or "").strip() or _fallback_local_accounting_confidence(item, correction_flags)
        item.pop("correction_flags_json", None)
    if review_only and not has_review_column:
        items = [item for item in items if item.get("review_required")][:limit]
    return {
        "count": len(items),
        "limit": limit,
        "document_type": document_type,
        "search": search,
        "review_only": review_only,
        "items": items,
    }


def _coerce_float(value) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def load_softdent_dashboard_rows() -> list[dict]:
    rows, _, _ = _load_optional_tabular_export(
        env_var="SOFTDENT_DASHBOARD_EXPORT_PATH",
        default_names=(
            "softdent_dashboard_data.json",
            "softdent_dashboard_export.json",
            "softdent_dashboard_data.csv",
            "softdent_dashboard_export.csv",
            "softdent_dashboard_data.xlsx",
            "softdent_dashboard_export.xlsx",
            "softdent_dashboard_data.xls",
            "softdent_dashboard_export.xls",
        ),
    )
    return rows


def fetch_softdent_dashboard_aggregate(snapshot_path: str | None = None) -> dict[str, object]:
    rows = load_softdent_dashboard_rows()
    source_file = Path(snapshot_path).name if snapshot_path else "softdent_dashboard_data.json"
    if not rows:
        return SoftDentAggregateSnapshot(
            source_file=source_file,
            period_start="",
            period_end="",
            provider_count=0,
            provider_rows=[],
            totals={
                "production": 0.0,
                "collections": 0.0,
                "insurance": 0.0,
                "patient": 0.0,
            },
            data_complete=False,
        ).model_dump()

    normalized_rows = []
    for index, row in enumerate(rows, start=1):
        provider_name = str(row.get("provider") or row.get("Provider") or "Unknown")
        normalized_rows.append(
            {
                "provider_id": _normalize_provider_id(provider_name, index=index),
                "provider_name": provider_name,
                "period": str(row.get("period") or row.get("Period") or ""),
                "production_amount": _coerce_float(row.get("production") or row.get("Production")),
                "collection_amount": _coerce_float(row.get("collections") or row.get("Collections")),
                "insurance_amount": _coerce_float(row.get("insurance") or row.get("Insurance")),
                "patient_amount": _coerce_float(row.get("patient") or row.get("Patient")),
            }
        )

    periods = [row["period"] for row in normalized_rows if row["period"]]
    period_value = max(periods) if periods else ""
    period_start, period_end = _derive_softdent_period_bounds(period_value)
    totals = {
        "production": round(sum(row["production_amount"] for row in normalized_rows), 2),
        "collections": round(sum(row["collection_amount"] for row in normalized_rows), 2),
        "insurance": round(sum(row["insurance_amount"] for row in normalized_rows), 2),
        "patient": round(sum(row["patient_amount"] for row in normalized_rows), 2),
    }
    data_complete = bool(period_start and period_end) and all(row["provider_name"] for row in normalized_rows)

    return SoftDentAggregateSnapshot(
        source_file=source_file,
        period_start=period_start,
        period_end=period_end,
        provider_count=len(normalized_rows),
        provider_rows=[
            ProviderProductionRow(
                provider_id=row["provider_id"],
                provider_name=row["provider_name"],
                production_amount=row["production_amount"],
                collection_amount=row["collection_amount"],
            )
            for row in normalized_rows
        ],
        totals=totals,
        data_complete=data_complete,
    ).model_dump()


def load_softdent_claim_rows() -> list[dict]:
    rows, _, _ = _load_optional_tabular_export(
        env_var="SOFTDENT_CLAIMS_EXPORT_PATH",
        default_names=(
            "softdent_claims_data.json",
            "softdent_claims_export.json",
            "softdent_claims_data.csv",
            "softdent_claims_export.csv",
            "softdent_claims_data.xlsx",
            "softdent_claims_export.xlsx",
            "softdent_claims_data.xls",
            "softdent_claims_export.xls",
        ),
    )
    return [_validate_softdent_claim_row(_normalize_softdent_claim_row(row)) for row in rows]


def load_softdent_clinical_note_rows() -> list[dict]:
    rows, _, _ = _load_optional_tabular_export(
        env_var="SOFTDENT_CLINICAL_NOTES_EXPORT_PATH",
        default_names=(
            "softdent_clinical_notes_data.json",
            "softdent_clinical_notes_export.json",
            "softdent_clinical_notes_data.csv",
            "softdent_clinical_notes_export.csv",
            "softdent_clinical_notes_data.xlsx",
            "softdent_clinical_notes_export.xlsx",
            "softdent_clinical_notes_data.xls",
            "softdent_clinical_notes_export.xls",
            "softdent_notes_data.json",
            "softdent_notes_export.csv",
            "softdent_notes_data.xlsx",
            "softdent_notes_export.xlsx",
            "softdent_notes_data.xls",
            "softdent_notes_export.xls",
        ),
    )
    return [_validate_softdent_clinical_note_row(_normalize_softdent_clinical_note_row(row)) for row in rows]


def get_softdent_source_status() -> dict[str, object]:
    _, source_path, source_backend = _load_optional_tabular_export(
        env_var="SOFTDENT_DASHBOARD_EXPORT_PATH",
        default_names=(
            "softdent_dashboard_data.json",
            "softdent_dashboard_export.json",
            "softdent_dashboard_data.csv",
            "softdent_dashboard_export.csv",
            "softdent_dashboard_data.xlsx",
            "softdent_dashboard_export.xlsx",
            "softdent_dashboard_data.xls",
            "softdent_dashboard_export.xls",
        ),
    )
    return _build_optional_source_status(source_path, source_backend)


def get_softdent_claim_source_status() -> dict[str, object]:
    _, source_path, source_backend = _load_optional_tabular_export(
        env_var="SOFTDENT_CLAIMS_EXPORT_PATH",
        default_names=(
            "softdent_claims_data.json",
            "softdent_claims_export.json",
            "softdent_claims_data.csv",
            "softdent_claims_export.csv",
            "softdent_claims_data.xlsx",
            "softdent_claims_export.xlsx",
            "softdent_claims_data.xls",
            "softdent_claims_export.xls",
        ),
    )
    return _build_optional_source_status(source_path, source_backend)


def get_softdent_clinical_note_source_status() -> dict[str, object]:
    _, source_path, source_backend = _load_optional_tabular_export(
        env_var="SOFTDENT_CLINICAL_NOTES_EXPORT_PATH",
        default_names=(
            "softdent_clinical_notes_data.json",
            "softdent_clinical_notes_export.json",
            "softdent_clinical_notes_data.csv",
            "softdent_clinical_notes_export.csv",
            "softdent_clinical_notes_data.xlsx",
            "softdent_clinical_notes_export.xlsx",
            "softdent_clinical_notes_data.xls",
            "softdent_clinical_notes_export.xls",
            "softdent_notes_data.json",
            "softdent_notes_export.csv",
            "softdent_notes_data.xlsx",
            "softdent_notes_export.xlsx",
            "softdent_notes_data.xls",
            "softdent_notes_export.xls",
        ),
    )
    return _build_optional_source_status(source_path, source_backend)


def build_softdent_snapshot() -> dict[str, object]:
    aggregate = fetch_softdent_dashboard_aggregate()
    raw_provider_rows = aggregate.get("provider_rows") if isinstance(aggregate, dict) else None
    if not isinstance(raw_provider_rows, list) or not raw_provider_rows:
        return {
            "available": False,
            "period": "",
            "provider_count": 0,
            "providers": [],
            "totals": {"production": 0.0, "collections": 0.0, "insurance": 0.0, "patient": 0.0},
        }

    provider_rows = [row for row in raw_provider_rows if isinstance(row, dict)]
    if not provider_rows:
        return {
            "available": False,
            "period": "",
            "provider_count": 0,
            "providers": [],
            "totals": {"production": 0.0, "collections": 0.0, "insurance": 0.0, "patient": 0.0},
        }

    raw_totals = aggregate.get("totals") if isinstance(aggregate.get("totals"), dict) else {}
    totals = {
        "production": round(_coerce_float(raw_totals.get("production")), 2),
        "collections": round(_coerce_float(raw_totals.get("collections")), 2),
        "insurance": round(_coerce_float(raw_totals.get("insurance")), 2),
        "patient": round(_coerce_float(raw_totals.get("patient")), 2),
    }
    period_start = str(aggregate.get("period_start") or "")
    period = period_start[:7] if period_start else ""
    provider_count = int(aggregate.get("provider_count") or len(provider_rows))
    return {
        "available": True,
        "period": period,
        "provider_count": provider_count,
        "providers": [
            {
                "provider": str(row.get("provider_name") or "Unknown"),
                "period": period,
                "production": _coerce_float(row.get("production_amount")),
                "collections": _coerce_float(row.get("collection_amount")),
                "insurance": _find_softdent_provider_insurance(str(row.get("provider_name") or "Unknown")),
                "patient": _find_softdent_provider_patient(str(row.get("provider_name") or "Unknown")),
            }
            for row in provider_rows
        ],
        "totals": totals,
    }


def get_softdent_data_coverage() -> dict[str, object]:
    dashboard_rows = load_softdent_dashboard_rows()
    dashboard_status = get_softdent_source_status()
    claim_rows = load_softdent_claim_rows()
    claim_status = get_softdent_claim_source_status()
    snapshot = build_softdent_snapshot()
    dashboard_source_available = bool(dashboard_status.get("available"))
    claim_source_available = bool(claim_status.get("available"))
    snapshot_available = bool(snapshot.get("available"))
    snapshot_period = str(snapshot.get("period") or "")

    rows = [
        _build_softdent_base_coverage_row(
            key="dashboardSnapshot",
            label="Dashboard Snapshot",
            status="available" if dashboard_source_available and snapshot_available else "missing",
            summary=(
                "Dashboard snapshot aggregate export is available."
                if dashboard_source_available and snapshot_available
                else "Dashboard snapshot aggregate export is not available from the canonical SoftDent import lane."
            ),
            required_report="softdent_dashboard_data.json in project workspace",
            action="Refresh the approved SoftDent dashboard snapshot export into the canonical import lane.",
            source_status=dashboard_status,
            row_count=len(dashboard_rows),
            last_period=snapshot_period,
        ),
        _build_softdent_base_coverage_row(
            key="claimsExport",
            label="Claims Export",
            status="available" if claim_source_available and bool(claim_rows) else "missing",
            summary=(
                "Claims export is available for HAL claim workflows."
                if claim_source_available and bool(claim_rows)
                else "Claims export is not available from the canonical SoftDent import lane."
            ),
            required_report="softdent_claims_export.csv in project workspace",
            action="Refresh the approved SoftDent claims export into the canonical import lane.",
            source_status=claim_status,
            row_count=len(claim_rows),
            last_period=_detect_last_period(claim_rows),
        ),
        _build_softdent_base_coverage_row(
            key="currentPeriodTotals",
            label="Current Period Totals",
            status="available" if snapshot_available else "missing",
            summary=(
                "Current period totals are available from the approved SoftDent aggregate snapshot."
                if snapshot_available
                else "Current period totals are unavailable because the approved SoftDent aggregate snapshot is missing."
            ),
            required_report="softdent_dashboard_data.json in project workspace",
            action="Refresh the approved SoftDent dashboard snapshot export into the canonical import lane.",
            source_status=dashboard_status,
            row_count=len(dashboard_rows),
            last_period=snapshot_period,
        ),
        _build_softdent_base_coverage_row(
            key="dailyProduction",
            label="Daily Production",
            status="available" if snapshot_available else "missing",
            summary=(
                "Daily production is available from the approved SoftDent aggregate snapshot."
                if snapshot_available
                else "Daily production is unavailable because the approved SoftDent aggregate snapshot is missing."
            ),
            required_report="softdent_dashboard_data.json in project workspace",
            action="Refresh the approved SoftDent dashboard snapshot export into the canonical import lane.",
            source_status=dashboard_status,
            row_count=len(dashboard_rows),
            last_period=snapshot_period,
        ),
        _build_softdent_base_coverage_row(
            key="dailyCollections",
            label="Daily Collections",
            status="available" if snapshot_available else "missing",
            summary=(
                "Daily collections are available from the approved SoftDent aggregate snapshot."
                if snapshot_available
                else "Daily collections are unavailable because the approved SoftDent aggregate snapshot is missing."
            ),
            required_report="softdent_dashboard_data.json in project workspace",
            action="Refresh the approved SoftDent dashboard snapshot export into the canonical import lane.",
            source_status=dashboard_status,
            row_count=len(dashboard_rows),
            last_period=snapshot_period,
        ),
        _build_softdent_base_coverage_row(
            key="totalAr",
            label="Total A/R",
            status="available" if snapshot_available else "missing",
            summary=(
                "Total A/R coverage is available from the approved SoftDent aggregate snapshot."
                if snapshot_available
                else "Total A/R coverage is unavailable because the approved SoftDent aggregate snapshot is missing."
            ),
            required_report="softdent_dashboard_data.json in project workspace",
            action="Refresh the approved SoftDent dashboard snapshot export into the canonical import lane.",
            source_status=dashboard_status,
            row_count=len(dashboard_rows),
            last_period=snapshot_period,
        ),
        _build_softdent_base_coverage_row(
            key="recareSummary",
            label="Recare Summary",
            status="available" if snapshot_available else "missing",
            summary=(
                "Recare summary coverage is available from the approved SoftDent aggregate snapshot."
                if snapshot_available
                else "Recare summary coverage is unavailable because the approved SoftDent aggregate snapshot is missing."
            ),
            required_report="softdent_dashboard_data.json in project workspace",
            action="Refresh the approved SoftDent dashboard snapshot export into the canonical import lane.",
            source_status=dashboard_status,
            row_count=len(dashboard_rows),
            last_period=snapshot_period,
        ),
        _build_softdent_optional_export_coverage_row(
            key="trueOutstandingClaims",
            label="True Outstanding Claims",
            env_var=SOFTDENT_OUTSTANDING_CLAIMS_EXPORT_ENV,
            default_names=("outstanding_claims_by_company.csv",),
            missing_summary="True Outstanding Claims aggregate export is missing from the canonical SoftDent import lane.",
            available_summary="True Outstanding Claims aggregate export is available.",
            required_report=r"outstanding_claims_by_company.csv in C:\SoftDentReportExports",
            action=r"Configure the automated SoftDent source pipeline to emit aggregate-only outstanding_claims_by_company.csv into C:\SoftDentReportExports.",
        ),
        _build_softdent_optional_export_coverage_row(
            key="unsubmittedClaims",
            label="Unsubmitted Claims",
            env_var=SOFTDENT_UNSUBMITTED_CLAIMS_EXPORT_ENV,
            default_names=("unsubmitted_claims.csv",),
            missing_summary="Unsubmitted Claims aggregate export is missing from the canonical SoftDent import lane.",
            available_summary="Unsubmitted Claims aggregate export is available.",
            required_report=r"unsubmitted_claims.csv in C:\SoftDentReportExports",
            action=r"Configure the automated SoftDent source pipeline to emit aggregate-only unsubmitted_claims.csv into C:\SoftDentReportExports.",
        ),
        _build_softdent_optional_export_coverage_row(
            key="insuranceIncome",
            label="Insurance Income",
            env_var=SOFTDENT_INSURANCE_INCOME_EXPORT_ENV,
            default_names=("insurance_income.csv",),
            missing_summary="Insurance Income aggregate export is missing from the canonical SoftDent import lane.",
            available_summary="Insurance Income aggregate export is available.",
            required_report=r"insurance_income.csv in C:\SoftDentReportExports",
            action=r"Configure the automated SoftDent source pipeline to emit aggregate-only insurance_income.csv into C:\SoftDentReportExports.",
        ),
        _build_softdent_optional_export_coverage_row(
            key="insurancePaymentDistribution",
            label="Insurance Payment Distribution",
            env_var=SOFTDENT_INSURANCE_PAYMENT_DISTRIBUTION_EXPORT_ENV,
            default_names=("insurance_payment_distribution.csv",),
            missing_summary="Insurance Payment Distribution aggregate export is missing from the canonical SoftDent import lane.",
            available_summary="Insurance Payment Distribution aggregate export is available.",
            required_report=r"insurance_payment_distribution.csv in C:\SoftDentReportExports",
            action=r"Configure the automated SoftDent source pipeline to emit aggregate-only insurance_payment_distribution.csv into C:\SoftDentReportExports.",
        ),
        _build_softdent_optional_export_coverage_row(
            key="insuranceCheckDistribution",
            label="Insurance Check Distribution",
            env_var=SOFTDENT_INSURANCE_CHECK_DISTRIBUTION_EXPORT_ENV,
            default_names=("insurance_check_distribution.csv",),
            missing_summary="Insurance Check Distribution aggregate export is missing from the canonical SoftDent import lane.",
            available_summary="Insurance Check Distribution aggregate export is available.",
            required_report=r"insurance_check_distribution.csv in C:\SoftDentReportExports",
            action=r"Configure the automated SoftDent source pipeline to emit aggregate-only insurance_check_distribution.csv into C:\SoftDentReportExports.",
        ),
        _build_softdent_optional_export_coverage_row(
            key="treatmentPlans",
            label="Treatment Plans",
            env_var=SOFTDENT_TREATMENT_PLAN_EXPORT_ENV,
            default_names=("treatment_plan_summary.csv",),
            missing_summary="Treatment Plans aggregate export is missing from the canonical SoftDent import lane.",
            available_summary="Treatment Plans aggregate export is available.",
            required_report="treatment_plan_summary.csv or Treatment plan entity/report or staged Database Extractor snapshot",
            action="Configure the automated SoftDent source pipeline to emit aggregate-only treatment_plan_summary.csv, a supported treatment plan entity report, or a staged Database Extractor snapshot with count/amount/status fields.",
        ),
        _build_softdent_optional_export_coverage_row(
            key="paymentPlans",
            label="Payment Plans",
            env_var=SOFTDENT_PAYMENT_PLAN_EXPORT_ENV,
            default_names=("payment_plans.csv",),
            missing_summary="Payment Plans aggregate export is missing from the canonical SoftDent import lane.",
            available_summary="Payment Plans aggregate export is available.",
            required_report="payment_plans.csv or Payment plan entity/report or staged Database Extractor snapshot",
            action="Configure the automated SoftDent source pipeline to emit aggregate-only payment_plans.csv, a supported payment plan entity report, or a staged Database Extractor snapshot with count/amount/status fields.",
        ),
        _build_softdent_limited_coverage_row(
            key="transactionFeed",
            label="Transaction Feed",
            rows=claim_rows,
            source_status=claim_status,
            action="Restore live Sensei DataSync transaction emission to unlock and validate ledger-backed collections, write-offs, insurance detail, and payment distribution.",
        ),
        _build_softdent_limited_coverage_row(
            key="fourYearHistory",
            label="Four-Year History",
            rows=dashboard_rows,
            source_status=dashboard_status,
            action="Configure automated historical yearly period reports for the SoftDent pipeline.",
        ),
        _build_softdent_limited_coverage_row(
            key="topAdaCodes",
            label="Top ADA Codes",
            rows=dashboard_rows,
            source_status=dashboard_status,
            action="Validate appointment-derived production against official SoftDent production reports or the after-hours snapshot before treating it as authoritative.",
            force_unknown_period=True,
        ),
        _build_softdent_limited_coverage_row(
            key="providerProduction",
            label="Provider Production",
            rows=dashboard_rows,
            source_status=dashboard_status,
            action="Validate appointment-derived production against official SoftDent production reports or the after-hours snapshot before treating it as authoritative.",
            force_unknown_period=True,
        ),
        _build_softdent_limited_coverage_row(
            key="writeOffsAdjustments",
            label="Write-offs & Adjustments",
            rows=claim_rows,
            source_status=claim_status,
            action="Connect Transaction feed or staged Database Extractor snapshot to populate ledger-backed write-offs and adjustments.",
        ),
    ]

    counts = {
        "missing": sum(1 for row in rows if row["status"] == "missing"),
        "limited": sum(1 for row in rows if row["status"] == "limited"),
        "available": sum(1 for row in rows if row["status"] == "available"),
    }
    return {
        "summary": "Missing and limited reports explain why some dashboard charts are unavailable.",
        "counts": counts,
        "rows": rows,
    }


def get_softdent_coverage_metrics() -> dict[str, object]:
    return {
        "trueOutstandingClaims": _build_softdent_claim_metric(
            label="True Outstanding Claims",
            env_var=SOFTDENT_OUTSTANDING_CLAIMS_EXPORT_ENV,
            default_names=("outstanding_claims_by_company.csv",),
            amount_aliases=("Outstanding_Amount", "OutstandingAmount", "Amount", "Balance"),
            count_aliases=("Claim_Count", "ClaimCount", "Count"),
        ),
        "unsubmittedClaims": _build_softdent_claim_metric(
            label="Unsubmitted Claims",
            env_var=SOFTDENT_UNSUBMITTED_CLAIMS_EXPORT_ENV,
            default_names=("unsubmitted_claims.csv",),
            amount_aliases=("Unsubmitted_Amount", "UnsubmittedAmount", "Amount", "Balance"),
            count_aliases=("Claim_Count", "ClaimCount", "Count"),
        ),
    }


def _build_softdent_base_coverage_row(
    *,
    key: str,
    label: str,
    status: str,
    summary: str,
    required_report: str,
    action: str,
    source_status: dict[str, object],
    row_count: int,
    last_period: str,
) -> dict[str, object]:
    return {
        "key": key,
        "label": label,
        "status": status,
        "summary": summary,
        "requiredReport": required_report,
        "action": action,
        "sourceFile": str(source_status.get("source_file") or ""),
        "sourceBackend": str(source_status.get("source_backend") or "missing"),
        "modifiedAtUtc": str(source_status.get("modified_at_utc") or ""),
        "rowCount": int(row_count or 0),
        "lastPeriod": str(last_period or ""),
    }


def _build_softdent_optional_export_coverage_row(
    *,
    key: str,
    label: str,
    env_var: str,
    default_names: tuple[str, ...],
    missing_summary: str,
    available_summary: str,
    required_report: str,
    action: str,
) -> dict[str, object]:
    rows, source_path, source_backend = _load_optional_tabular_export(env_var=env_var, default_names=default_names)
    source_status = _build_optional_source_status(source_path, source_backend)
    available = bool(rows) and source_status["available"]
    return {
        "key": key,
        "label": label,
        "status": "available" if available else "missing",
        "summary": available_summary if available else missing_summary,
        "requiredReport": required_report,
        "action": action,
        "sourceFile": str(source_status.get("source_file") or ""),
        "sourceBackend": str(source_status.get("source_backend") or "missing"),
        "modifiedAtUtc": str(source_status.get("modified_at_utc") or ""),
        "rowCount": len(rows),
        "lastPeriod": _detect_last_period(rows),
    }


def _build_softdent_limited_coverage_row(
    *,
    key: str,
    label: str,
    rows: list[dict],
    source_status: dict[str, object],
    action: str,
    force_unknown_period: bool = False,
) -> dict[str, object]:
    return {
        "key": key,
        "label": label,
        "status": "limited",
        "summary": f"{label} is only partially available from aggregate-only sources.",
        "requiredReport": "Aggregate-only staged SoftDent source",
        "action": action,
        "sourceFile": str(source_status.get("source_file") or ""),
        "sourceBackend": str(source_status.get("source_backend") or "missing"),
        "modifiedAtUtc": str(source_status.get("modified_at_utc") or ""),
        "rowCount": len(rows),
        "lastPeriod": "unknown" if force_unknown_period else _detect_last_period(rows),
    }


def _build_softdent_claim_metric(
    *,
    label: str,
    env_var: str,
    default_names: tuple[str, ...],
    amount_aliases: tuple[str, ...],
    count_aliases: tuple[str, ...],
) -> dict[str, object]:
    rows, source_path, source_backend = _load_optional_tabular_export(env_var=env_var, default_names=default_names)
    source_status = _build_optional_source_status(source_path, source_backend)
    breakdown = _build_softdent_metric_breakdown(rows, amount_aliases=amount_aliases, count_aliases=count_aliases)
    item_count = sum(int(item["count"]) for item in breakdown)
    total_amount = round(sum(float(item["amount"]) for item in breakdown), 2)
    available = bool(rows) and source_status["available"]
    return {
        "label": label,
        "available": available,
        "sourceFile": str(source_status.get("source_file") or ""),
        "sourceBackend": str(source_status.get("source_backend") or "missing"),
        "modifiedAtUtc": str(source_status.get("modified_at_utc") or ""),
        "rowCount": len(rows),
        "itemCount": item_count,
        "totalAmount": total_amount,
        "lastPeriod": _detect_last_period(rows),
        "summary": f"{label} exposure is available." if available else f"{label} exposure is not available yet.",
        "breakdown": breakdown,
    }


def _build_softdent_metric_breakdown(
    rows: list[dict], *, amount_aliases: tuple[str, ...], count_aliases: tuple[str, ...]
) -> list[dict[str, object]]:
    breakdown: list[dict[str, object]] = []
    for row in rows:
        label = _first_non_empty_value(row, "Payer", "payer", "Carrier", "carrier", "Plan", "plan") or "Unknown"
        count = int(_coerce_float(_first_non_empty_value(row, *count_aliases)))
        amount = round(_coerce_float(_first_non_empty_value(row, *amount_aliases)), 2)
        breakdown.append({"label": label, "amount": amount, "count": count})
    breakdown.sort(key=lambda item: (-float(item["amount"]), str(item["label"])))
    return breakdown


def _first_non_empty_value(row: dict[str, object], *keys: str) -> str:
    normalized_keys = {_normalize_alias(key) for key in keys}
    for key, value in row.items():
        if _normalize_alias(str(key)) in normalized_keys:
            text = str(value or "").strip()
            if text:
                return text
    return ""


def _detect_last_period(rows: list[dict]) -> str:
    candidate_values: list[str] = []
    for row in rows:
        for key, value in row.items():
            normalized_key = _normalize_alias(str(key))
            if normalized_key not in {
                "period",
                "reportperiod",
                "yearmonth",
                "servicedate",
                "notedate",
                "date",
                "entrydate",
                "transactiondate",
                "asofdate",
            }:
                continue
            normalized_value = _normalize_period_value(str(value or "").strip())
            if normalized_value:
                candidate_values.append(normalized_value)
    return max(candidate_values) if candidate_values else ""


def _normalize_period_value(value: str) -> str:
    if not value:
        return ""
    if len(value) >= 10:
        try:
            parsed = datetime.strptime(value[:10], "%Y-%m-%d")
            return parsed.strftime("%Y-%m-%d")
        except ValueError:
            pass
    if len(value) >= 7:
        try:
            parsed = datetime.strptime(value[:7], "%Y-%m")
            return parsed.strftime("%Y-%m")
        except ValueError:
            pass
    return ""


def _load_optional_tabular_export(*, env_var: str, default_names: tuple[str, ...]) -> tuple[list[dict], Path | None, str]:
    for path in _candidate_export_paths(env_var=env_var, default_names=default_names):
        if not path.exists() or not path.is_file():
            continue
        suffix = path.suffix.lower()
        if suffix == ".json":
            try:
                payload = json.loads(path.read_text(encoding="utf-8"))
            except (OSError, json.JSONDecodeError):
                continue
            rows = _extract_json_rows(payload)
            if rows:
                return rows, path, "json"
        elif suffix == ".csv":
            try:
                with path.open("r", encoding="utf-8-sig", newline="") as handle:
                    rows = [row for row in csv.DictReader(handle) if isinstance(row, dict)]
            except OSError:
                continue
            if rows:
                return rows, path, "csv"
        elif suffix in {".xlsx", ".xlsm", ".xls"}:
            rows = _load_excel_export_rows(path)
            if rows:
                return rows, path, "excel"
    return [], None, "missing"


def _load_excel_export_rows(path: Path) -> list[dict]:
    suffix = path.suffix.lower()
    try:
        if suffix == ".xls":
            return _load_xls_rows(path)
        return _load_openxml_rows(path)
    except OSError:
        return []


def _load_openxml_rows(path: Path) -> list[dict]:
    if load_workbook is None:
        return []
    workbook = load_workbook(filename=path, read_only=True, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]] if workbook.sheetnames else None
        if sheet is None:
            return []
        return _rows_from_iterable(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()


def _load_xls_rows(path: Path) -> list[dict]:
    if xlrd is None:
        return []
    workbook = xlrd.open_workbook(path)
    if workbook.nsheets <= 0:
        return []
    sheet = workbook.sheet_by_index(0)
    return _rows_from_iterable(sheet.row_values(index) for index in range(sheet.nrows))


def _rows_from_iterable(rows: Iterable[Iterable[object]]) -> list[dict]:
    materialized = [list(row) for row in rows if any(cell not in (None, "") for cell in row)]
    if not materialized:
        return []
    headers = [str(cell).strip() if cell is not None else "" for cell in materialized[0]]
    if not any(headers):
        return []

    normalized_headers: list[str] = []
    seen: dict[str, int] = {}
    for index, header in enumerate(headers, start=1):
        base = header or f"column_{index}"
        count = seen.get(base, 0)
        seen[base] = count + 1
        normalized_headers.append(base if count == 0 else f"{base}_{count + 1}")

    records: list[dict] = []
    for row in materialized[1:]:
        padded = row + [""] * max(0, len(normalized_headers) - len(row))
        record = {
            normalized_headers[index]: "" if padded[index] is None else str(padded[index]).strip()
            for index in range(len(normalized_headers))
        }
        if any(value not in (None, "") for value in record.values()):
            records.append(record)
    return records


def _candidate_export_paths(*, env_var: str, default_names: tuple[str, ...]) -> list[Path]:
    candidates: list[Path] = []
    configured = os.getenv(env_var, "").strip()
    quickbooks_default_dir = _quickbooks_import_dir()
    softdent_default_dir = _softdent_import_dir()
    normalized_names = {name.lower() for name in default_names}
    canonical_import_dir: Path | None = None

    if any("quickbooks" in name for name in normalized_names):
        canonical_import_dir = quickbooks_default_dir
    elif any("softdent" in name for name in normalized_names):
        canonical_import_dir = softdent_default_dir

    if configured and canonical_import_dir is not None:
        configured_path = Path(configured).expanduser().resolve()
        approved_names = {name.casefold() for name in default_names}
        try:
            within_import_dir = configured_path.relative_to(canonical_import_dir)
        except ValueError:
            within_import_dir = None
        if within_import_dir is not None and configured_path.name.casefold() in approved_names:
            candidates.append(configured_path)

    if canonical_import_dir is not None:
        candidates.extend(canonical_import_dir / name for name in default_names)

    deduped: list[Path] = []
    seen: set[Path] = set()
    for candidate in candidates:
        if candidate in seen:
            continue
        seen.add(candidate)
        deduped.append(candidate)
    return deduped


def _extract_json_rows(payload: object) -> list[dict]:
    if isinstance(payload, list):
        return [row for row in payload if isinstance(row, dict)]
    if isinstance(payload, dict):
        for key in ("rows", "items", "data", "claims", "notes"):
            value = payload.get(key)
            if isinstance(value, list):
                return [row for row in value if isinstance(row, dict)]
        for value in payload.values():
            if isinstance(value, list) and all(isinstance(row, dict) for row in value):
                return list(value)
    return []


def _build_optional_source_status(source_path: Path | None, source_backend: str) -> dict[str, object]:
    if source_path is None:
        return {
            "available": False,
            "source_backend": source_backend,
            "source_file": "",
            "modified_at_utc": "",
        }
    modified_at = datetime.fromtimestamp(source_path.stat().st_mtime, tz=timezone.utc).isoformat()
    return {
        "available": True,
        "source_backend": source_backend,
        "source_file": source_path.name,
        "modified_at_utc": modified_at,
    }


def _normalize_softdent_claim_row(row: dict) -> dict:
    normalized = dict(row)
    _apply_alias(normalized, row, "PatientName", "patientname", "patient_name", "patient", "name", "patname")
    _apply_alias(normalized, row, "MRN", "mrn", "patientid", "patient_id", "chart", "chartnumber")
    _apply_alias(normalized, row, "ClaimId", "claimid", "claim_id", "claimnumber", "claim_number", "claim")
    _apply_alias(normalized, row, "ClaimStatus", "claimstatus", "claim_status", "status")
    _apply_alias(normalized, row, "Payer", "payer", "carrier", "insurance", "insurancename", "plan")
    _apply_alias(normalized, row, "Procedure", "procedure", "procdesc", "description", "servicedescription", "treatment")
    _apply_alias(normalized, row, "ServiceDate", "servicedate", "dateofservice", "dos", "date")
    _apply_alias(normalized, row, "DenialReason", "denialreason", "reason", "remark", "claimnote", "note")
    _apply_alias(normalized, row, "ClaimAmount", "claimamount", "amount", "balance")
    normalized["ClaimAmount"] = _coerce_float(normalized.get("ClaimAmount"))
    return normalized


def _normalize_softdent_clinical_note_row(row: dict) -> dict:
    normalized = dict(row)
    _apply_alias(normalized, row, "PatientName", "patientname", "patient_name", "patient", "name", "patname")
    _apply_alias(normalized, row, "MRN", "mrn", "patientid", "patient_id", "chart", "chartnumber")
    _apply_alias(normalized, row, "NoteDate", "notedate", "date", "entrydate", "servicedate")
    _apply_alias(normalized, row, "Provider", "provider", "doctor", "clinician")
    _apply_alias(normalized, row, "Procedure", "procedure", "procdesc", "description", "treatment")
    _apply_alias(normalized, row, "ClinicalNote", "clinicalnote", "note", "narrative", "assessment", "chartnote")
    return normalized


def _validate_softdent_claim_row(row: dict) -> dict:
    return PatientClaimRecord(**row).model_dump()


def _validate_softdent_clinical_note_row(row: dict) -> dict:
    return ClinicalNoteEntry(**row).model_dump()


def _apply_alias(target: dict, source: dict, canonical_key: str, *aliases: str) -> None:
    if canonical_key in target and target.get(canonical_key) not in (None, ""):
        return
    normalized_aliases = {_normalize_alias(alias) for alias in aliases}
    for key, value in source.items():
        if value in (None, ""):
            continue
        if _normalize_alias(str(key)) in normalized_aliases:
            target[canonical_key] = value
            return


def _normalize_alias(value: str) -> str:
    return "".join(character for character in value.lower() if character.isalnum())


def _normalize_provider_id(provider_name: str, *, index: int) -> str:
    normalized = _normalize_alias(provider_name)
    return normalized or f"provider-{index}"


def _derive_softdent_period_bounds(period_value: str) -> tuple[str, str]:
    value = str(period_value or "").strip()
    if not value:
        return "", ""
    try:
        if len(value) == 7:
            period_start = datetime.strptime(value, "%Y-%m")
            last_day = monthrange(period_start.year, period_start.month)[1]
            return period_start.strftime("%Y-%m-01"), f"{period_start.year:04d}-{period_start.month:02d}-{last_day:02d}"
        if len(value) == 10:
            exact_date = datetime.strptime(value, "%Y-%m-%d")
            rendered = exact_date.strftime("%Y-%m-%d")
            return rendered, rendered
    except ValueError:
        return "", ""
    return "", ""


def _find_softdent_provider_insurance(provider_name: str) -> float:
    for row in load_softdent_dashboard_rows():
        candidate = str(row.get("provider") or row.get("Provider") or "Unknown")
        if candidate == provider_name:
            return _coerce_float(row.get("insurance") or row.get("Insurance"))
    return 0.0


def _find_softdent_provider_patient(provider_name: str) -> float:
    for row in load_softdent_dashboard_rows():
        candidate = str(row.get("provider") or row.get("Provider") or "Unknown")
        if candidate == provider_name:
            return _coerce_float(row.get("patient") or row.get("Patient"))
    return 0.0


def get_kpi_data():
    snapshot = build_softdent_snapshot()
    if bool(snapshot.get("available")):
        raw_totals = snapshot.get("totals") if isinstance(snapshot.get("totals"), dict) else {}
        production = _coerce_float(raw_totals.get("production"))
        collections = _coerce_float(raw_totals.get("collections"))
        ar_balance = round(max(production - collections, 0.0), 2)
        collection_ratio = round((collections / production), 4) if production else 0.0
        return [
            {"name": "production", "value": production},
            {"name": "collections", "value": collections},
            {"name": "ar", "value": ar_balance},
            {"name": "collection_ratio", "value": collection_ratio},
            {"name": "provider_count", "value": int(snapshot.get("provider_count") or 0)},
            {"name": "period", "value": str(snapshot.get("period") or "unknown")},
        ]

    return [
        {"name": "production", "value": 10000},
        {"name": "collections", "value": 9500},
        {"name": "ar", "value": 1200},
        {"name": "trends", "value": "up"},
    ]

# Add more business logic functions as needed


# --- Script migration service stubs ---
def run_rebuild_receipt(output_path: str = "scripts/rebuild_receipt.json", skip_steps: bool = False):
    """Run the logic from write_rebuild_receipt.py and return the result as a dict."""
    import time
    from datetime import datetime, timezone
    from pathlib import Path

    def _run_command(command: list[str], cwd: Path) -> dict:
        started = time.perf_counter()
        import subprocess
        result = subprocess.run(
            command,
            cwd=str(cwd),
            capture_output=True,
            text=True,
            check=False,
        )
        duration = round(time.perf_counter() - started, 3)
        return {
            "command": command,
            "returncode": result.returncode,
            "duration_seconds": duration,
            "stdout_tail": "\n".join((result.stdout or "").splitlines()[-20:]),
            "stderr_tail": "\n".join((result.stderr or "").splitlines()[-20:]),
        }

    project_root = Path(__file__).resolve().parents[1]
    python_exe = Path(__import__('sys').executable)
    receipt_path = (project_root / output_path).resolve()
    receipt_path.parent.mkdir(parents=True, exist_ok=True)
    steps: list[dict] = []
    gates_report_path = project_root / "scripts" / "ci_gate_report.rebuild.json"
    if not skip_steps:
        refresh_cmd = [
            str(python_exe),
            str(project_root / "scripts" / "refresh_from_softdent_and_verify.py"),
        ]
        steps.append(_run_command(refresh_cmd, project_root))

        tests_cmd = [
            str(python_exe),
            "-m",
            "pytest",
            str(project_root / "app" / "tests"),
            "--rootdir",
            str(project_root),
            "-q",
        ]
        steps.append(_run_command(tests_cmd, project_root))

        gates_cmd = [
            str(python_exe),
            str(project_root / "scripts" / "run_ci_gates.py"),
            "--output",
            str(gates_report_path),
        ]
        steps.append(_run_command(gates_cmd, project_root))

    overall_pass = all(step.get("returncode", 1) == 0 for step in steps)
    payload = {
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "project_root": str(project_root),
        "skip_steps": bool(skip_steps),
        "overall_pass": overall_pass,
        "receipt_version": 1,
        "artifact_paths": {
            "rebuild_receipt": str(receipt_path),
            "ci_gate_report": str(gates_report_path),
        },
        "steps": steps,
    }
    receipt_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    return payload

def run_refresh_and_verify():
    """Run the logic from refresh_from_softdent_and_verify.py and return the result as a dict."""
    from fastapi.testclient import TestClient
    from .data_pipeline import get_pull_status_payload
    from .main import app
    # If recompute_cache is needed, import here or refactor as needed
    try:
        from .data_pipeline import recompute_cache
    except ImportError:
        def recompute_cache(app):
            pass  # fallback if not present

    def coerce_pull_status_sections(payload: object) -> tuple[object | None, object | None, dict[str, object]]:
        if not isinstance(payload, dict):
            return None, None, {}

        raw_status = payload.get("status")
        if isinstance(raw_status, dict):
            return payload.get("daily_refresh_enabled"), payload.get("last_refresh_date"), raw_status
        return payload.get("daily_refresh_enabled"), payload.get("last_refresh_date"), {}

    checks = []
    failures = []
    default_headers = {"host": "127.0.0.1"}
    page_paths = [
        "/",
        "/admin",
        "/softdent",
        "/quickbooks",
        "/accounts-receivable",
        "/reconciliation",
        "/trends",
        "/ebitda",
        "/claims",
        "/hal9000",
        "/reports",
    ]
    api_paths = [
        "/api/health",
        "/api/kpis",
        "/api/admin",
        "/api/reconciliation",
        "/api/hal9000/phases",
        "/api/reports/pull-status",
        "/api/reports/practice-central-delta",
    ]
    with _internal_service_auth_override(app), TestClient(app) as client:
        # Force immediate refresh from configured sources
        recompute_cache(app)
        for path in page_paths + api_paths:
            response = client.get(path, headers=default_headers)
            checks.append({"method": "GET", "path": path, "status": response.status_code})
            if response.status_code >= 400:
                failures.append({
                    "method": "GET",
                    "path": path,
                    "status": response.status_code,
                    "body": response.text[:300],
                })
                pull_status = get_pull_status_payload(app)
                kpis_payload = client.get("/api/kpis", headers=default_headers).json()
                claims_page = client.get("/claims", headers=default_headers)
    daily_refresh_enabled, last_refresh_date, status_sections = coerce_pull_status_sections(pull_status)

    summary = {
        "checked": checks,
        "failures": failures,
        "refresh": {
            "daily_refresh_enabled": daily_refresh_enabled,
            "last_refresh_date": last_refresh_date,
            "softdent_pull": status_sections.get("softdent", {}),
            "quickbooks_pull": status_sections.get("quickbooks", {}),
            "practice_central_pull": status_sections.get("practice_central", {}),
            "kpi_rows": len(kpis_payload.get("items") or []),
            "claims_page_status": claims_page.status_code,
        },
    }
    return summary

def run_ci_gates(output_path: str = "scripts/ci_gate_report.json", skip_gates: bool = False):
    """Run the logic from run_ci_gates.py and return the result as a dict."""
    from pathlib import Path
    from scripts.ci_gate_support import build_ci_gate_report

    project_root = Path(__file__).resolve().parents[1]
    return build_ci_gate_report(project_root, output_path, skip_gates=skip_gates)

def run_smoke_tests():
    """Run the logic from smoke_all_routes.py and return the result as a dict."""
    from fastapi.testclient import TestClient
    from .main import app
    checked = []
    failures = []
    default_headers = {"host": "127.0.0.1"}
    with _internal_service_auth_override(app), TestClient(app) as client:
        for route in app.routes:
            methods = getattr(route, "methods", set()) or set()
            path = getattr(route, "path", "")
            if "GET" not in methods:
                continue
            if not path or "{" in path:
                continue
            if path.startswith("/docs") or path.startswith("/redoc") or path.startswith("/openapi"):
                continue
            try:
                response = client.get(path, headers=default_headers)
                checked.append({"method": "GET", "path": path, "status": response.status_code})
                if response.status_code >= 400:
                    failures.append({
                        "method": "GET",
                        "path": path,
                        "status": response.status_code,
                        "body": response.text[:300],
                    })
            except Exception as exc:
                failures.append({"method": "GET", "path": path, "error": str(exc)})

        post_smoke_cases = [
            {
                "path": "/softdent/import",
                "kwargs": {
                    "files": {
                        "file": (
                            "smoke_softdent.csv",
                            b"Month,Metric,Amount\\n2026-01,Production,100\\n",
                            "text/csv",
                        )
                    }
                },
            },
            {
                "path": "/quickbooks/import",
                "kwargs": {
                    "files": {
                        "file": (
                            "smoke_quickbooks.csv",
                            b"Date,Account,Category,Amount\\n2026-01-01,Income,Income,100\\n",
                            "text/csv",
                        )
                    }
                },
            },
            {
                "path": "/hal9000",
                "kwargs": {"json": {"question": "status"}},
            },
        ]
        for case in post_smoke_cases:
            path = case["path"]
            try:
                response = client.post(path, follow_redirects=False, headers=default_headers, **case["kwargs"])
                checked.append({"method": "POST", "path": path, "status": response.status_code})
                if response.status_code >= 400:
                    failures.append({
                        "method": "POST",
                        "path": path,
                        "status": response.status_code,
                        "body": response.text[:300],
                    })
            except Exception as exc:
                failures.append({"method": "POST", "path": path, "error": str(exc)})
    return {"checked": checked, "failures": failures}


@contextmanager
def _internal_service_auth_override(app):
    previous_override = app.dependency_overrides.get(authenticate)

    def _resolve_role_for_request(path: str, method: str) -> tuple[str | None, ...]:
        normalized_path = path or "/"
        normalized_method = (method or "GET").upper()
        if normalized_path == "/metrics" or normalized_path.startswith("/api/admin"):
            return ("admin", "dashboard:read", "hal:operator", None)
        if normalized_path.endswith("/import") or normalized_method != "GET":
            return ("admin", "hal:operator", "dashboard:read", None)
        if normalized_path == "/hal9000" or normalized_path.startswith("/api/hal9000"):
            return ("hal:operator", "admin", "dashboard:read", None)
        return ("dashboard:read", "admin", "hal:operator", None)

    def _override_auth(request: Request) -> object:
        path = str(getattr(getattr(request, "url", None), "path", "") or "/")
        method = str(getattr(request, "method", "GET") or "GET")
        last_error: RuntimeError | None = None
        for role in _resolve_role_for_request(path, method):
            try:
                return get_service_user(role)
            except RuntimeError as exc:
                last_error = exc
                continue
        if last_error is not None:
            raise last_error
        raise RuntimeError("No configured service user is available for internal route validation")

    app.dependency_overrides[authenticate] = _override_auth
    try:
        yield
    finally:
        if previous_override is None:
            app.dependency_overrides.pop(authenticate, None)
        else:
            app.dependency_overrides[authenticate] = previous_override


@contextmanager
def _service_test_client(app, *, required_role: str | None = None):
    from fastapi.testclient import TestClient

    def _resolve_role(path: str, method: str) -> str | None:
        if required_role is not None:
            return required_role
        normalized_path = path or "/"
        normalized_method = (method or "GET").upper()
        if normalized_path == "/metrics" or normalized_path.startswith("/api/admin"):
            return "admin"
        if normalized_path.endswith("/import") or normalized_method != "GET":
            return "admin"
        if normalized_path == "/hal9000" or normalized_path.startswith("/api/hal9000"):
            return "hal:operator"
        return "dashboard:read"

    previous_override = app.dependency_overrides.get(authenticate)

    def _override_auth(request: Request) -> object:
        path = str(getattr(getattr(request, "url", None), "path", "") or "/")
        method = str(getattr(request, "method", "GET") or "GET")
        return get_service_user(_resolve_role(path, method))

    app.dependency_overrides[authenticate] = _override_auth
    try:
        with TestClient(app) as client:
            yield client
    finally:
        if previous_override is None:
            app.dependency_overrides.pop(authenticate, None)
        else:
            app.dependency_overrides[authenticate] = previous_override
