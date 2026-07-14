"""Deep SoftDent transaction + register extract (Moonshot full-extract pack).

Reads live ``transactions_for_period.jsonl`` / ``register_for_period.jsonl`` from
``C:\\SoftDentFinancialExports`` (normalized JSONL wrapper shape) into:

- ``sd_transactions_full`` — every line-item transaction
- ``sd_register_detail`` — register summary / method rows
- ``sd_operatory_schedule`` — chairs/slots from ``operatory_schedule.json``
- refreshes ``sd_payments`` / ``sd_adjustments`` from typed transaction rows

Also upserts into legacy ``transactions`` when that table already exists so the
analytics DB stays in parity with the external SoftDent financial importer.
"""

from __future__ import annotations

import hashlib
import json
import os
import re
import sqlite3
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

REPO_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_EXPORTS = Path(r"C:\SoftDentFinancialExports")


def _utc_now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def resolve_exports_dir() -> Path:
    configured = os.environ.get("SOFTDENT_FINANCIAL_EXPORTS", "").strip()
    if configured:
        return Path(configured).expanduser()
    return DEFAULT_EXPORTS


def resolve_analytics_db() -> Path | None:
    try:
        from quickbooks_monthly_sync import resolve_analytics_db as _resolve

        path = _resolve()
        if path and path.is_file():
            return path
    except Exception:
        pass
    env = os.environ.get("NR2_FINANCIAL_ANALYTICS_DB", "").strip()
    if env:
        candidate = Path(env).expanduser()
        if candidate.is_file():
            return candidate
    default = resolve_exports_dir() / "softdent_financial_analytics.db"
    return default if default.is_file() else None


def resolve_account_transactions_db(db_path: Path | str | None = None) -> Path | None:
    """Resolve DB holding sd_account_transactions / sd_claims (HAL-10580 alias)."""
    if db_path:
        target = Path(db_path)
        return target if target.is_file() else None
    return resolve_analytics_db()


def parse_money(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("$", "").replace(",", "").replace("(", "-").replace(")", "")
    if not text or text in {"-", "—"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _normalize_code(code: Any) -> str:
    text = str(code or "").strip()
    if not text:
        return ""
    # SoftDent sometimes emits "2.00" for code 2
    if text.replace(".", "", 1).isdigit() and "." in text:
        try:
            as_float = float(text)
            if as_float == int(as_float):
                return str(int(as_float))
        except ValueError:
            pass
    return text


def _stable_id(*parts: Any) -> str:
    payload = "|".join("" if p is None else str(p) for p in parts)
    return hashlib.sha256(payload.encode("utf-8", errors="replace")).hexdigest()[:32]


def ensure_transactions_schema(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS sd_transactions_full (
            transaction_id TEXT PRIMARY KEY,
            patient_id TEXT,
            patient_name TEXT,
            provider_code TEXT,
            service_date TEXT,
            entry_date TEXT,
            ada_code TEXT,
            description TEXT,
            amount REAL,
            transaction_type TEXT,
            payment_method TEXT,
            payment_amount REAL,
            adjustment_code TEXT,
            adjustment_amount REAL,
            payer TEXT,
            claim_id TEXT,
            tooth TEXT,
            surface TEXT,
            quantity INTEGER DEFAULT 1,
            original_transaction_id TEXT,
            account_id TEXT,
            guarantor_id TEXT,
            source_file TEXT,
            row_number INTEGER,
            extracted_at TEXT
        );
        CREATE INDEX IF NOT EXISTS idx_sd_trans_patient ON sd_transactions_full(patient_id);
        CREATE INDEX IF NOT EXISTS idx_sd_trans_date ON sd_transactions_full(service_date);
        CREATE INDEX IF NOT EXISTS idx_sd_trans_type ON sd_transactions_full(transaction_type);
        CREATE INDEX IF NOT EXISTS idx_sd_trans_code ON sd_transactions_full(ada_code);

        -- Moonshot account-tx Excel ledger (distinct from sd_transactions_full)
        CREATE TABLE IF NOT EXISTS sd_account_transactions (
            stable_id TEXT PRIMARY KEY,
            source_file TEXT NOT NULL,
            row_number INTEGER NOT NULL,
            account_num TEXT NOT NULL,
            patient_name TEXT,
            service_date TEXT,
            provider TEXT,
            procedure TEXT,
            note_flag TEXT,
            amount REAL,
            prod REAL,
            charges REAL,
            prod_adj REAL,
            cash REAL,
            "check" REAL,
            credit REAL,
            pay_adj REAL,
            period_start TEXT,
            period_end TEXT,
            extracted_at TEXT
        );
        CREATE INDEX IF NOT EXISTS idx_acctx_account ON sd_account_transactions(account_num);
        CREATE INDEX IF NOT EXISTS idx_acctx_date ON sd_account_transactions(service_date);
        CREATE INDEX IF NOT EXISTS idx_acctx_period ON sd_account_transactions(period_start, period_end);
        CREATE INDEX IF NOT EXISTS idx_acctx_source ON sd_account_transactions(source_file);

        CREATE TABLE IF NOT EXISTS sd_register_detail (
            register_id TEXT PRIMARY KEY,
            transaction_id TEXT,
            patient_id TEXT,
            payment_date TEXT,
            payment_method TEXT,
            payment_amount REAL,
            check_number TEXT,
            batch_number TEXT,
            label TEXT,
            deposited INTEGER DEFAULT 0,
            source_file TEXT,
            extracted_at TEXT
        );

        CREATE TABLE IF NOT EXISTS sd_operatory_schedule (
            appointment_id TEXT PRIMARY KEY,
            operatory TEXT,
            patient_id TEXT,
            patient_name TEXT,
            provider_code TEXT,
            appt_date TEXT,
            start_time TEXT,
            end_time TEXT,
            status TEXT,
            appointment_type TEXT,
            extracted_at TEXT
        );

        CREATE TABLE IF NOT EXISTS sd_extract_meta (
            key TEXT PRIMARY KEY,
            value TEXT
        );
        """
    )


def _unwrap_normalized(payload: dict[str, Any]) -> dict[str, Any]:
    norm = payload.get("normalized")
    if isinstance(norm, dict):
        return norm
    return payload


def normalize_transaction(
    payload: dict[str, Any],
    *,
    source_file: str,
    extracted_at: str,
) -> dict[str, Any] | None:
    """Map SoftDentFinancialExports JSONL row → sd_transactions_full."""
    if not isinstance(payload, dict):
        return None
    raw = _unwrap_normalized(payload)
    if raw.get("source_summary_type"):
        return None  # period totals header — not a line item

    row_number = payload.get("row_number")
    tx_date = str(raw.get("transaction_date") or raw.get("serviceDate") or raw.get("date") or "")[:10]
    post_date = str(raw.get("posting_date") or raw.get("entryDate") or "")[:10]
    service_date = str(raw.get("service_date") or tx_date or "")[:10]
    code = _normalize_code(raw.get("transaction_code") or raw.get("code") or raw.get("adaCode"))
    description = str(raw.get("transaction_description") or raw.get("description") or "")
    tx_type = str(raw.get("transaction_type") or raw.get("type") or "transaction").strip().lower()
    amount = parse_money(raw.get("amount") or raw.get("production") or raw.get("chargeAmount"))
    patient_id = str(raw.get("patient_id") or raw.get("patientId") or "")
    provider_code = str(raw.get("provider_id") or raw.get("provider_code") or raw.get("providerId") or "")
    account_id = str(raw.get("account_id") or "")
    guarantor_id = str(raw.get("guarantor_id") or "")

    explicit_id = str(raw.get("transactionId") or raw.get("id") or "").strip()
    transaction_id = explicit_id or _stable_id(
        source_file,
        row_number,
        tx_date,
        post_date,
        code,
        amount,
        patient_id,
        provider_code,
        tx_type,
        description,
    )

    payment_amount = None
    adjustment_amount = None
    adjustment_code = ""
    if tx_type in {"payment", "credit", "refund"}:
        payment_amount = abs(amount) if amount is not None else None
    if tx_type in {"adjustment", "writeoff", "write-off"} or code in {"51", "52"}:
        adjustment_code = code
        adjustment_amount = abs(amount) if amount is not None else None

    return {
        "transaction_id": transaction_id,
        "patient_id": patient_id,
        "patient_name": str(raw.get("patient_name") or raw.get("patientName") or ""),
        "provider_code": provider_code,
        "service_date": service_date,
        "entry_date": post_date or tx_date,
        "ada_code": code,
        "description": description,
        "amount": amount,
        "transaction_type": tx_type or "transaction",
        "payment_method": str(raw.get("paymentMethod") or raw.get("payment_method") or ""),
        "payment_amount": payment_amount,
        "adjustment_code": adjustment_code,
        "adjustment_amount": adjustment_amount,
        "payer": str(raw.get("payer") or ""),
        "claim_id": str(raw.get("claimId") or raw.get("claim_id") or ""),
        "tooth": str(raw.get("toothNumber") or raw.get("tooth") or ""),
        "surface": str(raw.get("surface") or ""),
        "quantity": int(raw.get("unitQuantity") or raw.get("quantity") or 1),
        "original_transaction_id": str(raw.get("originalTransactionId") or ""),
        "account_id": account_id,
        "guarantor_id": guarantor_id,
        "source_file": str(payload.get("source_file") or source_file),
        "row_number": int(row_number) if str(row_number or "").isdigit() else None,
        "extracted_at": extracted_at,
        # legacy transactions table helpers
        "_provider_name": str(raw.get("provider_name") or ""),
        "_collecting_provider_id": str(raw.get("collecting_provider_id") or ""),
        "_transaction_date": tx_date,
        "_posting_date": post_date,
    }


def load_transactions_jsonl(path: Path) -> list[dict[str, Any]]:
    transactions: list[dict[str, Any]] = []
    if not path.is_file():
        return transactions
    extracted_at = _utc_now()
    with path.open("r", encoding="utf-8", errors="replace") as handle:
        for line in handle:
            line = line.strip()
            if not line:
                continue
            try:
                payload = json.loads(line)
            except json.JSONDecodeError:
                continue
            items = payload if isinstance(payload, list) else [payload]
            for item in items:
                if not isinstance(item, dict):
                    continue
                details = item.get("transactionDetails") or item.get("details") or item.get("transactions")
                if isinstance(details, list):
                    for detail in details:
                        if isinstance(detail, dict):
                            wrapped = {"normalized": detail, "source_file": item.get("source_file"), "row_number": item.get("row_number")}
                            tx = normalize_transaction(wrapped, source_file=path.name, extracted_at=extracted_at)
                            if tx:
                                transactions.append(tx)
                    continue
                tx = normalize_transaction(item, source_file=path.name, extracted_at=extracted_at)
                if tx:
                    transactions.append(tx)
    return transactions


def load_register_jsonl(path: Path) -> list[dict[str, Any]]:
    registers: list[dict[str, Any]] = []
    if not path.is_file():
        return registers
    extracted_at = _utc_now()
    with path.open("r", encoding="utf-8", errors="replace") as handle:
        for line in handle:
            line = line.strip()
            if not line:
                continue
            try:
                payload = json.loads(line)
            except json.JSONDecodeError:
                continue
            if not isinstance(payload, dict):
                continue
            if str(payload.get("dataset_name") or "") not in {"", "register_for_period"}:
                # still accept unlabeled rows
                if payload.get("dataset_name") and payload.get("dataset_name") != "register_for_period":
                    continue
            raw_row = payload.get("raw_row")
            norm = payload.get("normalized") if isinstance(payload.get("normalized"), dict) else {}
            cells = [str(c or "").strip() for c in raw_row] if isinstance(raw_row, list) else []
            while len(cells) < 4:
                cells.append("")
            label = (cells[0] or cells[1] or str(norm.get("adjustments_to_prod") or "")).strip()
            method = cells[1].strip() if cells[0] else ""
            amount = parse_money(cells[2]) or parse_money(cells[3])
            if amount is None:
                # try odd normalized money keys
                for key, value in norm.items():
                    if key in {"adjustments_to_prod", "calendar_year", "calendar_month", "year_month"}:
                        continue
                    amount = parse_money(value)
                    if amount is not None:
                        break
            if not label and amount is None:
                continue
            row_number = payload.get("row_number")
            register_id = _stable_id(path.name, row_number, label, method, amount)
            registers.append(
                {
                    "register_id": register_id,
                    "transaction_id": "",
                    "patient_id": "",
                    "payment_date": "",
                    "payment_method": method or label,
                    "payment_amount": amount,
                    "check_number": "",
                    "batch_number": "",
                    "label": label,
                    "deposited": 0,
                    "source_file": str(payload.get("source_file") or path.name),
                    "extracted_at": extracted_at,
                }
            )
    return registers


def persist_transactions(conn: sqlite3.Connection, transactions: list[dict[str, Any]]) -> int:
    if not transactions:
        return 0
    rows = [
        {
            k: v
            for k, v in tx.items()
            if not str(k).startswith("_")
        }
        for tx in transactions
    ]
    conn.executemany(
        """
        INSERT INTO sd_transactions_full (
            transaction_id, patient_id, patient_name, provider_code, service_date, entry_date,
            ada_code, description, amount, transaction_type, payment_method, payment_amount,
            adjustment_code, adjustment_amount, payer, claim_id, tooth, surface, quantity,
            original_transaction_id, account_id, guarantor_id, source_file, row_number, extracted_at
        ) VALUES (
            :transaction_id, :patient_id, :patient_name, :provider_code, :service_date, :entry_date,
            :ada_code, :description, :amount, :transaction_type, :payment_method, :payment_amount,
            :adjustment_code, :adjustment_amount, :payer, :claim_id, :tooth, :surface, :quantity,
            :original_transaction_id, :account_id, :guarantor_id, :source_file, :row_number, :extracted_at
        )
        ON CONFLICT(transaction_id) DO UPDATE SET
            patient_id=excluded.patient_id,
            patient_name=excluded.patient_name,
            provider_code=excluded.provider_code,
            service_date=excluded.service_date,
            entry_date=excluded.entry_date,
            ada_code=excluded.ada_code,
            description=excluded.description,
            amount=excluded.amount,
            transaction_type=excluded.transaction_type,
            payment_method=excluded.payment_method,
            payment_amount=excluded.payment_amount,
            adjustment_code=excluded.adjustment_code,
            adjustment_amount=excluded.adjustment_amount,
            payer=excluded.payer,
            account_id=excluded.account_id,
            guarantor_id=excluded.guarantor_id,
            extracted_at=excluded.extracted_at
        """,
        rows,
    )
    # Do not upsert into legacy `transactions` — that table is owned by the
    # SoftDent financial importer (different business_key scheme). NR2 depth
    # lives in sd_transactions_full; payments/adjustments are refreshed below.
    _upsert_payments_adjustments_from_transactions(conn, transactions)
    conn.commit()
    return len(rows)


def _table_exists(conn: sqlite3.Connection, name: str) -> bool:
    cur = conn.execute(
        "SELECT 1 FROM sqlite_master WHERE type='table' AND name=? LIMIT 1",
        (name,),
    )
    return cur.fetchone() is not None


def _ensure_sd_pay_adj(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS sd_payments (
            practice_id TEXT NOT NULL DEFAULT '',
            patient_id TEXT NOT NULL DEFAULT '',
            payment_date TEXT NOT NULL DEFAULT '',
            amount REAL,
            payer TEXT,
            method TEXT,
            extracted_at TEXT,
            PRIMARY KEY (practice_id, patient_id, payment_date, amount, method)
        );
        CREATE TABLE IF NOT EXISTS sd_adjustments (
            practice_id TEXT NOT NULL DEFAULT '',
            patient_id TEXT NOT NULL DEFAULT '',
            adj_date TEXT NOT NULL DEFAULT '',
            ada_code TEXT NOT NULL DEFAULT '',
            amount REAL,
            description TEXT,
            extracted_at TEXT,
            PRIMARY KEY (practice_id, patient_id, adj_date, ada_code, amount)
        );
        """
    )


def _upsert_payments_adjustments_from_transactions(
    conn: sqlite3.Connection,
    transactions: list[dict[str, Any]],
) -> dict[str, int]:
    _ensure_sd_pay_adj(conn)
    counts = {"sd_payments": 0, "sd_adjustments": 0}
    for tx in transactions:
        tx_type = str(tx.get("transaction_type") or "").lower()
        patient_id = str(tx.get("patient_id") or "").strip() or f"tx:{tx.get('transaction_id')}"
        date = str(tx.get("service_date") or tx.get("entry_date") or "").strip()
        amount = tx.get("amount")
        if amount is None:
            continue
        if tx_type == "payment" or (tx.get("payment_amount") is not None and tx_type in {"payment", "credit", "refund"}):
            if not date:
                continue
            conn.execute(
                """
                INSERT OR REPLACE INTO sd_payments
                (practice_id, patient_id, payment_date, amount, payer, method, extracted_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    "",
                    patient_id,
                    date,
                    abs(float(amount)),
                    tx.get("payer") or "",
                    tx.get("payment_method") or tx.get("description") or tx.get("ada_code") or "payment",
                    tx.get("extracted_at"),
                ),
            )
            counts["sd_payments"] += 1
        elif tx_type in {"adjustment", "writeoff", "write-off"} or str(tx.get("adjustment_code") or "") in {"51", "52"}:
            if not date:
                continue
            conn.execute(
                """
                INSERT OR REPLACE INTO sd_adjustments
                (practice_id, patient_id, adj_date, ada_code, amount, description, extracted_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    "",
                    patient_id,
                    date,
                    tx.get("adjustment_code") or tx.get("ada_code") or "adj",
                    abs(float(amount)),
                    tx.get("description") or "adjustment",
                    tx.get("extracted_at"),
                ),
            )
            counts["sd_adjustments"] += 1
    return counts


def persist_register(conn: sqlite3.Connection, registers: list[dict[str, Any]]) -> int:
    if not registers:
        return 0
    conn.executemany(
        """
        INSERT INTO sd_register_detail (
            register_id, transaction_id, patient_id, payment_date, payment_method,
            payment_amount, check_number, batch_number, label, deposited, source_file, extracted_at
        ) VALUES (
            :register_id, :transaction_id, :patient_id, :payment_date, :payment_method,
            :payment_amount, :check_number, :batch_number, :label, :deposited, :source_file, :extracted_at
        )
        ON CONFLICT(register_id) DO UPDATE SET
            payment_amount=excluded.payment_amount,
            payment_method=excluded.payment_method,
            label=excluded.label,
            extracted_at=excluded.extracted_at
        """,
        registers,
    )
    conn.commit()
    return len(registers)


def load_operatory_schedule(path: Path | None = None) -> list[dict[str, Any]]:
    """Parse operatory_schedule.json (operatoryChairs[] NR2 shape or operatories[])."""
    if path is None:
        exports = resolve_exports_dir()
        candidates = [
            exports / "operatory_schedule.json",
        ]
        try:
            from import_loader import softdent_import_dir

            candidates.insert(0, softdent_import_dir() / "operatory_schedule.json")
        except Exception:
            pass
        path = next((p for p in candidates if p.is_file()), None)
    if not path or not path.is_file():
        return []
    try:
        data = json.loads(path.read_text(encoding="utf-8-sig"))
    except (OSError, json.JSONDecodeError):
        return []
    if not isinstance(data, dict):
        return []
    extracted_at = _utc_now()
    slots: list[dict[str, Any]] = []

    chairs = data.get("operatoryChairs")
    if isinstance(chairs, list):
        for chair in chairs:
            if not isinstance(chair, dict):
                continue
            op_name = str(chair.get("name") or chair.get("operatory") or "")
            for idx, appt in enumerate(chair.get("slots") or []):
                if not isinstance(appt, dict):
                    continue
                time_label = str(appt.get("time") or "")
                patient = str(appt.get("patient") or "")
                status = str(appt.get("procedure") or appt.get("status") or "")
                appt_id = _stable_id(op_name, time_label, patient, idx, status)
                slots.append(
                    {
                        "appointment_id": appt_id,
                        "operatory": op_name,
                        "patient_id": str(appt.get("patientId") or ""),
                        "patient_name": patient,
                        "provider_code": str(appt.get("provider") or appt.get("providerCode") or ""),
                        "appt_date": time_label[:10],
                        "start_time": time_label,
                        "end_time": "",
                        "status": status,
                        "appointment_type": str(appt.get("type") or ""),
                        "extracted_at": extracted_at,
                    }
                )
        return slots

    for op in data.get("operatories") or []:
        if not isinstance(op, dict):
            continue
        op_name = str(op.get("operatoryName") or op.get("name") or "")
        for appt in op.get("appointments") or []:
            if not isinstance(appt, dict):
                continue
            appt_id = str(appt.get("id") or _stable_id(op_name, appt.get("patientId"), appt.get("startTime")))
            slots.append(
                {
                    "appointment_id": appt_id,
                    "operatory": op_name,
                    "patient_id": str(appt.get("patientId") or ""),
                    "patient_name": str(appt.get("patientName") or ""),
                    "provider_code": str(appt.get("providerCode") or ""),
                    "appt_date": str(appt.get("date") or "")[:10],
                    "start_time": str(appt.get("startTime") or ""),
                    "end_time": str(appt.get("endTime") or ""),
                    "status": str(appt.get("status") or ""),
                    "appointment_type": str(appt.get("type") or ""),
                    "extracted_at": extracted_at,
                }
            )
    return slots


def persist_operatory_schedule(conn: sqlite3.Connection, slots: list[dict[str, Any]]) -> int:
    if not slots:
        return 0
    conn.executemany(
        """
        INSERT INTO sd_operatory_schedule (
            appointment_id, operatory, patient_id, patient_name, provider_code,
            appt_date, start_time, end_time, status, appointment_type, extracted_at
        ) VALUES (
            :appointment_id, :operatory, :patient_id, :patient_name, :provider_code,
            :appt_date, :start_time, :end_time, :status, :appointment_type, :extracted_at
        )
        ON CONFLICT(appointment_id) DO UPDATE SET
            operatory=excluded.operatory,
            patient_name=excluded.patient_name,
            status=excluded.status,
            extracted_at=excluded.extracted_at
        """,
        slots,
    )
    conn.commit()
    return len(slots)


def verify_completeness(conn: sqlite3.Connection, trans_path: Path) -> dict[str, Any]:
    stats: dict[str, Any] = {
        "file_bytes": trans_path.stat().st_size if trans_path.is_file() else 0,
        "jsonl_rows": 0,
        "db_rows": 0,
        "date_range": {"min": None, "max": None},
        "amount_sum": 0.0,
        "by_type": {},
        "parity_ratio": None,
    }
    if trans_path.is_file():
        with trans_path.open("r", encoding="utf-8", errors="replace") as handle:
            stats["jsonl_rows"] = sum(1 for line in handle if line.strip())
    cur = conn.execute(
        "SELECT COUNT(*), MIN(service_date), MAX(service_date), SUM(amount) FROM sd_transactions_full"
    )
    row = cur.fetchone()
    if row:
        stats["db_rows"] = int(row[0] or 0)
        stats["date_range"]["min"] = row[1]
        stats["date_range"]["max"] = row[2]
        stats["amount_sum"] = float(row[3] or 0.0)
    for tx_type, count in conn.execute(
        "SELECT transaction_type, COUNT(*) FROM sd_transactions_full GROUP BY transaction_type"
    ):
        stats["by_type"][str(tx_type)] = int(count)
    # Exclude 1 summary header line when present
    expected = max(0, int(stats["jsonl_rows"]) - 1)
    if expected > 0:
        stats["parity_ratio"] = round(stats["db_rows"] / expected, 4)
    return stats


def extract_all_transactions(
    db_path: Path | None = None,
    *,
    force: bool = False,
) -> dict[str, Any]:
    """Main entry — full transaction / register / operatory extract."""
    del force  # always re-read source files; upserts are idempotent
    result: dict[str, Any] = {
        "ok": False,
        "transactions": 0,
        "register": 0,
        "operatory": 0,
        "warnings": [],
        "mode": "jsonl-financial-exports",
    }
    db_path = db_path or resolve_analytics_db()
    if not db_path:
        result["warnings"].append("No analytics DB found")
        return result

    exports = resolve_exports_dir()
    trans_path = exports / "transactions_for_period.jsonl"
    register_path = exports / "register_for_period.jsonl"

    conn = sqlite3.connect(str(db_path))
    try:
        ensure_transactions_schema(conn)
        if trans_path.is_file():
            txs = load_transactions_jsonl(trans_path)
            result["transactions"] = persist_transactions(conn, txs)
        else:
            result["warnings"].append(f"Missing {trans_path}")

        if register_path.is_file():
            regs = load_register_jsonl(register_path)
            result["register"] = persist_register(conn, regs)
        else:
            result["warnings"].append(f"Missing {register_path}")

        slots = load_operatory_schedule()
        if slots:
            result["operatory"] = persist_operatory_schedule(conn, slots)

        result["verification"] = verify_completeness(conn, trans_path)
        conn.execute(
            "INSERT OR REPLACE INTO sd_extract_meta (key, value) VALUES (?, ?)",
            ("last_transaction_extract", _utc_now()),
        )
        conn.commit()
        parity = (result.get("verification") or {}).get("parity_ratio")
        result["ok"] = bool(result["transactions"] > 0 and (parity is None or parity >= 0.9))
        if parity is not None and parity < 0.9:
            result["warnings"].append(
                f"Transaction parity {parity} below 0.9 (jsonl vs sd_transactions_full)"
            )
    finally:
        conn.close()
    return result


def sync_softdent_transactions() -> dict[str, Any]:
    """Import-sync hook."""
    return extract_all_transactions(force=True)


# ---------------------------------------------------------------------------
# SoftDent Trans-for-a-Period Excel (TXN*.xls) — read-only parse / HAL ledger
# Moonshot AFTER_ACCOUNT_TX_EXCEL (2026-07-12). empty ≠ $0; no SoftDent write-back.
# ---------------------------------------------------------------------------

DEFAULT_TXN_INBOX = Path(r"C:\SoftDentReportExports")
DEFAULT_TX_PARSED_DIR = DEFAULT_EXPORTS / "tx_parsed"
TXN_HEADER_MARKERS = {"date", "id", "name"}


def resolve_tx_parsed_dir() -> Path:
    configured = os.environ.get("SOFTDENT_TX_PARSED_DIR", "").strip()
    if configured:
        return Path(configured).expanduser()
    return DEFAULT_TX_PARSED_DIR


def _load_account_tx_csv_rows(path: Path) -> list[list[Any]]:
    """Load SoftDent TXN CSV (often saved as ``.xls``) as row lists."""
    import csv

    text = path.read_text(encoding="latin-1", errors="ignore")
    return [list(row) for row in csv.reader(text.splitlines())]


def _load_account_tx_excel_rows(path: Path) -> list[list[Any]]:
    """Load SoftDent Trans-for-a-Period .xls/.xlsx/CSV as row lists (no PHI logging).

    SoftDent often writes CSV bytes under a ``.xls`` name; detect OLE magic first.
    """
    suffix = path.suffix.lower()
    head = b""
    try:
        with path.open("rb") as handle:
            head = handle.read(8)
    except OSError:
        head = b""
    is_ole = head[:4] == b"\xd0\xcf\x11\xe0"
    if suffix == ".csv" or (suffix == ".xls" and not is_ole and head[:4] in {b"TRAN", b"Date", b'"Dat'}):
        return _load_account_tx_csv_rows(path)
    if suffix == ".xls" or is_ole:
        try:
            import xlrd  # type: ignore
        except ImportError as exc:  # pragma: no cover
            raise RuntimeError("xlrd required to parse SoftDent TXN .xls exports") from exc
        try:
            book = xlrd.open_workbook(str(path))
        except Exception:
            # Misnamed CSV or truncated OLE — fall back to CSV text parse
            return _load_account_tx_csv_rows(path)
        sheet = book.sheet_by_index(0)
        return [[sheet.cell_value(r, c) for c in range(sheet.ncols)] for r in range(sheet.nrows)]
    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook  # type: ignore
        except ImportError as exc:  # pragma: no cover
            raise RuntimeError("openpyxl required to parse SoftDent TXN .xlsx exports") from exc
        book = load_workbook(str(path), data_only=True, read_only=True)
        try:
            sheet = book.active
            return [list(row) for row in sheet.iter_rows(values_only=True)]
        finally:
            book.close()
    # Bare SoftDent export without suffix — try CSV then fail
    if head and not is_ole:
        return _load_account_tx_csv_rows(path)
    raise ValueError(f"unsupported excel suffix: {suffix}")


def _excel_serial_to_iso(value: Any) -> str | None:
    """Convert SoftDent Excel date cell to YYYY-MM-DD; never invent a date."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%Y/%m/%d"):
            try:
                return datetime.strptime(text[:10], fmt).date().isoformat()
            except ValueError:
                continue
        return None
    if isinstance(value, (int, float)):
        serial = float(value)
        if serial < 20000 or serial > 80000:
            return None
        try:
            import xlrd  # type: ignore

            dt = xlrd.xldate_as_datetime(serial, 0)
            return dt.date().isoformat()
        except Exception:
            try:
                # Excel 1900-epoch fallback (SoftDent .xls)
                from datetime import date, timedelta

                base = date(1899, 12, 30)
                return (base + timedelta(days=int(serial))).isoformat()
            except Exception:
                return None
    return None


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value).strip()


def _account_num_str(value: Any) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    text = str(value).strip()
    if not text:
        return None
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def _first_money(*values: Any) -> float | None:
    """First present money cell; empty stays None (never coerce to 0)."""
    for value in values:
        if value is None or value == "":
            continue
        if isinstance(value, str) and not value.strip():
            continue
        parsed = parse_money(value)
        if parsed is not None:
            return parsed
    return None


def _is_header_row(row: list[Any]) -> bool:
    labels = {_cell_str(c).lower() for c in (row or [])[:6]}
    return TXN_HEADER_MARKERS.issubset(labels)


def _is_data_row(row: list[Any]) -> bool:
    if not row or len(row) < 3:
        return False
    if _is_header_row(row):
        return False
    date_iso = _excel_serial_to_iso(row[0] if row else None)
    name = _cell_str(row[2] if len(row) > 2 else None)
    account = _account_num_str(row[1] if len(row) > 1 else None)
    return bool(date_iso and (name or account))


def _row_to_account_tx_record(row: list[Any], *, row_number: int, source_file: str) -> dict[str, Any]:
    # SoftDent header: Date | ID | Name | D$ | Dr | Code | (flag) | Prod | * | Charges |
    # Prod Adj | Cash | Check | Credit | Pay Adj
    cells = list(row or [])
    while len(cells) < 15:
        cells.append("")
    amount = _first_money(
        cells[7],  # Prod
        cells[9],  # Charges
        cells[10],  # Prod Adj
        cells[11],  # Cash
        cells[12],  # Check
        cells[13],  # Credit
        cells[14],  # Pay Adj
    )
    note_raw = cells[6]
    note_flag = _cell_str(note_raw) or None
    procedure = _normalize_code(cells[5]) or None
    provider = _cell_str(cells[4]) or None
    return {
        "date": _excel_serial_to_iso(cells[0]),
        "account_num": _account_num_str(cells[1]),
        "patient_name": _cell_str(cells[2]) or None,
        "provider": provider,
        "procedure": procedure,
        "amount": amount,
        "note_flag": note_flag,
        "row_number": row_number,
        "source_file": source_file,
        # Extra typed money legs (still empty≠$0) for HAL detail without inventing totals
        "prod": parse_money(cells[7]),
        "charges": parse_money(cells[9]),
        "prod_adj": parse_money(cells[10]),
        "cash": parse_money(cells[11]),
        "check": parse_money(cells[12]),
        "credit": parse_money(cells[13]),
        "pay_adj": parse_money(cells[14]),
    }


def parse_account_transactions_xls(path: Path | str) -> dict[str, Any]:
    """Parse SoftDent Trans-for-a-Period Excel into typed account-tx records.

    Returns meta + records. ``rowCount`` is the sheet row count (matches TXN*.xls nrows).
    ``amount`` is null when all money columns are empty — never invented $0.
    """
    target = Path(path)
    result: dict[str, Any] = {
        "ok": False,
        "path": str(target),
        "rowCount": 0,
        "recordCount": 0,
        "periodHint": None,
        "records": [],
        "warnings": [],
    }
    if not target.is_file():
        result["warnings"].append("file missing")
        return result
    try:
        rows = _load_account_tx_excel_rows(target)
    except Exception as exc:  # noqa: BLE001
        result["warnings"].append(f"excel_open_{type(exc).__name__}")
        return result

    result["rowCount"] = len(rows)
    period_hint = None
    for raw in rows[:20]:
        joined = " ".join(_cell_str(c) for c in (raw or []) if c not in ("", None))
        if not joined:
            continue
        # e.g. "02/01/26 TO 02/28/26"
        m = re.search(
            r"\b(\d{1,2})/(\d{1,2})/(\d{2,4})\s*TO\s*(\d{1,2})/(\d{1,2})/(\d{2,4})\b",
            joined,
            re.I,
        )
        if m:
            def _y(yy: str) -> int:
                n = int(yy)
                if n < 100:
                    return 2000 + n if n < 80 else 1900 + n
                return n

            period_hint = (
                f"{_y(m.group(3)):04d}-{int(m.group(1)):02d}-01:"
                f"{_y(m.group(6)):04d}-{int(m.group(4)):02d}-{int(m.group(5)):02d}"
            )
            break
    result["periodHint"] = period_hint

    records: list[dict[str, Any]] = []
    for idx, raw in enumerate(rows):
        if not _is_data_row(raw):
            continue
        records.append(
            _row_to_account_tx_record(raw, row_number=idx, source_file=target.name)
        )
    result["records"] = records
    result["recordCount"] = len(records)
    result["ok"] = bool(records)
    return result


def write_account_transactions_jsonl(
    parsed: dict[str, Any],
    *,
    out_dir: Path | None = None,
    source_path: Path | str | None = None,
) -> dict[str, Any]:
    """Emit one JSONL file under SoftDentFinancialExports/tx_parsed/."""
    dest_dir = out_dir or resolve_tx_parsed_dir()
    dest_dir.mkdir(parents=True, exist_ok=True)
    src = Path(source_path or parsed.get("path") or "TXN.xls")
    out_path = dest_dir / f"{src.stem}.jsonl"
    records = list(parsed.get("records") or [])
    with out_path.open("w", encoding="utf-8", newline="\n") as handle:
        meta = {
            "_meta": True,
            "sourcePath": str(src),
            "rowCount": parsed.get("rowCount"),
            "recordCount": parsed.get("recordCount"),
            "periodHint": parsed.get("periodHint"),
            "extractedAt": _utc_now(),
            "honesty": "empty != $0; read-only SoftDent Excel parse",
        }
        handle.write(json.dumps(meta, ensure_ascii=False) + "\n")
        for rec in records:
            handle.write(json.dumps(rec, ensure_ascii=False, default=str) + "\n")
    return {
        "ok": True,
        "path": str(out_path),
        "rowCount": parsed.get("rowCount"),
        "recordCount": len(records),
    }


def ingest_account_transactions_xls(
    path: Path | str | None = None,
    *,
    out_dir: Path | None = None,
    db_path: Path | None = None,
) -> dict[str, Any]:
    """Parse TXN*.xls, write JSONL under tx_parsed/, upsert into analytics DB."""
    target = Path(path) if path else DEFAULT_TXN_INBOX / "TXN260201.xls"
    parsed = parse_account_transactions_xls(target)
    if not parsed.get("ok"):
        return {
            "ok": False,
            "path": str(target),
            "warnings": parsed.get("warnings") or ["parse failed"],
            "rowCount": parsed.get("rowCount"),
            "recordCount": parsed.get("recordCount"),
        }
    written = write_account_transactions_jsonl(parsed, out_dir=out_dir, source_path=target)
    db_result = upsert_account_transactions_jsonl(written["path"], db_path=db_path)
    return {
        "ok": bool(written.get("ok") and db_result.get("ok")),
        "sourcePath": str(target),
        "jsonlPath": written["path"],
        "rowCount": parsed.get("rowCount"),
        "recordCount": parsed.get("recordCount"),
        "periodHint": parsed.get("periodHint"),
        "db": db_result,
        "donnaNickelLines": sum(
            1
            for r in (parsed.get("records") or [])
            if str(r.get("account_num") or "") == "27002"
            or "nickel, donna" in str(r.get("patient_name") or "").lower()
        ),
        "nickelMentions": sum(
            1
            for r in (parsed.get("records") or [])
            if "nickel" in str(r.get("patient_name") or "").lower()
        ),
    }


YEAR_CHUNK_MANIFEST = DEFAULT_EXPORTS / "softdent_account_tx_year_chunks.json"
YEAR_CHUNK_INGEST_LOG = DEFAULT_EXPORTS / "softdent_account_tx_year_chunks_ingest.json"
# Sample Feb export superseded by TXN2026YTD once year chunks are loaded.
YEAR_CHUNK_SUPERSEDED_SOURCES = ("TXN260201.xls", "TXN260201.XLS")


def resolve_txn_export_path(stem: str, inbox: Path | None = None) -> Path | None:
    """Resolve SoftDent TXN export path for a stem (Windows case-insensitive)."""
    root = inbox or DEFAULT_TXN_INBOX
    for name in (f"{stem}.xls", f"{stem}.XLS", f"{stem}.csv", f"{stem}.CSV"):
        candidate = root / name
        if candidate.is_file():
            return candidate
    matches = sorted(
        [p for p in root.glob(f"{stem}.*") if p.suffix.lower() in {".xls", ".csv"}],
        key=lambda p: (0 if p.suffix.lower() == ".xls" else 1, -p.stat().st_size),
    )
    return matches[0] if matches else None


def _purge_account_tx_sources(db_path: Path, source_files: tuple[str, ...]) -> int:
    if not source_files or not db_path.is_file():
        return 0
    conn = sqlite3.connect(str(db_path))
    try:
        ensure_account_transactions_schema(conn)
        deleted = 0
        for name in source_files:
            cur = conn.execute(
                "DELETE FROM sd_account_transactions WHERE source_file = ?",
                (name,),
            )
            deleted += int(cur.rowcount or 0)
        conn.commit()
        return deleted
    finally:
        conn.close()


def ingest_account_transactions_year_chunks(
    *,
    inbox: Path | None = None,
    out_dir: Path | None = None,
    db_path: Path | None = None,
    manifest_path: Path | None = None,
    include_txnall: bool = True,
) -> dict[str, Any]:
    """Ingest verified year-chunk TX Excel/CSV into JSONL + sd_account_transactions.

    Uses ``softdent_account_tx_year_chunks.json`` as validation manifest (not business
    truth). Idempotent via purge-by-source_file upsert. Purges TXN260201 after
    TXN2026YTD so Feb sample does not duplicate YTD rows. empty ≠ $0.
    """
    root = inbox or DEFAULT_TXN_INBOX
    manifest = Path(manifest_path) if manifest_path else YEAR_CHUNK_MANIFEST
    target_db = Path(db_path) if db_path else resolve_analytics_db()
    if not target_db:
        target_db = resolve_exports_dir() / "softdent_financial_analytics.db"

    expected_by_stem: dict[str, int] = {}
    if manifest.is_file():
        try:
            raw = json.loads(manifest.read_text(encoding="utf-8"))
            for chunk in raw.get("chunks") or []:
                stem = str(chunk.get("stem") or "")
                rows = int(chunk.get("rows") or 0)
                if stem and rows > 0:
                    expected_by_stem[stem] = rows
        except (OSError, json.JSONDecodeError, TypeError, ValueError):
            expected_by_stem = {}

    stems: list[str] = []
    if include_txnall:
        stems.append("TXNALL260712")
    default_stems = [
        "TXN2017H2",
        "TXN2018",
        "TXN2019",
        "TXN2020",
        "TXN2021",
        "TXN2022",
        "TXN2023",
        "TXN2024",
        "TXN2025",
        "TXN2026YTD",
    ]
    for stem in default_stems:
        if stem not in stems:
            stems.append(stem)

    results: list[dict[str, Any]] = []
    for stem in stems:
        path = resolve_txn_export_path(stem, root)
        if path is None:
            results.append(
                {
                    "ok": False,
                    "stem": stem,
                    "error": "file missing",
                    "expectedRows": expected_by_stem.get(stem),
                }
            )
            continue
        ingest = ingest_account_transactions_xls(path, out_dir=out_dir, db_path=target_db)
        expected = expected_by_stem.get(stem)
        record_count = int(ingest.get("recordCount") or 0)
        parity_ok = True
        warnings = list((ingest.get("db") or {}).get("warnings") or [])
        if expected is not None:
            # Allow small parse variance (header/blank rows) but flag large gaps.
            if abs(record_count - expected) > max(25, int(expected * 0.02)):
                parity_ok = False
                warnings.append(f"manifest_rows={expected} ingest_rows={record_count}")
        results.append(
            {
                "ok": bool(ingest.get("ok") and parity_ok),
                "stem": stem,
                "sourcePath": ingest.get("sourcePath") or str(path),
                "jsonlPath": ingest.get("jsonlPath"),
                "recordCount": record_count,
                "expectedRows": expected,
                "periodHint": ingest.get("periodHint"),
                "dbCount": (ingest.get("db") or {}).get("dbCount"),
                "nullAmountCount": (ingest.get("db") or {}).get("nullAmountCount"),
                "warnings": warnings,
            }
        )

    purged = 0
    if any(r.get("ok") and r.get("stem") == "TXN2026YTD" for r in results):
        purged = _purge_account_tx_sources(target_db, YEAR_CHUNK_SUPERSEDED_SOURCES)

    conn = sqlite3.connect(str(target_db))
    try:
        ensure_account_transactions_schema(conn)
        total = int(conn.execute("SELECT COUNT(*) FROM sd_account_transactions").fetchone()[0])
        by_source = dict(
            conn.execute(
                "SELECT source_file, COUNT(*) FROM sd_account_transactions GROUP BY 1 ORDER BY 1"
            ).fetchall()
        )
        year_min = conn.execute(
            "SELECT MIN(substr(service_date,1,4)) FROM sd_account_transactions "
            "WHERE service_date IS NOT NULL AND length(service_date) >= 4"
        ).fetchone()[0]
        year_max = conn.execute(
            "SELECT MAX(substr(service_date,1,4)) FROM sd_account_transactions "
            "WHERE service_date IS NOT NULL AND length(service_date) >= 4"
        ).fetchone()[0]
    finally:
        conn.close()

    ok = all(r.get("ok") for r in results) and total > 100_000
    summary = {
        "ok": ok,
        "at": _utc_now(),
        "mode": "ingest-year-chunks",
        "honesty": "empty != $0; read-only SoftDent Excel/CSV ingest; no write-back",
        "manifestPath": str(manifest) if manifest.is_file() else None,
        "dbPath": str(target_db),
        "chunkCount": len(results),
        "okCount": sum(1 for r in results if r.get("ok")),
        "failCount": sum(1 for r in results if not r.get("ok")),
        "dbTotal": total,
        "dbBySource": by_source,
        "serviceYearMin": year_min,
        "serviceYearMax": year_max,
        "purgedSupersededRows": purged,
        "purgedSources": list(YEAR_CHUNK_SUPERSEDED_SOURCES) if purged else [],
        "account_tx_multi_year_available": bool(ok and year_min and year_max and int(year_min) <= 2017 and int(year_max) >= 2026),
        "chunks": results,
    }
    YEAR_CHUNK_INGEST_LOG.parent.mkdir(parents=True, exist_ok=True)
    YEAR_CHUNK_INGEST_LOG.write_text(json.dumps(summary, indent=2, default=str), encoding="utf-8")
    summary["ingestLogPath"] = str(YEAR_CHUNK_INGEST_LOG)
    return summary


def _split_period_hint(period_hint: Any) -> tuple[str | None, str | None]:
    text = str(period_hint or "").strip()
    if not text:
        return None, None
    if ":" in text:
        left, right = text.split(":", 1)
        return left.strip()[:10] or None, right.strip()[:10] or None
    return text[:10] or None, text[:10] or None


def _money_or_none(value: Any) -> float | None:
    """Preserve null honesty — never coerce empty to 0.0."""
    if value is None or value == "":
        return None
    if isinstance(value, str) and not value.strip():
        return None
    return parse_money(value)


def ensure_account_transactions_schema(conn: sqlite3.Connection) -> None:
    """Ensure sd_account_transactions exists (Moonshot account-tx DB design)."""
    ensure_transactions_schema(conn)


def upsert_account_transactions_jsonl(
    jsonl_path: Path | str,
    db_path: Path | str | None = None,
) -> dict[str, Any]:
    """Idempotent load of TXN JSONL into sd_account_transactions.

    Purges prior rows for the same source_file, then inserts.
    stable_id = ``{source_file}:{row_number}``. empty money -> SQL NULL.
    """
    path = Path(jsonl_path)
    result: dict[str, Any] = {
        "ok": False,
        "jsonlPath": str(path),
        "dbPath": None,
        "sourceFile": None,
        "inserted": 0,
        "warnings": [],
    }
    if not path.is_file():
        result["warnings"].append("jsonl missing")
        return result

    target_db = Path(db_path) if db_path else resolve_analytics_db()
    if not target_db:
        target_db = resolve_exports_dir() / "softdent_financial_analytics.db"
    target_db.parent.mkdir(parents=True, exist_ok=True)
    result["dbPath"] = str(target_db)

    meta: dict[str, Any] = {}
    records: list[dict[str, Any]] = []
    try:
        for line in path.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line:
                continue
            obj = json.loads(line)
            if not isinstance(obj, dict):
                continue
            if obj.get("_meta"):
                meta = obj
                continue
            records.append(obj)
    except (OSError, json.JSONDecodeError) as exc:
        result["warnings"].append(f"jsonl_read_{type(exc).__name__}")
        return result

    source_file = None
    if meta.get("sourcePath"):
        source_file = Path(str(meta["sourcePath"])).name
    if not source_file and records:
        source_file = str(records[0].get("source_file") or path.stem + ".xls")
    if not source_file:
        source_file = path.stem + ".xls"
    result["sourceFile"] = source_file
    period_start, period_end = _split_period_hint(meta.get("periodHint"))
    extracted_at = _utc_now()

    conn = sqlite3.connect(str(target_db))
    try:
        ensure_account_transactions_schema(conn)
        conn.execute("BEGIN")
        conn.execute(
            "DELETE FROM sd_account_transactions WHERE source_file = ?",
            (source_file,),
        )
        insert_sql = """
            INSERT INTO sd_account_transactions (
                stable_id, source_file, row_number, account_num, patient_name,
                service_date, provider, procedure, note_flag,
                amount, prod, charges, prod_adj, cash, "check", credit, pay_adj,
                period_start, period_end, extracted_at
            ) VALUES (
                ?, ?, ?, ?, ?,
                ?, ?, ?, ?,
                ?, ?, ?, ?, ?, ?, ?, ?,
                ?, ?, ?
            )
        """
        rows_out = 0
        for rec in records:
            acct = _account_num_str(rec.get("account_num"))
            if not acct:
                continue
            row_number = int(rec.get("row_number") or 0)
            src = str(rec.get("source_file") or source_file)
            stable_id = f"{src}:{row_number}"
            conn.execute(
                insert_sql,
                (
                    stable_id,
                    src,
                    row_number,
                    acct,
                    rec.get("patient_name"),
                    rec.get("date"),
                    rec.get("provider"),
                    rec.get("procedure"),
                    rec.get("note_flag"),
                    _money_or_none(rec.get("amount")),
                    _money_or_none(rec.get("prod")),
                    _money_or_none(rec.get("charges")),
                    _money_or_none(rec.get("prod_adj")),
                    _money_or_none(rec.get("cash")),
                    _money_or_none(rec.get("check")),
                    _money_or_none(rec.get("credit")),
                    _money_or_none(rec.get("pay_adj")),
                    period_start,
                    period_end,
                    extracted_at,
                ),
            )
            rows_out += 1
        conn.commit()
        result["inserted"] = rows_out

        counted = conn.execute(
            "SELECT COUNT(*) FROM sd_account_transactions WHERE source_file = ?",
            (source_file,),
        ).fetchone()[0]
        result["dbCount"] = int(counted)
        expected = meta.get("recordCount")
        if expected is not None and int(expected) != int(counted):
            result["warnings"].append(
                f"parity recordCount={expected} dbCount={counted}"
            )
        donna = conn.execute(
            """
            SELECT COUNT(*) FROM sd_account_transactions
            WHERE source_file = ?
              AND account_num = '27002'
              AND (period_start = '2026-02-01' OR service_date LIKE '2026-02-%')
            """,
            (source_file,),
        ).fetchone()[0]
        result["donna27002Count"] = int(donna)
        # Null honesty spot-check: count rows with amount IS NULL for this source
        null_amt = conn.execute(
            """
            SELECT COUNT(*) FROM sd_account_transactions
            WHERE source_file = ? AND amount IS NULL
            """,
            (source_file,),
        ).fetchone()[0]
        result["nullAmountCount"] = int(null_amt)
        result["ok"] = int(counted) == rows_out and rows_out > 0
        if donna and source_file.upper().startswith("TXN260201"):
            # Live gate for known Feb export
            if int(donna) != 5:
                result["warnings"].append(f"donna_expected_5_got_{donna}")
    except Exception as exc:  # noqa: BLE001
        conn.rollback()
        result["warnings"].append(f"upsert_{type(exc).__name__}:{exc}")
        result["ok"] = False
    finally:
        conn.close()
    return result


def _account_tx_db_row_count(db_path: Path) -> int:
    try:
        conn = sqlite3.connect(f"file:{db_path}?mode=ro", uri=True)
        try:
            row = conn.execute(
                "SELECT COUNT(*) FROM sqlite_master WHERE type='table' AND name='sd_account_transactions'"
            ).fetchone()
            if not row or int(row[0]) == 0:
                return 0
            return int(conn.execute("SELECT COUNT(*) FROM sd_account_transactions").fetchone()[0])
        finally:
            conn.close()
    except Exception:
        return 0


def _query_account_transactions_db(
    account_num: str | int | None = None,
    patient_name: str | None = None,
    date_range: Any = None,
    *,
    db_path: Path | None = None,
    limit: int = 200,
) -> dict[str, Any] | None:
    """SQL query against sd_account_transactions. Returns None if DB unavailable/empty."""
    target = Path(db_path) if db_path else resolve_analytics_db()
    if not target or not target.is_file():
        return None
    if _account_tx_db_row_count(target) <= 0:
        return None

    start, end = _parse_date_range(date_range)
    acct = _account_num_str(account_num) if account_num not in (None, "") else None
    name_q = str(patient_name or "").strip().lower()
    name_tokens = [t for t in name_q.replace(",", " ").split() if t]

    sql = """
        SELECT stable_id, source_file, row_number, account_num, patient_name,
               service_date, provider, procedure, note_flag,
               amount, prod, charges, prod_adj, cash, "check", credit, pay_adj,
               period_start, period_end
        FROM sd_account_transactions
        WHERE 1=1
    """
    params: list[Any] = []
    if acct:
        sql += " AND account_num = ?"
        params.append(acct)
    if name_tokens:
        for tok in name_tokens:
            sql += " AND LOWER(COALESCE(patient_name,'')) LIKE ?"
            params.append(f"%{tok}%")
    if start:
        sql += " AND service_date >= ?"
        params.append(start[:10])
    if end:
        # YYYY-MM expand uses exclusive next-month-01; keep that behavior
        if (
            len(end) == 10
            and end.endswith("-01")
            and start
            and start.endswith("-01")
            and start[:7] != end[:7]
        ):
            sql += " AND service_date < ?"
            params.append(end[:10])
        else:
            sql += " AND service_date <= ?"
            params.append(end[:10])
    sql += " ORDER BY service_date DESC, row_number DESC LIMIT ?"
    params.append(max(1, int(limit)))

    conn = sqlite3.connect(f"file:{target}?mode=ro", uri=True)
    conn.row_factory = sqlite3.Row
    try:
        rows = conn.execute(sql, params).fetchall()
    finally:
        conn.close()

    matches: list[dict[str, Any]] = []
    for row in rows:
        matches.append(
            {
                "date": row["service_date"],
                "account_num": row["account_num"],
                "patient_name": row["patient_name"],
                "provider": row["provider"],
                "procedure": row["procedure"],
                "amount": row["amount"],
                "note_flag": row["note_flag"],
                "row_number": row["row_number"],
                "source_file": row["source_file"],
                "prod": row["prod"],
                "charges": row["charges"],
                "prod_adj": row["prod_adj"],
                "cash": row["cash"],
                "check": row["check"],
                "credit": row["credit"],
                "pay_adj": row["pay_adj"],
                "period_start": row["period_start"],
                "period_end": row["period_end"],
                "stable_id": row["stable_id"],
                "_source": "sd_account_transactions",
            }
        )
    # Newest-first from SQL; HAL samples often prefer chronological — reverse for display
    matches.reverse()

    return {
        "ok": True,
        "reason": None,
        "matchCount": len(matches),
        "matches": matches,
        "filters": {
            "account_num": acct,
            "patient_name": patient_name,
            "date_range": date_range,
        },
        "source": "sd_account_transactions",
    }


def load_parsed_account_transactions(    *,
    parsed_dir: Path | None = None,
    source_stem: str | None = None,
) -> list[dict[str, Any]]:
    """Load typed records from tx_parsed JSONL (skips _meta line)."""
    dest = parsed_dir or resolve_tx_parsed_dir()
    if not dest.is_dir():
        return []
    files = sorted(dest.glob("*.jsonl"), key=lambda p: p.stat().st_mtime, reverse=True)
    if source_stem:
        files = [p for p in files if p.stem.lower() == source_stem.lower()] or files
    if not files:
        return []
    records: list[dict[str, Any]] = []
    for path in files:
        try:
            for line in path.read_text(encoding="utf-8").splitlines():
                line = line.strip()
                if not line:
                    continue
                obj = json.loads(line)
                if isinstance(obj, dict) and obj.get("_meta"):
                    continue
                if isinstance(obj, dict):
                    records.append(obj)
        except (OSError, json.JSONDecodeError):
            continue
    return records


def load_txn_jsonl(
    *,
    parsed_dir: Path | None = None,
    source_stem: str | None = None,
) -> list[dict[str, Any]]:
    """Moonshot alias — stream typed TXN rows from SoftDentFinancialExports/tx_parsed."""
    return load_parsed_account_transactions(parsed_dir=parsed_dir, source_stem=source_stem)

def _parse_date_range(
    date_range: Any,
) -> tuple[str | None, str | None]:
    """Accept '2026-02', '2018', '2018:2019', '2026-02-01:2026-02-28', (start, end), or None."""
    if date_range is None or date_range == "":
        return None, None
    if isinstance(date_range, (list, tuple)) and len(date_range) >= 2:
        start = str(date_range[0] or "").strip() or None
        end = str(date_range[1] or "").strip() or None
        return _normalize_range_bound(start, end=False), _normalize_range_bound(end, end=True)
    text = str(date_range).strip()
    if ":" in text:
        left, right = text.split(":", 1)
        return _normalize_range_bound(left.strip(), end=False), _normalize_range_bound(
            right.strip(), end=True
        )
    if re.fullmatch(r"20\d{2}", text):
        y = int(text)
        return f"{y:04d}-01-01", f"{y:04d}-12-31"
    if len(text) == 7 and text[4] == "-":
        # YYYY-MM → month bounds
        year, month = text.split("-", 1)
        try:
            y, m = int(year), int(month)
            if m == 12:
                return f"{y:04d}-12-01", f"{y:04d}-12-31"
            return f"{y:04d}-{m:02d}-01", f"{y:04d}-{m + 1:02d}-01"
        except ValueError:
            return text, text
    return _normalize_range_bound(text, end=False), _normalize_range_bound(text, end=True)


def _normalize_range_bound(value: str | None, *, end: bool) -> str | None:
    """Expand bare YYYY to year start/end so ISO service_date compares work."""
    if not value:
        return None
    text = str(value).strip()
    if re.fullmatch(r"20\d{2}", text) or re.fullmatch(r"19\d{2}", text):
        y = int(text)
        return f"{y:04d}-12-31" if end else f"{y:04d}-01-01"
    return text


def _date_in_range(date_iso: str | None, start: str | None, end: str | None) -> bool:
    if not date_iso:
        return False
    d = date_iso[:10]
    if start and d < start[:10]:
        return False
    if end:
        end_s = end[:10]
        # Exclusive end when caller passed next-month day from YYYY-MM expand
        if len(end) == 10 and end.endswith("-01") and start and start.endswith("-01") and start[:7] != end[:7]:
            if d >= end_s:
                return False
        elif d > end_s:
            return False
    return True


def account_tx_ledger_coverage(*, db_path: Path | None = None) -> dict[str, Any]:
    """Honest multi-year coverage for HAL (from ingest log + DB; empty ≠ $0)."""
    coverage: dict[str, Any] = {
        "account_tx_multi_year_available": False,
        "dbTotal": 0,
        "serviceDateMin": None,
        "serviceDateMax": None,
        "source": None,
        "honesty": "empty != $0; read-only SoftDent Excel/CSV ledger",
    }
    ingest_log = YEAR_CHUNK_INGEST_LOG
    if ingest_log.is_file():
        try:
            raw = json.loads(ingest_log.read_text(encoding="utf-8"))
            coverage.update(
                {
                    "account_tx_multi_year_available": bool(
                        raw.get("account_tx_multi_year_available")
                    ),
                    "dbTotal": int(raw.get("dbTotal") or 0),
                    "serviceDateMin": (
                        f"{raw.get('serviceYearMin')}-01-01"
                        if raw.get("serviceYearMin")
                        else None
                    ),
                    "serviceDateMax": (
                        f"{raw.get('serviceYearMax')}-12-31"
                        if raw.get("serviceYearMax")
                        else None
                    ),
                    "source": "softdent_account_tx_year_chunks_ingest.json",
                }
            )
        except (OSError, json.JSONDecodeError, TypeError, ValueError):
            pass

    target = Path(db_path) if db_path else resolve_analytics_db()
    if not target or not target.is_file():
        return coverage
    try:
        conn = sqlite3.connect(f"file:{target}?mode=ro", uri=True)
        try:
            if _account_tx_db_row_count(target) <= 0:
                return coverage
            total = int(conn.execute("SELECT COUNT(*) FROM sd_account_transactions").fetchone()[0])
            row = conn.execute(
                """
                SELECT MIN(service_date), MAX(service_date)
                FROM sd_account_transactions
                WHERE service_date IS NOT NULL AND length(service_date) >= 10
                """
            ).fetchone()
            dmin = row[0] if row else None
            dmax = row[1] if row else None
            coverage["dbTotal"] = total
            coverage["serviceDateMin"] = dmin or coverage.get("serviceDateMin")
            coverage["serviceDateMax"] = dmax or coverage.get("serviceDateMax")
            coverage["source"] = "sd_account_transactions"
            ymin = int(str(dmin)[:4]) if dmin else None
            ymax = int(str(dmax)[:4]) if dmax else None
            coverage["account_tx_multi_year_available"] = bool(
                total > 100_000 and ymin and ymax and ymin <= 2017 and ymax >= 2026
            )
        finally:
            conn.close()
    except Exception:
        return coverage
    return coverage


def _attach_account_tx_coverage(result: dict[str, Any], *, db_path: Path | None = None) -> dict[str, Any]:
    coverage = account_tx_ledger_coverage(db_path=db_path)
    result["coverage"] = coverage
    result["account_tx_multi_year_available"] = bool(
        coverage.get("account_tx_multi_year_available")
    )
    result["dbTotal"] = coverage.get("dbTotal")
    result["availableRange"] = {
        "min": coverage.get("serviceDateMin"),
        "max": coverage.get("serviceDateMax"),
    }
    return result


def query_account_transactions(
    account_num: str | int | None = None,
    patient_name: str | None = None,
    date_range: Any = None,
    *,
    parsed_dir: Path | None = None,
    db_path: Path | None = None,
    prefer_db: bool = True,
    limit: int = 200,
) -> dict[str, Any]:
    """Query SoftDent account transactions for HAL / widgets.

    Prefers ``sd_account_transactions`` in the analytics DB; falls back to
    tx_parsed JSONL / live TXN*.xls when the table is empty. empty != $0.
    """
    if prefer_db and parsed_dir is None:
        db_hit = _query_account_transactions_db(
            account_num=account_num,
            patient_name=patient_name,
            date_range=date_range,
            db_path=db_path,
            limit=limit,
        )
        if db_hit is not None:
            return _attach_account_tx_coverage(db_hit, db_path=db_path)

    records = load_parsed_account_transactions(parsed_dir=parsed_dir)
    if not records:
        # Best-effort: parse live inbox file if present but not yet ingested
        inbox = DEFAULT_TXN_INBOX / "TXN260201.xls"
        candidates = sorted(
            DEFAULT_TXN_INBOX.glob("TXN*.xls*"),
            key=lambda p: p.stat().st_mtime,
            reverse=True,
        )
        target = candidates[0] if candidates else inbox
        if target.is_file():
            parsed = parse_account_transactions_xls(target)
            records = list(parsed.get("records") or [])
        if not records:
            out = {
                "ok": False,
                "reason": "data not yet exported",
                "message": (
                    "Account transaction data not yet exported. "
                    "Pull SoftDent Reports → Accounting → Trans for a Period → Excel "
                    r"into C:\SoftDentReportExports, then ingest to "
                    r"C:\SoftDentFinancialExports\tx_parsed\ and "
                    r"softdent_financial_analytics.db."
                ),
                "matches": [],
                "matchCount": 0,
            }
            return _attach_account_tx_coverage(out, db_path=db_path)

    start, end = _parse_date_range(date_range)
    acct = _account_num_str(account_num) if account_num not in (None, "") else None
    name_q = str(patient_name or "").strip().lower()
    name_tokens = [t for t in name_q.replace(",", " ").split() if t]

    matches: list[dict[str, Any]] = []
    for rec in records:
        if acct and str(rec.get("account_num") or "") != acct:
            continue
        if name_tokens:
            pname = str(rec.get("patient_name") or "").lower()
            if not all(tok in pname for tok in name_tokens):
                continue
        if start or end:
            if not _date_in_range(rec.get("date"), start, end):
                continue
        matches.append(rec)
        if len(matches) >= max(1, int(limit)):
            break

    result = {
        "ok": True,
        "reason": None,
        "matchCount": len(matches),
        "matches": matches,
        "filters": {
            "account_num": acct,
            "patient_name": patient_name,
            "date_range": date_range,
        },
        "source": "jsonl_or_xls",
    }
    return _attach_account_tx_coverage(result, db_path=db_path)


def format_account_transactions_hal_reply(result: dict[str, Any], *, max_lines: int = 12) -> str:
    """HAL-facing reply from query_account_transactions (honesty: empty ≠ $0)."""
    coverage = (result or {}).get("coverage") or {}
    avail_min = coverage.get("serviceDateMin") or ((result or {}).get("availableRange") or {}).get(
        "min"
    )
    avail_max = coverage.get("serviceDateMax") or ((result or {}).get("availableRange") or {}).get(
        "max"
    )
    db_total = (result or {}).get("dbTotal") or coverage.get("dbTotal")
    multi = bool(
        (result or {}).get("account_tx_multi_year_available")
        or coverage.get("account_tx_multi_year_available")
    )
    coverage_bits = []
    if multi:
        coverage_bits.append("account_tx_multi_year_available=true")
    if db_total:
        coverage_bits.append(f"db_total={db_total}")
    if avail_min and avail_max:
        coverage_bits.append(f"available_range={avail_min} to {avail_max}")
    coverage_s = f" [{'; '.join(coverage_bits)}]" if coverage_bits else ""

    if not result or not result.get("ok"):
        base = str(
            (result or {}).get("message")
            or "Account transaction data not yet exported."
        )
        return f"{base}{coverage_s}".strip()
    matches = list(result.get("matches") or [])
    if not matches:
        filters = result.get("filters") or {}
        return (
            "No matching account transactions for these filters "
            f"(filters={filters}; source={result.get('source') or 'unknown'}). "
            f"Ledger is read-only SoftDent ingest — empty != $0.{coverage_s}"
        )
    lines = [
        f"SoftDent account transactions "
        f"({result.get('source') or 'parsed TXN'}; {len(matches)} match(es); empty != $0)"
        f"{coverage_s}:"
    ]
    for rec in matches[: max(1, int(max_lines))]:
        amt = rec.get("amount")
        amt_s = "null" if amt is None else f"{amt:g}"
        lines.append(
            f"- {rec.get('date') or '?'} | acct {rec.get('account_num') or '?'} | "
            f"{(rec.get('patient_name') or '').strip()} | Dr {rec.get('provider') or '—'} | "
            f"code {rec.get('procedure') or '—'} | amt {amt_s}"
            + (f" | flag {rec.get('note_flag')}" if rec.get("note_flag") else "")
        )
    if len(matches) > max_lines:
        lines.append(f"… +{len(matches) - max_lines} more")
    return " ".join(lines) if len(lines) == 1 else "\n".join(lines)


if __name__ == "__main__":
    import sys

    if len(sys.argv) > 1 and str(sys.argv[1]).lower().endswith((".xls", ".xlsx", ".xlsm")):
        print(json.dumps(ingest_account_transactions_xls(sys.argv[1]), indent=2, default=str))
    else:
        db = Path(sys.argv[1]) if len(sys.argv) > 1 else None
        print(json.dumps(extract_all_transactions(db_path=db, force=True), indent=2, default=str))
