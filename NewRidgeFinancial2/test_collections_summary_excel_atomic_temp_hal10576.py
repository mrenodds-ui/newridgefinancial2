"""Moonshot Collections Summary Excel-temp atomic finalize (hal-10576)."""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from apex_backend import BUILD_ID
from softdent_excel_temp import (
    call_with_excel_temp_retry,
    collections_export_health,
    is_excel_temp_lock_error,
)
from softdent_practice_exports import atomic_copy_export, atomic_write_excel_export


class Hal10576CollectionsExcelAtomicTempTests(unittest.TestCase):
    def test_build_id(self) -> None:
        self.assertEqual(BUILD_ID, "hal-10576")

    def test_atomic_write_excel_export_success(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            dest = Path(tmp) / "COL_SUMMARY.xlsx"

            def _write(path: Path) -> None:
                try:
                    from openpyxl import Workbook
                except ImportError:
                    path.write_bytes(b"PK\x03\x04minimal-xlsx-stub-not-empty")
                    return
                wb = Workbook()
                ws = wb.active
                ws.append(["Period", "Collections", "Ins Plan", "Patient"])
                ws.append(["2026-07", 30626.42, 0.0, 0.0])  # Ins Plan $0 is SoftDent truth
                wb.save(str(path))

            meta = atomic_write_excel_export(
                dest,
                _write,
                event="collections_summary_export_success",
            )
            self.assertTrue(meta.get("ok"))
            self.assertEqual(meta.get("event"), "collections_summary_export_success")
            self.assertTrue(meta.get("temp_cleanup"))
            self.assertFalse(meta.get("writeBack"))
            self.assertTrue(dest.is_file())
            self.assertGreater(int(meta.get("bytes") or 0), 0)
            # No leftover sibling temps
            leftovers = [p for p in Path(tmp).iterdir() if p.suffix.endswith(".tmp") or ".tmp" in p.name]
            self.assertEqual(leftovers, [])
            try:
                from openpyxl import load_workbook

                book = load_workbook(str(dest), data_only=True, read_only=True)
                try:
                    rows = list(book.active.iter_rows(values_only=True))
                finally:
                    book.close()
                self.assertGreaterEqual(len(rows), 2)
                self.assertEqual(rows[1][2], 0.0)  # Ins Plan stays $0 — honesty
            except ImportError:
                pass

    def test_atomic_write_rejects_zero_byte_and_cleans_temp(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            dest = Path(tmp) / "empty.xls"

            def _write(path: Path) -> None:
                path.write_bytes(b"")

            with self.assertRaises(RuntimeError):
                atomic_write_excel_export(dest, _write)
            self.assertFalse(dest.exists())
            leftovers = list(Path(tmp).glob("*.tmp"))
            self.assertEqual(leftovers, [])

    def test_atomic_copy_export(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            src = Path(tmp) / "src.csv"
            dest = Path(tmp) / "dest.csv"
            src.write_text("a,b\n1,2\n", encoding="utf-8")
            meta = atomic_copy_export(src, dest)
            self.assertTrue(meta.get("ok"))
            self.assertTrue(meta.get("temp_cleanup"))
            self.assertEqual(dest.read_text(encoding="utf-8"), "a,b\n1,2\n")

    def test_retry_passes_through_non_lock_errors(self) -> None:
        calls = {"n": 0}

        def _boom() -> None:
            calls["n"] += 1
            raise ValueError("not a lock")

        with self.assertRaises(ValueError):
            call_with_excel_temp_retry(_boom)
        self.assertEqual(calls["n"], 1)

    def test_lock_error_detection(self) -> None:
        self.assertTrue(is_excel_temp_lock_error(PermissionError(13, "Access is denied")))
        self.assertFalse(is_excel_temp_lock_error(ValueError("nope")))

    def test_collections_export_health_shape(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            health = collections_export_health(dest_root=Path(tmp), temp_dir=Path(tmp))
            self.assertTrue(health.get("ok"))
            self.assertEqual(health.get("phase"), "hal-10576")
            self.assertFalse(health.get("writeBack"))
            self.assertEqual(health.get("honesty"), "empty_not_zero")
            self.assertFalse(health.get("collectionsExportReady"))
            self.assertEqual(health.get("errorCode"), "no_exports")


if __name__ == "__main__":
    unittest.main()
