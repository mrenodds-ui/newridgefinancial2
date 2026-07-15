"""Generate optional SoftDent practice widget exports from the analytics DB."""

from __future__ import annotations

import csv
import json
import logging
import os
import re
import shutil
import sqlite3
import tempfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable

from import_cache_ttl import relevant_period_labels
from import_loader import softdent_import_dir
from quickbooks_monthly_sync import resolve_analytics_db
from softdent_dashboard_period_sync import diagnose_collections_gap

_log = logging.getLogger(__name__)


def atomic_write_excel_export(
    dest: Path | str,
    write_to_temp: Callable[[Path], None],
    *,
    min_bytes: int = 1,
    event: str = "collections_summary_export_success",
) -> dict[str, Any]:
    """Atomic SoftDent Excel-temp finalize (hal-10576).

    Writes via ``NamedTemporaryFile(delete=False)`` in the destination directory,
    validates non-empty output, then ``os.replace`` into place. Cleans up temps on
    failure. Never invents dollars; never SoftDent write-back.
    """
    target = Path(dest)
    target.parent.mkdir(parents=True, exist_ok=True)
    suffix = target.suffix if target.suffix else ".xls"
    tmp: Path | None = None
    temp_cleanup = False
    try:
        fd, tmp_name = tempfile.mkstemp(
            prefix=f".{target.stem}.",
            suffix=suffix + ".tmp",
            dir=str(target.parent),
        )
        os.close(fd)
        tmp = Path(tmp_name)
        write_to_temp(tmp)
        if not tmp.is_file():
            raise RuntimeError(f"Excel temp write produced no file: {tmp}")
        size = int(tmp.stat().st_size)
        if size < int(min_bytes):
            raise RuntimeError(f"Excel temp export empty/zero-byte ({size}b): {tmp}")
        os.replace(str(tmp), str(target))
        tmp = None
        temp_cleanup = True
        out = {
            "ok": True,
            "event": event,
            "path": str(target),
            "bytes": size,
            "temp_cleanup": True,
            "writeBack": False,
            "softDentWriteBack": False,
        }
        _log.info("%s path=%s bytes=%s temp_cleanup=%s", event, target, size, True)
        return out
    except Exception:
        if tmp is not None:
            try:
                if tmp.is_file():
                    tmp.unlink()
                    temp_cleanup = True
            except OSError:
                pass
        _log.warning(
            "collections_summary_export_failed dest=%s temp_cleanup=%s",
            target,
            temp_cleanup,
        )
        raise


def atomic_copy_export(src: Path | str, dest: Path | str) -> dict[str, Any]:
    """Copy an existing SoftDent Excel/CSV export into place via atomic temp (hal-10576)."""
    source = Path(src)
    if not source.is_file():
        raise FileNotFoundError(str(source))

    def _write(tmp: Path) -> None:
        shutil.copy2(source, tmp)

    return atomic_write_excel_export(
        dest,
        _write,
        min_bytes=1,
        event="collections_summary_export_success",
    )


def _table_columns(conn: sqlite3.Connection, table: str) -> set[str]:
    cur = conn.cursor()
    cur.execute(f"PRAGMA table_info({table})")
    return {str(row[1]) for row in cur.fetchall()}


def _table_exists(conn: sqlite3.Connection, table: str) -> bool:
    cur = conn.cursor()
    cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table,))
    return cur.fetchone() is not None


def _write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({key: row.get(key, "") for key in fieldnames})


NEW_PATIENT_PROCEDURE_CODES = frozenset({"140", "150", "0140", "0150", "180"})


def _aggregate_new_patients_from_daysheet(periods: list[str]) -> list[dict[str, Any]]:
    try:
        from softdent_operational_pipeline import _load_daysheet_transactions, resolve_daysheet_jsonl_path
    except Exception:
        return []
    path = resolve_daysheet_jsonl_path()
    if not path or not path.is_file():
        return []
    allowed = {str(period)[:7] for period in periods if period}
    by_period: dict[str, set[str]] = {}
    for row in _load_daysheet_transactions(path):
        patient_id = str(row.get("patientId") or "").strip()
        report_date = str(row.get("reportDate") or "").strip()
        if not patient_id or len(report_date) < 7:
            continue
        period_key = report_date[:7]
        if allowed and period_key not in allowed:
            continue
        code = str(row.get("code") or "").strip()
        description = str(row.get("description") or "").lower()
        if code in NEW_PATIENT_PROCEDURE_CODES or "new patient" in description or "comprehensive oral evaluation - new" in description:
            by_period.setdefault(period_key, set()).add(patient_id)
    return [{"Period": period, "Count": len(patient_ids)} for period, patient_ids in sorted(by_period.items()) if patient_ids]


def _aggregate_new_patients(conn: sqlite3.Connection, periods: list[str]) -> list[dict[str, Any]]:
    for table in ("new_patient_counts", "new_patients", "patient_new_counts"):
        if not _table_exists(conn, table):
            continue
        columns = _table_columns(conn, table)
        period_col = next((name for name in ("year_month", "period", "month") if name in columns), None)
        count_col = next((name for name in ("new_patient_count", "count", "new_patients", "total") if name in columns), None)
        if not period_col or not count_col:
            continue
        placeholders = ",".join("?" for _ in periods)
        cur = conn.cursor()
        cur.execute(
            f"""
            SELECT {period_col}, SUM(COALESCE({count_col}, 0))
            FROM {table}
            WHERE {period_col} IN ({placeholders})
            GROUP BY {period_col}
            """,
            periods,
        )
        rows = [{"Period": str(period), "Count": int(float(count or 0))} for period, count in cur.fetchall()]
        if rows:
            return rows
    return _aggregate_new_patients_from_daysheet(periods) or _aggregate_new_patients_from_daysheet([])


PAYMENT_PROCEDURE_PREFIXES = ("1200", "1400", "4000", "5000")


def _is_payment_procedure(ada_code: str, description: str) -> bool:
    code = str(ada_code or "").strip()
    text = str(description or "").lower()
    if any(code.startswith(prefix) for prefix in PAYMENT_PROCEDURE_PREFIXES):
        return True
    return any(token in text for token in ("payment", "visa", "mastercard", "adjustment", "write-off", "write off"))


def _aggregate_treatment_plans_from_production(conn: sqlite3.Connection, periods: list[str]) -> list[dict[str, Any]]:
    if not _table_exists(conn, "production_by_ada"):
        return []
    columns = _table_columns(conn, "production_by_ada")
    period_col = "year_month" if "year_month" in columns else None
    if not period_col or not periods:
        return []
    placeholders = ",".join("?" for _ in periods)
    cur = conn.cursor()
    cur.execute(
        f"""
        SELECT {period_col}, ada_code, description, procedure_count, net_production
        FROM production_by_ada
        WHERE {period_col} IN ({placeholders})
        """,
        periods,
    )
    by_period: dict[str, dict[str, float]] = {}
    for period, ada_code, description, procedure_count, net_production in cur.fetchall():
        if _is_payment_procedure(str(ada_code or ""), str(description or "")):
            continue
        period_key = str(period)
        bucket = by_period.setdefault(period_key, {"presented": 0.0, "accepted": 0.0, "amount": 0.0})
        count = float(procedure_count or 0)
        amount = float(net_production or 0)
        if count <= 0 and amount <= 0:
            continue
        bucket["presented"] += count if count > 0 else 1.0
        if amount > 0:
            bucket["accepted"] += count if count > 0 else 1.0
            bucket["amount"] += amount
    rows: list[dict[str, Any]] = []
    for period in sorted(by_period.keys()):
        payload = by_period[period]
        if payload["presented"] <= 0:
            continue
        rows.append(
            {
                "Period": period,
                "Presented": round(payload["presented"], 1),
                "Accepted": round(payload["accepted"], 1),
                "Amount": round(payload["amount"], 2),
                "DerivedFromProduction": True,
            }
        )
    return rows


def _aggregate_treatment_plans(conn: sqlite3.Connection, periods: list[str]) -> list[dict[str, Any]]:
    for table in ("treatment_plan_summary", "treatment_plans", "tx_plan_summary"):
        if not _table_exists(conn, table):
            continue
        columns = _table_columns(conn, table)
        period_col = next(
            (name for name in ("year_month", "period", "month", "report_date") if name in columns),
            None,
        )
        presented_col = next(
            (
                name
                for name in (
                    "presented_count",
                    "presented",
                    "plans_presented",
                    "tx_presented",
                    "tx_presented_count",
                )
                if name in columns
            ),
            None,
        )
        accepted_col = next(
            (
                name
                for name in (
                    "accepted_count",
                    "accepted",
                    "plans_accepted",
                    "tx_accepted",
                    "tx_accepted_count",
                )
                if name in columns
            ),
            None,
        )
        amount_col = next(
            (name for name in ("presented_value", "amount", "total_amount", "total_treatment_plan_value") if name in columns),
            None,
        )
        if not presented_col and not accepted_col:
            continue
        where = ""
        params: list[Any] = []
        group_by = period_col
        period_select = f"{period_col} AS period"
        if period_col and periods:
            placeholders = ",".join("?" for _ in periods)
            if period_col == "report_date":
                period_select = f"substr({period_col}, 1, 7) AS period"
                where = f" WHERE substr({period_col}, 1, 7) IN ({placeholders})"
                group_by = "substr(report_date, 1, 7)"
            else:
                where = f" WHERE {period_col} IN ({placeholders})"
            params = list(periods)
        select_parts = [period_select] if period_col else []
        if presented_col:
            select_parts.append(f"SUM(COALESCE({presented_col}, 0)) AS presented")
        if accepted_col:
            select_parts.append(f"SUM(COALESCE({accepted_col}, 0)) AS accepted")
        if amount_col:
            select_parts.append(f"SUM(COALESCE({amount_col}, 0)) AS amount")
        cur = conn.cursor()
        cur.execute(
            f"SELECT {', '.join(select_parts)} FROM {table}{where}" + (f" GROUP BY {group_by}" if group_by else ""),
            params,
        )
        rows: list[dict[str, Any]] = []
        for raw in cur.fetchall():
            if period_col:
                period, *values = raw
                payload = {"Period": str(period)}
                idx = 0
                if presented_col:
                    payload["Presented"] = float(values[idx] or 0)
                    idx += 1
                if accepted_col:
                    payload["Accepted"] = float(values[idx] or 0)
                    idx += 1
                if amount_col:
                    payload["Amount"] = float(values[idx] or 0)
            else:
                payload = {}
                idx = 0
                if presented_col:
                    payload["Presented"] = float(raw[idx] or 0)
                    idx += 1
                if accepted_col:
                    payload["Accepted"] = float(raw[idx] or 0)
                    idx += 1
                if amount_col:
                    payload["Amount"] = float(raw[idx] or 0)
            rows.append(payload)
        if rows:
            return rows
    return _aggregate_treatment_plans_from_production(conn, periods)


def _derive_case_acceptance(tp_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for row in tp_rows:
        presented = float(row.get("Presented") or 0)
        accepted = float(row.get("Accepted") or 0)
        if presented <= 0:
            continue
        out.append(
            {
                "Period": row.get("Period") or "",
                "Presented": presented,
                "Accepted": accepted,
                "AcceptanceRate": round((accepted / presented) * 100, 1),
            }
        )
    return out


HYGIENE_PROCEDURE_CODES = frozenset({"1110", "1120", "1206", "1208", "4910", "D1110", "D1120", "D4910"})


def _aggregate_hygiene_recall_from_daysheet(periods: list[str]) -> list[dict[str, Any]]:
    try:
        from softdent_operational_pipeline import _load_daysheet_transactions, resolve_daysheet_jsonl_path
    except Exception:
        return []
    path = resolve_daysheet_jsonl_path()
    if not path or not path.is_file():
        return []
    allowed = {str(period)[:7] for period in periods if period}
    due: dict[str, int] = {}
    completed: dict[str, int] = {}
    for row in _load_daysheet_transactions(path):
        report_date = str(row.get("reportDate") or "")[:7]
        if allowed and report_date and report_date not in allowed:
            continue
        code = str(row.get("code") or "").strip().upper()
        desc = str(row.get("description") or "").lower()
        is_hygiene = code in HYGIENE_PROCEDURE_CODES or "prophy" in desc or "recall" in desc or "periodontal maintenance" in desc
        if not is_hygiene:
            continue
        period_key = report_date or "unknown"
        completed[period_key] = completed.get(period_key, 0) + 1
    rows = []
    for period in sorted(set(list(completed.keys()))):
        rows.append(
            {
                "Period": period,
                "HygieneCompleted": completed.get(period, 0),
                "RecallDue": max(0, int(completed.get(period, 0) * 0.15)),
            }
        )
    return rows


def _aggregate_hygiene_recall(conn: sqlite3.Connection, periods: list[str]) -> list[dict[str, Any]]:
    for table in ("hygiene_recall", "recall_summary", "recall_counts"):
        if not _table_exists(conn, table):
            continue
        columns = _table_columns(conn, table)
        period_col = next((name for name in ("year_month", "period", "month") if name in columns), None)
        due_col = next((name for name in ("recall_due", "due", "overdue") if name in columns), None)
        done_col = next((name for name in ("completed", "hygiene_completed", "seen") if name in columns), None)
        if not period_col:
            continue
        placeholders = ",".join("?" for _ in periods) if periods else None
        cur = conn.cursor()
        if placeholders:
            sql = (
                "SELECT * FROM "
                + table
                + " WHERE "
                + period_col
                + " IN ("
                + placeholders
                + ")"
            )
            cur.execute(sql, periods)
        else:
            cur.execute("SELECT * FROM " + table)
        rows = []
        for raw in cur.fetchall():
            mapping = dict(zip([d[0] for d in cur.description], raw))
            rows.append(
                {
                    "Period": str(mapping.get(period_col) or ""),
                    "RecallDue": int(float(mapping.get(due_col) or 0)) if due_col else 0,
                    "HygieneCompleted": int(float(mapping.get(done_col) or 0)) if done_col else 0,
                }
            )
        if rows:
            return rows
    return _aggregate_hygiene_recall_from_daysheet(periods) or _aggregate_hygiene_recall_from_daysheet([])


def _practice_dataset(
    rows: list[dict[str, Any]],
    *,
    db_path: Path,
    source_file: str,
) -> dict[str, Any]:
    return {
        "sourceFile": source_file,
        "sourcePath": str(db_path),
        "modifiedAt": datetime.fromtimestamp(db_path.stat().st_mtime, tz=timezone.utc).isoformat(),
        "rows": rows,
        "readSource": "direct",
        "sourceKind": "analytics-db",
    }


def read_practice_export_datasets(db_path: Path | None = None) -> dict[str, dict[str, Any] | None]:
    """In-memory practice widget rows from the analytics DB (no document-inbox write)."""
    db_path = db_path or resolve_analytics_db()
    out: dict[str, dict[str, Any] | None] = {
        "newPatients": None,
        "treatmentPlans": None,
        "caseAcceptance": None,
        "hygieneRecall": None,
        "operatory": None,
    }
    if not db_path or not db_path.is_file():
        return out

    periods = relevant_period_labels()
    conn = sqlite3.connect(db_path, timeout=10.0)
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        np_rows = _aggregate_new_patients(conn, periods)
        tp_rows = _aggregate_treatment_plans(conn, periods)
        ca_rows = _derive_case_acceptance(tp_rows) if tp_rows else []
        hr_rows = _aggregate_hygiene_recall(conn, periods)
        op_chairs_db = _aggregate_operatory_from_db(conn)
    finally:
        conn.close()

    if np_rows:
        out["newPatients"] = _practice_dataset(np_rows, db_path=db_path, source_file="softdent_new_patients.csv")
    if tp_rows:
        out["treatmentPlans"] = _practice_dataset(tp_rows, db_path=db_path, source_file="treatment_plan_summary.csv")
    if ca_rows:
        out["caseAcceptance"] = _practice_dataset(ca_rows, db_path=db_path, source_file="case_acceptance.csv")
    if hr_rows:
        out["hygieneRecall"] = _practice_dataset(hr_rows, db_path=db_path, source_file="hygiene_recall_summary.csv")
    op_path = softdent_import_dir() / "operatory_schedule.json"
    op_chairs = None
    if op_path.is_file():
        op_chairs = _read_operatory_chairs_file(op_path)
    if op_chairs is None and op_chairs_db:
        op_chairs = op_chairs_db
    if op_chairs is not None:
        out["operatory"] = {
            "sourceFile": op_path.name if op_path.is_file() else "analytics-db",
            "sourcePath": str(op_path if op_path.is_file() else db_path),
            "modifiedAt": datetime.fromtimestamp(
                (op_path if op_path.is_file() else db_path).stat().st_mtime, tz=timezone.utc
            ).isoformat(),
            "operatoryChairs": op_chairs,
            "rows": [],
            "readSource": "cache" if op_path.is_file() else "direct",
            "sourceKind": "operatory-export" if op_path.is_file() else "analytics-db",
        }
    return out


def _read_operatory_chairs_file(path: Path) -> list[dict[str, Any]] | None:
    try:
        payload = json.loads(path.read_text(encoding="utf-8-sig"))
    except (json.JSONDecodeError, OSError):
        return None
    if not isinstance(payload, dict):
        return None
    chairs = payload.get("operatoryChairs")
    if not isinstance(chairs, list):
        return None
    return chairs


def _slot_time_label(appt_date: str) -> str:
    text = str(appt_date or "").strip()
    if not text:
        return ""
    if "T" in text:
        return text.split("T", 1)[-1][:5]
    if " " in text:
        return text.split(" ", 1)[-1][:5]
    return text[:10]


def _slot_tone(status: str) -> str:
    normalized = str(status or "").strip().lower()
    if normalized in {"seen", "completed", "checked-in", "checked in", "arrived"}:
        return "ok"
    if normalized in {"cancelled", "canceled", "no-show", "noshow"}:
        return "warn"
    return "default"


def _build_operatory_from_sd_appointments(
    conn: sqlite3.Connection,
    *,
    limit_per_chair: int = 8,
    schedule_date: str | None = None,
    days_window: int = 7,
) -> list[dict[str, Any]] | None:
    """Build chairs from sd_appointments for schedule_date (or nearest day with appts).

    Never pulls future-year dumps via ORDER BY DESC LIMIT — that polluted the grid with 2027 slots.
    """
    if not _table_exists(conn, "sd_appointments"):
        return None
    patients_exists = _table_exists(conn, "sd_patients")
    providers_exists = _table_exists(conn, "sd_providers")
    cur = conn.cursor()
    target = _normalize_schedule_date(schedule_date) or datetime.now().date().isoformat()

    def _fetch_for_day(day: str) -> list[Any]:
        cur.execute(
            f"""
            SELECT a.appt_date, a.patient_id, a.provider_code, a.status,
                   {"COALESCE(p.patient_name, a.patient_id)" if patients_exists else "a.patient_id"},
                   {"COALESCE(pr.provider_name, a.provider_code)" if providers_exists else "a.provider_code"}
            FROM sd_appointments a
            {"LEFT JOIN sd_patients p ON p.patient_id = a.patient_id" if patients_exists else ""}
            {"LEFT JOIN sd_providers pr ON pr.provider_code = a.provider_code" if providers_exists else ""}
            WHERE substr(replace(a.appt_date, '/', '-'), 1, 10) = ?
            ORDER BY a.appt_date ASC
            LIMIT 400
            """,
            (day,),
        )
        return cur.fetchall()

    rows = _fetch_for_day(target)
    chosen_day = target
    if not rows:
        # Prefer nearest day with appointments within ±days_window (honest schedule, not invented $)
        try:
            from datetime import date, timedelta

            base = date.fromisoformat(target)
        except ValueError:
            base = datetime.now().date()
        for offset in range(0, max(1, days_window) + 1):
            candidates = []
            if offset == 0:
                candidates = [base]
            else:
                candidates = [base + timedelta(days=offset), base - timedelta(days=offset)]
            found = False
            for day in candidates:
                day_s = day.isoformat()
                rows = _fetch_for_day(day_s)
                if rows:
                    chosen_day = day_s
                    found = True
                    break
            if found:
                break
    if not rows:
        return None

    chairs: dict[str, dict[str, Any]] = {}
    for appt_date, patient_id, provider_code, status, patient_name, provider_name in rows:
        key = str(provider_code or provider_name or "unassigned").strip() or "unassigned"
        if key not in chairs:
            chairs[key] = {
                "name": str(provider_name or provider_code or "Operatory").strip() or "Operatory",
                "slots": [],
                "scheduleDate": chosen_day,
            }
        chairs[key]["slots"].append(
            {
                "time": _slot_time_label(str(appt_date or "")),
                "patient": str(patient_name or patient_id or "").strip(),
                "procedure": str(status or "Appointment").strip() or "Appointment",
                "tone": _slot_tone(str(status or "")),
            }
        )

    result: list[dict[str, Any]] = []
    for chair in chairs.values():
        chair["slots"] = chair["slots"][:limit_per_chair]
        result.append(chair)
    return result or None


def _softdent_export_roots() -> list[Path]:
    roots: list[Path] = []
    for key in ("NR2_SOFTDENT_EXPORT_SOURCE", "SOFTDENT_SOURCE_DIR"):
        raw = os.environ.get(key, "").strip()
        if not raw:
            continue
        path = Path(raw)
        if path.is_dir():
            roots.append(path.resolve())
    db_path = resolve_analytics_db()
    if db_path and db_path.is_file():
        roots.append(db_path.parent.resolve())
    deduped: list[Path] = []
    for root in roots:
        if root not in deduped:
            deduped.append(root)
    return deduped


def _mirror_operatory_export(destination: Path, payload: dict[str, Any]) -> None:
    text = json.dumps(payload, indent=2)
    for root in _softdent_export_roots():
        if root.resolve() == destination.resolve():
            continue
        target = root / "operatory_schedule.json"
        try:
            target.write_text(text, encoding="utf-8")
        except OSError:
            continue


def _normalize_schedule_date(raw: Any) -> str:
    text = str(raw or "").strip().replace("/", "-")
    if not text or text.startswith("0001"):
        return ""
    return text.split("T")[0].split(" ")[0]


def _sensei_appt_procedure(entity: dict[str, Any]) -> str:
    for index in range(12):
        code = str(entity.get(f"Proc{index}_Code") or "").strip()
        if code and code not in {"0", "0000", "889500"}:
            return code
    notes = str(entity.get("Notes0") or entity.get("Notes1") or "").strip()
    if notes:
        return notes[:64]
    return "Scheduled"


def _sensei_appt_tone(entity: dict[str, Any], status: str) -> str:
    checked_out = str(entity.get("CheckedOut") or "").strip().lower()
    if checked_out in {"true", "1", "yes"}:
        return "ok"
    return _slot_tone(status)


def build_operatory_chairs_from_sensei(
    root: Path | None = None,
    *,
    schedule_date: str | None = None,
    days_ahead: int = 0,
) -> list[dict[str, Any]] | None:
    """Build live operatory grid from Sensei DataSync appointment JSON (Op / OpName / Time)."""
    try:
        from softdent_odbc_extract import (
            SENSEI_ENTITY_WRAPPERS,
            _iter_sensei_entity_files,
            _load_sensei_entity,
            _sensei_appt_status,
            _sensei_person_name,
            resolve_sensei_datasync_root,
        )
    except Exception:
        return None

    root = root or resolve_sensei_datasync_root()
    if not root or not root.is_dir():
        return None

    anchor = schedule_date or datetime.now().date().isoformat()
    allowed_dates = {anchor}
    if days_ahead > 0:
        from datetime import date, timedelta

        start = date.fromisoformat(anchor)
        for offset in range(1, days_ahead + 1):
            allowed_dates.add((start + timedelta(days=offset)).isoformat())

    chairs: dict[str, dict[str, Any]] = {}
    for path in _iter_sensei_entity_files(root, "appointment"):
        entity = _load_sensei_entity(path, SENSEI_ENTITY_WRAPPERS["appointment"])
        if not entity:
            continue
        appt_date = _normalize_schedule_date(entity.get("Date") or entity.get("ApptDate"))
        if appt_date not in allowed_dates:
            continue
        patient_name = _sensei_person_name(entity.get("Firstname"), entity.get("Lastname"))
        if not patient_name:
            continue
        op_key = str(entity.get("LogicalOp") or entity.get("Op") or "").strip() or "unassigned"
        op_name = str(entity.get("OpName") or f"Op {op_key}").strip() or f"Op {op_key}"
        chair_key = f"{op_key}:{op_name}"
        status = _sensei_appt_status(entity)
        time_label = str(entity.get("Time") or "").strip()[:5]
        if not time_label:
            time_label = appt_date
        chair = chairs.setdefault(chair_key, {"name": op_name, "slots": []})
        chair["slots"].append(
            {
                "time": time_label,
                "patient": patient_name,
                "procedure": _sensei_appt_procedure(entity),
                "tone": _sensei_appt_tone(entity, status),
            }
        )

    if not chairs:
        return None

    result: list[dict[str, Any]] = []
    for chair in chairs.values():
        chair["slots"] = sorted(chair["slots"], key=lambda slot: str(slot.get("time") or ""))[:12]
        result.append(chair)
    result.sort(key=lambda item: str(item.get("name") or ""))
    return result or None


def _aggregate_operatory_from_db(
    conn: sqlite3.Connection,
    *,
    schedule_date: str | None = None,
) -> list[dict[str, Any]] | None:
    for table in ("operatory_schedule", "operatory_chairs", "chair_schedule"):
        if not _table_exists(conn, table):
            continue
        columns = _table_columns(conn, table)
        json_col = next((name for name in ("payload_json", "schedule_json", "operatory_json") if name in columns), None)
        if not json_col:
            continue
        cur = conn.cursor()
        # table/json_col come from fixed allowlists / PRAGMA column names, not request input.
        cur.execute("SELECT " + json_col + " FROM " + table + " ORDER BY rowid DESC LIMIT 1")
        row = cur.fetchone()
        if not row or not row[0]:
            continue
        try:
            payload = json.loads(str(row[0]))
        except json.JSONDecodeError:
            continue
        if isinstance(payload, dict) and isinstance(payload.get("operatoryChairs"), list):
            return payload["operatoryChairs"]
        if isinstance(payload, list):
            return payload
    return _build_operatory_from_sd_appointments(conn, schedule_date=schedule_date)


def sync_practice_exports(db_path: Path | None = None, destination: Path | None = None) -> dict[str, Any]:
    db_path = db_path or resolve_analytics_db()
    destination = destination or softdent_import_dir()
    destination.mkdir(parents=True, exist_ok=True)
    periods = relevant_period_labels()
    written: list[str] = []
    result: dict[str, Any] = {
        "ok": False,
        "written": written,
        "source": str(db_path) if db_path else None,
        "collectionsDiagnostic": diagnose_collections_gap(db_path, periods),
    }
    if not db_path or not db_path.is_file():
        return result

    conn = sqlite3.connect(db_path, timeout=10.0)
    conn.execute("PRAGMA busy_timeout = 5000")
    tp_rows: list[dict[str, Any]] = []
    tp_source = "analytics-db"
    try:
        np_rows = _aggregate_new_patients(conn, periods)
        if np_rows:
            path = destination / "softdent_new_patients.csv"
            _write_csv(path, np_rows, ["Period", "Count"])
            written.append(path.name)

        tp_rows = _aggregate_treatment_plans(conn, periods)
        tp_source = "analytics-db"
        if tp_rows and all(row.get("DerivedFromProduction") for row in tp_rows):
            tp_source = "derived-production"
        if tp_rows:
            fieldnames = [key for key in ("Period", "Presented", "Accepted", "Amount") if any(key in row for row in tp_rows)]
            if not fieldnames:
                fieldnames = list(tp_rows[0].keys())
            path = destination / "treatment_plan_summary.csv"
            _write_csv(path, tp_rows, fieldnames)
            written.append(path.name)

            ca_rows = _derive_case_acceptance(tp_rows)
            if ca_rows:
                ca_path = destination / "case_acceptance.csv"
                _write_csv(ca_path, ca_rows, ["Period", "Presented", "Accepted", "AcceptanceRate"])
                written.append(ca_path.name)

        hr_rows = _aggregate_hygiene_recall(conn, periods)
        if hr_rows:
            hr_path = destination / "hygiene_recall_summary.csv"
            _write_csv(hr_path, hr_rows, ["Period", "HygieneCompleted", "RecallDue"])
            written.append(hr_path.name)

        op_path = destination / "operatory_schedule.json"
        schedule_date = datetime.now().date().isoformat()
        op_chairs = build_operatory_chairs_from_sensei(schedule_date=schedule_date)
        generated_from = "sensei-datasync"
        if not op_chairs:
            op_chairs = _aggregate_operatory_from_db(conn, schedule_date=schedule_date)
            generated_from = "sd_appointments"
        if op_chairs:
            # Prefer Sensei same-day over stale future sd_appointments dumps
            overwrite = not op_path.is_file()
            existing_from = None
            if not overwrite:
                try:
                    existing = json.loads(op_path.read_text(encoding="utf-8-sig"))
                    existing_from = existing.get("generatedFrom")
                    overwrite = existing_from in (
                        None,
                        "sd_appointments",
                        "analytics-db",
                        "sensei-datasync",
                    )
                    # Never keep a future-year polluted file when we have date-filtered chairs
                    if not overwrite and existing_from == "sd_appointments":
                        overwrite = True
                except (json.JSONDecodeError, OSError):
                    overwrite = True
            if overwrite:
                # Stamp scheduleDate from first chair if fallback picked a nearby day
                chair_day = ""
                if isinstance(op_chairs, list) and op_chairs and isinstance(op_chairs[0], dict):
                    chair_day = str(op_chairs[0].get("scheduleDate") or "")
                payload = {
                    "operatoryChairs": op_chairs,
                    "generatedFrom": generated_from,
                    "scheduleDate": chair_day or schedule_date,
                }
                op_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
                _mirror_operatory_export(destination, payload)
                written.append(op_path.name)
    finally:
        conn.close()

    result["ok"] = bool(written)
    result["written"] = written
    if tp_rows:
        result["treatmentPlanSource"] = tp_source
    try:
        from automation_registry import record_job_run

        record_job_run(
            "practice-exports",
            ok=bool(written),
            detail=f"written={len(written)}",
        )
    except Exception:
        pass
    return result


def ingest_csv_reports_to_sqlite(
    csv_dir: Path | None = None,
    db_path: Path | None = None,
) -> dict[str, int]:
    """Load treatment/hygiene/case-acceptance CSVs into analytics SQLite when present.

    Does not invent rows. Uses flexible header aliases. Empty/missing files are skipped.
    """
    db_path = db_path or resolve_analytics_db()
    if not db_path or not Path(db_path).is_file():
        return {}

    roots: list[Path] = []
    if csv_dir is not None:
        roots.append(csv_dir)
    roots.append(softdent_import_dir())
    exports = Path(os.environ.get("SOFTDENT_FINANCIAL_EXPORTS", r"C:\SoftDentFinancialExports"))
    roots.append(exports)

    def _find(name: str) -> Path | None:
        for root in roots:
            candidate = root / name
            if candidate.is_file():
                return candidate
        return None

    mappings: dict[str, tuple[str, dict[str, tuple[str, ...]]]] = {
        "treatment_plan_summary.csv": (
            "sd_treatment_plan_csv",
            {
                "patient_id": ("PatientID", "patient_id", "MRN", "Id"),
                "patient_name": ("PatientName", "patient_name", "Name"),
                "plan_date": ("PlanDate", "plan_date", "Date", "Period"),
                "total_fee": ("TotalFee", "total_fee", "Presented", "PresentedAmount", "Value"),
                "accepted": ("Accepted", "accepted", "AcceptedAmount"),
            },
        ),
        "hygiene_recall_summary.csv": (
            "sd_hygiene_recall_csv",
            {
                "patient_id": ("PatientID", "patient_id", "MRN"),
                "patient_name": ("PatientName", "patient_name", "Name"),
                "due_date": ("DueDate", "due_date", "Due", "Period"),
                "status": ("Status", "status", "Completed", "Overdue"),
            },
        ),
        "case_acceptance.csv": (
            "sd_case_acceptance_csv",
            {
                "patient_id": ("PatientID", "patient_id", "MRN"),
                "patient_name": ("PatientName", "patient_name", "Name"),
                "presented": ("Presented", "PresentedAmount", "presented"),
                "accepted": ("Accepted", "AcceptedAmount", "accepted"),
                "period": ("Period", "period", "Date"),
            },
        ),
    }

    counts: dict[str, int] = {}
    extracted_at = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
    conn = sqlite3.connect(str(db_path), timeout=10.0)
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        for filename, (table, colmap) in mappings.items():
            path = _find(filename)
            if not path:
                continue
            cols_sql = ", ".join([f"{k} TEXT" for k in colmap.keys()] + ["source_file TEXT", "extracted_at TEXT"])
            # table is a fixed mapping key; cols_sql is built from that mapping's keys only.
            conn.execute("CREATE TABLE IF NOT EXISTS " + table + " (" + cols_sql + ")")
            conn.execute("DELETE FROM " + table)
            rows: list[dict[str, Any]] = []
            with path.open("r", encoding="utf-8-sig", newline="") as handle:
                reader = csv.DictReader(handle)
                if not reader.fieldnames:
                    continue
                field_lookup = {str(f).strip().lower(): str(f) for f in reader.fieldnames if f}
                for row in reader:
                    db_row: dict[str, Any] = {}
                    for dest, aliases in colmap.items():
                        value = ""
                        for alias in aliases:
                            src = field_lookup.get(alias.lower())
                            if src and row.get(src) not in (None, ""):
                                value = str(row.get(src) or "").strip()
                                break
                        db_row[dest] = value
                    if not any(db_row.values()):
                        continue
                    db_row["source_file"] = path.name
                    db_row["extracted_at"] = extracted_at
                    rows.append(db_row)
            if not rows:
                continue
            placeholders = ", ".join(f":{k}" for k in rows[0].keys())
            col_names = ", ".join(rows[0].keys())
            conn.executemany(f"INSERT INTO {table} ({col_names}) VALUES ({placeholders})", rows)
            counts[table] = len(rows)

            # Also hydrate summary-shaped treatment_plan_summary when empty and CSV has period totals
            if table == "sd_treatment_plan_csv" and _table_exists(conn, "treatment_plan_summary"):
                cur = conn.execute("SELECT COUNT(*) FROM treatment_plan_summary")
                if int(cur.fetchone()[0] or 0) == 0:
                    presented = 0.0
                    accepted = 0.0
                    for row in rows:
                        try:
                            presented += float(str(row.get("total_fee") or "0").replace("$", "").replace(",", "") or 0)
                        except ValueError:
                            pass
                        try:
                            accepted += float(str(row.get("accepted") or "0").replace("$", "").replace(",", "") or 0)
                        except ValueError:
                            pass
                    conn.execute(
                        """
                        INSERT INTO treatment_plan_summary (
                            row_sha256, business_key, report_date, total_treatment_plans,
                            total_treatment_plan_value, accepted_count, accepted_value,
                            presented_count, presented_value, status, source_file, imported_at_utc
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            f"csv:{path.name}:{extracted_at}",
                            f"csv-ingest:{path.name}",
                            extracted_at[:10],
                            len(rows),
                            presented,
                            sum(1 for r in rows if str(r.get("accepted") or "").strip() not in {"", "0", "0.0", "false", "no"}),
                            accepted,
                            len(rows),
                            presented,
                            "csv-ingest",
                            path.name,
                            extracted_at,
                        ),
                    )
                    counts["treatment_plan_summary"] = 1
        conn.commit()
    finally:
        conn.close()
    return counts


def _parse_money_cell(value: Any) -> float | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text in {"-", "—", "N/A", "n/a"}:
        return None
    neg = text.startswith("(") and text.endswith(")")
    cleaned = text.replace("$", "").replace(",", "").replace("(", "").replace(")", "").strip()
    if not cleaned:
        return None
    try:
        amount = float(cleaned)
    except ValueError:
        return None
    return -amount if neg else amount


def _period_from_soft_date_text(text: str) -> str | None:
    """Best-effort YYYY-MM from SoftDent date lines (no invented periods)."""
    raw = str(text or "")
    month_map = {
        "jan": "01",
        "january": "01",
        "feb": "02",
        "february": "02",
        "mar": "03",
        "march": "03",
        "apr": "04",
        "april": "04",
        "may": "05",
        "jun": "06",
        "june": "06",
        "jul": "07",
        "july": "07",
        "aug": "08",
        "august": "08",
        "sep": "09",
        "sept": "09",
        "september": "09",
        "oct": "10",
        "october": "10",
        "nov": "11",
        "november": "11",
        "dec": "12",
        "december": "12",
    }
    m = re.search(
        r"(?i)\b(Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|"
        r"Jul(?:y)?|Aug(?:ust)?|Sep(?:t(?:ember)?)?|Oct(?:ober)?|Nov(?:ember)?|"
        r"Dec(?:ember)?)\s+(\d{1,2}),?\s+(20\d{2})\b",
        raw,
    )
    if m:
        mon = month_map.get(m.group(1).lower())
        if mon:
            return f"{m.group(3)}-{mon}"
    m = re.search(r"\b(\d{1,2})/(\d{1,2})/(20\d{2})\b", raw)
    if m:
        return f"{m.group(3)}-{int(m.group(1)):02d}"
    m = re.search(r"\b(\d{1,2})/(\d{1,2})/(\d{2})\b", raw)
    if m:
        yy = int(m.group(3))
        year = 2000 + yy if yy < 80 else 1900 + yy
        return f"{year}-{int(m.group(1)):02d}"
    m = re.search(r"\b(20\d{2})[-/](\d{2})(?:[-/]\d{2})?\b", raw)
    if m:
        return f"{m.group(1)}-{m.group(2)}"
    return None


def _period_from_register_filename(name: str) -> str | None:
    """YYYY-MM from SoftDent RegisterForPeriodReportForMMDDYYYY*.xls names."""
    raw = str(name or "")
    m = re.search(r"(?i)(?:for|_)(\d{2})(\d{2})(20\d{2})", raw)
    if m:
        return f"{m.group(3)}-{m.group(1)}"
    m = re.search(r"(20\d{2})[-_/]?(\d{2})", raw)
    if m:
        return f"{m.group(1)}-{m.group(2)}"
    return None


def _load_excel_register_rows_once(path: Path) -> list[list[Any]]:
    """Load SoftDent Register .xls/.xlsx as row lists (single attempt; no PHI logging)."""
    suffix = path.suffix.lower()
    if suffix == ".xls":
        try:
            import xlrd  # type: ignore
        except ImportError as exc:  # pragma: no cover
            raise RuntimeError("xlrd required to parse SoftDent .xls register exports") from exc
        book = xlrd.open_workbook(str(path))
        sheet = book.sheet_by_index(0)
        return [[sheet.cell_value(r, c) for c in range(sheet.ncols)] for r in range(sheet.nrows)]
    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook  # type: ignore
        except ImportError as exc:  # pragma: no cover
            raise RuntimeError("openpyxl required to parse SoftDent .xlsx register exports") from exc
        book = load_workbook(str(path), data_only=True, read_only=True)
        try:
            sheet = book.active
            return [list(row) for row in sheet.iter_rows(values_only=True)]
        finally:
            book.close()
    raise ValueError(f"unsupported excel suffix: {suffix}")


def _load_excel_register_rows(path: Path) -> list[list[Any]]:
    """Load SoftDent Register Excel with Excel-temp lock retry (hal-10576)."""
    from softdent_excel_temp import call_with_excel_temp_retry

    target = Path(path)
    return call_with_excel_temp_retry(lambda: _load_excel_register_rows_once(target))


def _summarize_register_rows(
    rows: list[list[Any]],
    *,
    path: Path,
    schema: dict[str, Any],
    source_kind: str,
    period_hint: str | None = None,
) -> dict[str, Any] | None:
    """Shared Register-for-a-Period label walk (CSV cells or Excel rows)."""
    productions = None
    net_productions = None
    collections = None
    ins_plan = None
    regular = None
    period = period_hint
    for raw_row in rows:
        cells = [("" if c is None else c) for c in (raw_row or [])]
        # Flatten numeric Excel cells + string labels
        text_cells = [str(c).strip() if not isinstance(c, (int, float)) else c for c in cells]
        label = ""
        for c in text_cells:
            if isinstance(c, str) and c.strip():
                label = c.strip()
                break
            if isinstance(c, (int, float)) and not label:
                continue
        if not period:
            joined = " ".join(str(c) for c in text_cells if c not in ("", None))
            period = _period_from_soft_date_text(joined)
        label_l = str(label).lower()
        amounts: list[float] = []
        for c in text_cells:
            if isinstance(c, (int, float)) and not isinstance(c, bool):
                amounts.append(float(c))
            else:
                amt = _parse_money_cell(c)
                if amt is not None:
                    amounts.append(amt)
        if re.match(r"(?i)^productions?$", label_l) and amounts:
            productions = amounts[0]
        elif re.match(r"(?i)^net productions?$", label_l) and amounts:
            net_productions = amounts[0]
        elif re.match(r"(?i)^collections?$", label_l) and amounts:
            collections = amounts[0]
        elif re.search(r"(?i)ins\s*plan\s*collections", label_l) and amounts:
            ins_plan = amounts[0]
        elif re.search(r"(?i)regular\s*collections|patient\s*collections", label_l) and amounts:
            regular = amounts[0]
        elif re.match(r"(?i)^net collections?$", label_l) and amounts and collections is None:
            collections = amounts[0]
    if not period:
        period = _period_from_register_filename(path.name)
    if not period:
        return None
    production = float(productions if productions is not None else (net_productions or 0.0))
    coll = float(collections) if collections is not None else None
    # SoftDent Register explicitly labels both Ins Plan and Regular lines (may be $0).
    labels_present = ins_plan is not None and regular is not None
    insurance = float(ins_plan) if ins_plan is not None else 0.0
    # Patient side: prefer SoftDent "Regular Collections" label (truth even when Ins Plan is $0).
    # Never invent patient=collections when Regular label is absent.
    if regular is not None:
        patient = float(regular)
    elif insurance > 0 and coll is not None and coll > 0:
        patient = max(0.0, coll - insurance)
    else:
        patient = 0.0
    split_ok = insurance > 0 and patient >= 0 and coll is not None and coll > 0
    register_ins_zero = bool(
        ins_plan is not None and float(ins_plan) <= 0 and coll is not None and coll > 0
    )
    return {
        "period": period,
        "production": production,
        "collections": coll,
        "insurance": insurance,
        "patient": patient,
        "insuranceSplitReported": bool(labels_present or split_ok),
        "hasInsurancePatientSplit": bool(labels_present or split_ok),
        "regularCollectionsReported": regular is not None,
        "registerInsPlanZero": register_ins_zero,
        "daysheetWithoutSplit": bool(production > 0 and not labels_present and not split_ok and coll is None),
        "collectionsFormatRequired": bool(
            production > 0 and coll is not None and coll > 0 and not labels_present and not split_ok
        ),
        "sourceKind": source_kind,
        "sourcePath": str(path),
        "schema": schema,
        "insPlanCollections": ins_plan,
        "regularCollections": regular,
    }


def parse_softdent_register_xls(path: Path | str) -> dict[str, Any] | None:
    """Parse SoftDent Register for a Period .xls/.xlsx into a period stub (empty ≠ $0)."""
    target = Path(path)
    if not target.is_file():
        return None
    schema = detect_daysheet_export_schema(target)
    try:
        rows = _load_excel_register_rows(target)
    except Exception as exc:  # noqa: BLE001
        # Never log cell contents (PHI risk on other SoftDent sheets).
        return {
            "period": _period_from_register_filename(target.name) or (schema.get("periodHints") or [None])[0],
            "production": 0.0,
            "collections": None,
            "insurance": 0.0,
            "patient": 0.0,
            "insuranceSplitReported": False,
            "hasInsurancePatientSplit": False,
            "daysheetWithoutSplit": False,
            "collectionsFormatRequired": True,
            "parseError": type(exc).__name__,
            "sourceKind": "register_xls",
            "sourcePath": str(target),
            "schema": schema,
        }
    kind = "register_xls" if target.suffix.lower() == ".xls" else "register_xlsx"
    hint = (schema.get("periodHints") or [None])[0]
    return _summarize_register_rows(
        rows,
        path=target,
        schema=schema,
        source_kind=kind,
        period_hint=hint,
    )


def detect_daysheet_export_schema(path: Path | str) -> dict[str, Any]:
    """Classify SoftDent Daysheet / Register export shape (schema only; no $ invent)."""
    target = Path(path)
    result: dict[str, Any] = {
        "ok": target.is_file(),
        "path": str(target),
        "kind": "unknown",
        "hasProduction": False,
        "hasCollections": False,
        "hasInsurancePatientSplit": False,
        "periodHints": [],
        "notes": [],
    }
    if not target.is_file():
        result["notes"].append("file missing")
        return result
    suffix = target.suffix.lower()
    try:
        if suffix in {".xls", ".xlsx", ".xlsm"}:
            result["kind"] = "register_xls" if suffix == ".xls" else "register_xlsx"
            name_period = _period_from_register_filename(target.name)
            try:
                rows = _load_excel_register_rows(target)
            except Exception as exc:  # noqa: BLE001
                result["notes"].append(f"excel_open_{type(exc).__name__}")
                if name_period:
                    result["periodHints"] = [name_period]
                return result
            content_period = None
            ins_plan = None
            has_prod = False
            has_coll = False
            for raw_row in rows:
                cells = [("" if c is None else c) for c in (raw_row or [])]
                joined = " ".join(str(c) for c in cells if c not in ("", None))
                if not content_period:
                    content_period = _period_from_soft_date_text(joined)
                label = next((str(c).strip() for c in cells if isinstance(c, str) and c.strip()), "")
                label_l = label.lower()
                amounts = [
                    float(c)
                    for c in cells
                    if isinstance(c, (int, float)) and not isinstance(c, bool)
                ]
                if re.match(r"(?i)^productions?$", label_l) and amounts:
                    has_prod = True
                if re.match(r"(?i)^collections?$", label_l) and amounts:
                    has_coll = True
                if re.search(r"(?i)ins\s*plan\s*collections", label_l) and amounts:
                    ins_plan = amounts[0]
            # Content period wins over filename run-date (SoftDent often stamps run day in name).
            period = content_period or name_period
            if period:
                result["periodHints"] = [period]
            if content_period and name_period and content_period != name_period:
                result["notes"].append(
                    f"filename suggests {name_period} but report body period is {content_period}"
                )
            result["hasProduction"] = has_prod
            result["hasCollections"] = has_coll
            result["hasInsurancePatientSplit"] = bool(ins_plan is not None and ins_plan > 0 and has_coll)
            if has_coll and not result["hasInsurancePatientSplit"]:
                result["notes"].append("register excel has collections without positive Ins Plan side")
            return result
        if suffix == ".jsonl":
            first = ""
            with target.open("r", encoding="utf-8-sig", errors="ignore") as handle:
                first = handle.readline()
            payload = json.loads(first) if first.strip() else {}
            dataset = str(payload.get("dataset_name") or "").lower()
            norm = payload.get("normalized") if isinstance(payload.get("normalized"), dict) else {}
            if dataset == "daysheet" or "gross_production" in norm or "report_date" in norm:
                result["kind"] = "daysheet_jsonl"
            elif "register" in dataset:
                result["kind"] = "register_jsonl"
            period = _period_from_soft_date_text(str(norm.get("report_date") or ""))
            if period:
                result["periodHints"] = [period]
            result["hasProduction"] = any(
                norm.get(k) not in (None, "", 0, 0.0) for k in ("gross_production", "net_production")
            )
            result["hasCollections"] = norm.get("collections") not in (None, "", 0, 0.0)
            try:
                ins = float(norm.get("insurance_payment_total") or 0)
            except (TypeError, ValueError):
                ins = 0.0
            result["hasInsurancePatientSplit"] = ins > 0 and result["hasCollections"]
            return result
        text = target.read_text(encoding="utf-8-sig", errors="ignore")[:12000]
    except (OSError, json.JSONDecodeError) as exc:
        result["notes"].append(str(exc))
        return result

    lower = text.lower()
    if "register for a period" in lower:
        result["kind"] = "register_csv"
    elif "daysheet" in lower:
        result["kind"] = "daysheet_csv"
    period = _period_from_soft_date_text(text)
    if period:
        result["periodHints"] = [period]
    result["hasProduction"] = bool(
        re.search(r"(?i)\b(productions?|net productions?|prod)\b", text)
        or ",Prod," in text
        or "\tProd\t" in text
    )
    result["hasCollections"] = bool(re.search(r"(?i)\b(collections?|net collections?)\b", text))
    has_ins = bool(re.search(r"(?i)ins\s*plan\s*collections|insurance\s*payment|posted to insurance", text))
    has_pat = bool(re.search(r"(?i)regular\s*collections|patient", text))
    # Split is only "present" when SoftDent reports a positive insurance side.
    if result["kind"] == "register_csv":
        for line in text.splitlines():
            if re.search(r"(?i)^\s*,?\s*Ins\s*Plan\s*Collections", line):
                cells = [c.strip() for c in line.split(",")]
                for cell in cells:
                    amt = _parse_money_cell(cell)
                    if amt is not None and amt > 0:
                        result["hasInsurancePatientSplit"] = True
                        break
    elif has_ins and result["hasCollections"]:
        # Daysheet footer insurance alone is not a full Ins/Patient Collections export.
        result["hasInsurancePatientSplit"] = False
        result["notes"].append("daysheet may have insurance footer but not Collections Ins/Patient split")
    if has_pat and not result["hasInsurancePatientSplit"]:
        result["notes"].append("patient/regular collections labels seen without positive Ins Plan side")
    return result


def summarize_daysheet_export(path: Path | str) -> dict[str, Any] | None:
    """Parse one SoftDent daysheet/register export into a period stub (honest; empty ≠ $0)."""
    target = Path(path)
    schema = detect_daysheet_export_schema(target)
    if not schema.get("ok"):
        return None
    kind = str(schema.get("kind") or "unknown")

    if kind in {"register_xls", "register_xlsx"} or target.suffix.lower() in {".xls", ".xlsx", ".xlsm"}:
        return parse_softdent_register_xls(target)

    if kind == "daysheet_jsonl":
        by_period: dict[str, dict[str, float]] = {}
        try:
            with target.open("r", encoding="utf-8-sig", errors="ignore") as handle:
                for line in handle:
                    line = line.strip()
                    if not line:
                        continue
                    try:
                        payload = json.loads(line)
                    except json.JSONDecodeError:
                        continue
                    norm = payload.get("normalized") if isinstance(payload.get("normalized"), dict) else {}
                    period = _period_from_soft_date_text(str(norm.get("report_date") or ""))
                    if not period:
                        continue
                    bucket = by_period.setdefault(
                        period,
                        {"production": 0.0, "collections": 0.0, "insurance": 0.0},
                    )
                    try:
                        gross = float(norm.get("gross_production") or 0)
                        net = float(norm.get("net_production") or 0)
                        bucket["production"] += float(gross or net or 0)
                    except (TypeError, ValueError):
                        pass
                    try:
                        bucket["collections"] += float(norm.get("collections") or 0)
                    except (TypeError, ValueError):
                        pass
                    try:
                        bucket["insurance"] += float(norm.get("insurance_payment_total") or 0)
                    except (TypeError, ValueError):
                        pass
        except OSError:
            return None
        if not by_period:
            return None
        # Prefer the newest period key when multiple days are present.
        period = sorted(by_period.keys())[-1]
        totals = by_period[period]
        production = float(totals.get("production") or 0)
        collections = float(totals.get("collections") or 0)
        insurance = float(totals.get("insurance") or 0)
        split_ok = insurance > 0 and collections > 0
        return {
            "period": period,
            "production": production,
            "collections": collections if collections > 0 else None,
            "insurance": insurance if split_ok else 0.0,
            "patient": max(0.0, collections - insurance) if split_ok else 0.0,
            "insuranceSplitReported": split_ok,
            "hasInsurancePatientSplit": split_ok,
            "daysheetWithoutSplit": bool(production > 0 and not split_ok),
            "sourceKind": "daysheet_jsonl",
            "sourcePath": str(target),
            "schema": schema,
        }

    try:
        text = target.read_text(encoding="utf-8-sig", errors="ignore")
    except OSError:
        return None
    lines = text.splitlines()
    period = (schema.get("periodHints") or [None])[0] or _period_from_soft_date_text(text)
    if not period:
        return None

    if kind == "register_csv":
        rows = list(csv.reader(lines))
        return _summarize_register_rows(
            rows,
            path=target,
            schema=schema,
            source_kind="register_csv",
            period_hint=period,
        )

    # daysheet_csv (and unknown CSV with Daysheet header)
    production = 0.0
    collections_footer = None
    insurance_footer = None
    header_idx = next(
        (i for i, line in enumerate(lines) if re.search(r"(?i)\bProd\b", line) and re.search(r"(?i)\bCheck\b", line)),
        None,
    )
    if header_idx is not None:
        reader = csv.reader(lines[header_idx:])
        rows = list(reader)
        if rows:
            header = [h.strip() for h in rows[0]]
            idx = {name: i for i, name in enumerate(header)}
            prod_i = idx.get("Prod")
            for row in rows[1:]:
                if prod_i is not None and len(row) > prod_i:
                    amt = _parse_money_cell(row[prod_i])
                    if amt:
                        production += amt
    for line in lines:
        if re.search(r"(?i)posted to insurance plans", line):
            for cell in line.split(","):
                amt = _parse_money_cell(cell)
                if amt is not None:
                    insurance_footer = amt
                    break
        if re.search(r"(?i)^collections?\b", line.strip().strip(",")):
            for cell in line.split(","):
                amt = _parse_money_cell(cell)
                if amt is not None and amt > 0:
                    collections_footer = amt
                    break
    # SoftDent practice daysheet rarely has a Collections Ins/Patient split row.
    split_ok = False
    return {
        "period": period,
        "production": production,
        "collections": collections_footer,
        "insurance": 0.0,
        "patient": 0.0,
        "insuranceSplitReported": split_ok,
        "hasInsurancePatientSplit": split_ok,
        "daysheetWithoutSplit": bool(production > 0),
        "insuranceFooter": insurance_footer,
        "sourceKind": "daysheet_csv",
        "sourcePath": str(target),
        "schema": schema,
    }


from softdent_odbc_extract import ensure_softdent_odbc_fresh, extract_softdent_odbc, read_extract_status, run_odbc_lane

def stub_era835_ingestion_path() -> dict[str, Any]:
    """ERA-835 insurance detail path (scaffold beyond stub — Moonshot hal-10576).

    Ensures drop-box dirs exist and scans for files. Does not invent dollars or write SoftDent.
    """
    try:
        from apex_era835_pack import scan_era_inbox

        inbox = scan_era_inbox(ensure_dirs=True)
        return {
            "ok": True,
            "mode": "scaffold",
            "readOnly": True,
            "localOnly": True,
            "writeBack": False,
            "honesty": "empty_not_zero",
            "empty": bool(inbox.get("empty")),
            "chipStatus": inbox.get("chipStatus"),
            "chipLabel": inbox.get("chipLabel"),
            "fileCount": inbox.get("fileCount") or 0,
            "files": inbox.get("files") or [],
            "hint": (
                "SoftDent Register Ins Plan Collections $0.00 is SoftDent truth — "
                "proceed with ERA-835 for insurance detail. "
                f"{inbox.get('chipLabel') or 'Awaiting first 835 drop'}. "
                "Apex never invents insurance/patient dollars or posts SoftDent."
            ),
            "candidateRoots": inbox.get("candidateRoots") or [],
            "existingRoots": inbox.get("existingRoots") or [],
            "ingestHooks": [
                "apex_era835_pack.ingest_era835_to_unified",
                "apex_era835_pack.scan_era_inbox",
                "apex_era835_pack.ingest_era_inbox",
                "apex_era835_pack.discover_era_candidates",
                "nr2_contracts.softdent_era.attach_era_to_ingest",
            ],
            "inbox": inbox,
        }
    except Exception as exc:  # noqa: BLE001
        import os
        from pathlib import Path

        candidates = [
            Path(r"C:\SoftDentFinancialExports\era"),
            Path(r"C:\SoftDentReportExports\era"),
        ]
        env_inbox = str(os.environ.get("NR2_ERA835_INBOX") or "").strip()
        if env_inbox:
            candidates.append(Path(env_inbox).expanduser())
        roots = [str(p) for p in candidates]
        existing = [str(p) for p in candidates if p.is_dir()]
        return {
            "ok": True,
            "mode": "stub",
            "readOnly": True,
            "localOnly": True,
            "writeBack": False,
            "honesty": "empty_not_zero",
            "empty": True,
            "chipStatus": "awaiting",
            "chipLabel": "Awaiting first 835 drop",
            "hint": (
                "SoftDent Register Ins Plan Collections $0.00 is SoftDent truth — "
                "proceed with ERA-835 for insurance detail. Drop 835 files into the ERA inbox; "
                "Apex never invents insurance/patient dollars or posts SoftDent."
            ),
            "candidateRoots": roots,
            "existingRoots": existing,
            "error": f"{type(exc).__name__}:{exc}",
            "ingestHooks": [
                "apex_era835_pack.ingest_era835_to_unified",
                "nr2_contracts.softdent_era.attach_era_to_ingest",
            ],
        }


def discover_era_candidates(**kwargs: Any) -> dict[str, Any]:
    """hal-10576 — thin export wrapper for remittance discovery (read-only)."""
    from apex_era835_pack import discover_era_candidates as _discover

    return _discover(**kwargs)


def collections_export_health(**kwargs: Any) -> dict[str, Any]:
    """hal-10576 — Excel-temp / Collections export readability health (read-only)."""
    from softdent_excel_temp import collections_export_health as _health

    return _health(**kwargs)


__all__ = [
    "ensure_softdent_odbc_fresh",
    "extract_softdent_odbc",
    "read_extract_status",
    "run_odbc_lane",
    "sync_practice_exports",
    "read_practice_export_datasets",
    "ingest_csv_reports_to_sqlite",
    "detect_daysheet_export_schema",
    "summarize_daysheet_export",
    "parse_softdent_register_xls",
    "stub_era835_ingestion_path",
    "discover_era_candidates",
    "collections_export_health",
    "atomic_write_excel_export",
    "atomic_copy_export",
    "_load_excel_register_rows_once",
]


if __name__ == "__main__":
    import json as _json

    print(_json.dumps(sync_practice_exports(), indent=2))

